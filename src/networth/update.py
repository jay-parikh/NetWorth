"""Updater — one command replaces the legacy three .bat scripts (SPEC §7).

Round trip: read inputs → fetch AMFI + bhavcopy → compute XIRR → regenerate
the workbook (backup first, atomic replace). Each source failure degrades
gracefully: previous values stay in place and the summary says so.
"""

from __future__ import annotations

import argparse
import glob
import os
import shutil
import sys
from datetime import date
from pathlib import Path

import io

from . import __version__
from . import crypto
from . import model as M
from .compute.cashflows import compute_all_xirr, ppf_ledger_by_account
from .compute.projections import fy_expected_by_person
from .compute.ppf import current_rate, load_ppf_rates, ppf_cashflows, ppf_value
from .compute.snapshot import net_worth_snapshot, upsert_snapshot
from .compute.xirr import xirr
from .fetch import amfi as amfi_mod
from .fetch import bhavcopy as bhav_mod
from .fetch import corporate_actions as ca_mod
from .compute.restructures import (apply_demergers, consumed_label,
                                   integrate_restructures)
from .model import (ASSET_CLASSES, MANUAL_CLASS_LABELS, DividendRow,
                    chained_adjustment_factor, class_has_data,
                    cost_adjustment_factor, effective_enabled, fy_label,
                    load_fmv, load_restructures, off_with_data_classes,
                    resolve_isin)
from .generate import build_workbook
from .reader import read_workbook

KEEP_BACKUPS = 10


# ------------------------------------------------------------- console UI --
# The updater is many users' ONLY window into what happened — it should feel
# alive, not like a log file. Colours degrade gracefully: plain text when the
# terminal can't (old cmd.exe, redirected output), emoji stripped when the
# console encoding can't carry them.

def _supports_color() -> bool:
    if os.environ.get("NO_COLOR"):
        return False
    if not hasattr(sys.stdout, "isatty") or not sys.stdout.isatty():
        return False
    if os.name == "nt":
        os.system("")          # enables ANSI escape processing on Windows 10+
    return True


_COLOR = _supports_color()


def _c(code: str, text: str) -> str:
    return f"\033[{code}m{text}\033[0m" if _COLOR else text


def _say(text: str = "", **kw) -> None:
    """print() that never crashes on a console that can't render emoji."""
    try:
        print(text, **kw)
    except UnicodeEncodeError:
        print(text.encode("ascii", "ignore").decode(), **kw)


def _step(text: str) -> None:
    _say(_c("36", f"  ⏳ {text}"))


def _banner(title: str) -> None:
    line = "─" * 62
    _say(_c("36", line))
    _say(_c("1;36", f"  💰 {title}"))
    _say(_c("36", line))


def _fail(msg: str) -> "SystemExit":
    print(f"ERROR: {msg}", file=sys.stderr)
    return SystemExit(1)


def locate_workbook(arg: str | None) -> Path:
    if arg:
        p = Path(arg)
        if not p.exists():
            raise _fail(f"workbook not found: {p}")
        return p
    default = Path(M.TEMPLATE_FILENAME)
    if default.exists():
        return default
    candidates = [p for p in glob.glob("*.xlsx") if not p.startswith("~$")]
    if len(candidates) == 1:
        return Path(candidates[0])
    raise _fail("no workbook given and no single .xlsx found here — "
                "run next to your tracker file or pass its path")


def ensure_closed(path: Path) -> None:
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise _fail(f"{path.name} is open in Excel — close it and run again")
    lock = path.with_name("~$" + path.name)
    if lock.exists():
        raise _fail(f"{path.name} looks open in Excel (found {lock.name}) — "
                    "close it and run again")


def _refuse_overfull(data) -> None:
    """v1.6.2: regeneration writes only each sheet's row budget, so an
    overfull sheet would silently LOSE rows. Refuse up front — before the
    fetch, before the backup — while everything is still intact (the
    ensure_closed pattern: stop in plain words, change nothing)."""
    def _count(attr) -> int:
        if attr == "tax_rules":
            # the sheet is written MERGED (bundled defaults upserted with
            # the user's rows, §3.22) — measure what will be written, not
            # what was read, or an app upgrade could push it past the cap
            valid, invalid, _w = M.effective_tax_rules(data.tax_rules)
            return len(valid) + len(invalid)
        return len(getattr(data, attr))

    over = [(sheet, _count(attr), last - M.FIRST_DATA_ROW + 1)
            for attr, last, sheet in M.CAPACITIES
            if _count(attr) > last - M.FIRST_DATA_ROW + 1]
    if over:
        raise _fail("; ".join(
            f"{sheet} holds {n} rows but the sheet can only save {cap}"
            for sheet, n, cap in over)
            + " - the update stopped so nothing is lost. Move the extra "
            "rows to another file and run again.")


def make_backup(path: Path, *, unmasked: bool = False) -> Path:
    """Byte-copy the at-rest file. A Locked file yields an encrypted backup
    for free. `unmasked=True` (mask on but the file was left open for
    viewing) names the copy distinctly so the next masked run can purge the
    readable copies (SPEC §3.19)."""
    from datetime import datetime
    bdir = path.parent / "backups"
    bdir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    kind = "unmasked-backup" if unmasked else "backup"
    dest = bdir / f"{path.stem}.{kind}-{stamp}{path.suffix}"
    shutil.copy2(path, dest)
    # each kind rotates on its own: a view-preferring user who never
    # re-masks would otherwise accumulate readable copies without bound
    # (v1.6.2; the masked-run purge still removes unmasked ones earlier)
    for kind_glob in (f"{path.stem}.backup-*{path.suffix}",
                      f"{path.stem}.unmasked-backup-*{path.suffix}"):
        backups = sorted(bdir.glob(kind_glob))
        for old in backups[:-KEEP_BACKUPS]:
            old.unlink()
    return dest


def _purge_unmasked_backups(path: Path, keep: Path | None = None) -> int:
    """Remove the readable view-run backups once the file is masked/locked
    again — they were the one place hidden numbers could linger. `keep`
    (v1.6.2) spares THIS run's pre-run copy so one rollback backup always
    survives one cycle; the next successful masked run removes it."""
    gone = 0
    for p in path.parent.glob(f"backups/{path.stem}.unmasked-backup-*"):
        if keep is not None and p == keep:
            continue
        p.unlink()
        gone += 1
    return gone


def _dividend_qty(owner: str, isin: str, ex: date, equity_rows,
                  isin_by_name: dict[str, str], actions) -> float:
    """Estimated shares held at the ex-date (SPEC §6.12): lots bought before
    the ex-date, at the CA-adjusted count as of the day before. A lot still
    keyed to a merged-away ISIN earns the SUCCESSOR's dividends — the chain
    resolves as of the ex-date and the merger ratio folds into the count.
    There is no sell ledger — the estimate projects the CURRENT rows
    backwards, which the sheet hint and Guide state plainly (hence the amber
    '(est.)' columns)."""
    from datetime import timedelta
    as_of = ex - timedelta(days=1)
    total = 0.0
    for row in equity_rows:
        row_isin = row.isin_override or isin_by_name.get(row.scrip, "")
        if not row_isin or row.owner != owner or not row.qty:
            continue
        if row_isin != isin and resolve_isin(row_isin, actions, as_of) != isin:
            continue
        if row.cost_date and row.cost_date >= ex:
            continue
        total += row.qty * chained_adjustment_factor(row_isin, row.cost_date,
                                                     as_of, actions)
    return total


def _merge_stock_master(existing: list[tuple[str, str, str]],
                        fetched: list[tuple[str, str, str]]) -> tuple[list, int]:
    """Add-only (SPEC §6.4): known ISINs keep their names so user rows survive."""
    known = {isin for _s, _n, isin in existing}
    merged = list(existing)
    added = 0
    for sym, name, isin in fetched:
        if isin not in known:
            merged.append((sym, name, isin))
            known.add(isin)
            added += 1
    merged.sort(key=lambda r: r[1].casefold())
    return merged, added


def _replace_mf_master(existing: list[tuple[str, str, str]],
                       fetched: list[tuple[str, str, str]],
                       referenced: set[str]) -> list[tuple[str, str, str]]:
    """Wholesale AMFI refresh, preserving ISINs a user row still references."""
    new_isins = {isin for _f, _s, isin in fetched}
    merged = list(fetched)
    merged.extend(row for row in existing
                  if row[2] in referenced and row[2] not in new_isins)
    merged.sort(key=lambda r: r[1].casefold())
    return merged


def run(path: Path, *, price_data=None, amfi_data=None, ca_data=None,
        div_data=None, bullion_rates=None, nps_data=None, restructures=None,
        add_persons: list[str] | None = None,
        toggle_classes: list[str] | None = None,
        password: str | None = None, reveal: bool = False,
        reset_privacy: bool = False,
        today: date | None = None, do_backup: bool = True) -> dict:
    today = today or date.today()
    summary: dict = {"warnings": []}

    ensure_closed(path)
    # ---- privacy read path (SPEC §3.19): a Locked file is ciphertext and
    # cannot even be read without the password — which is the point
    locked_at_rest = crypto.is_encrypted(path)
    if locked_at_rest:
        if not password:
            raise _fail(f"{path.name} is locked — run by double-click and "
                        "type your password (a scheduled run can't)")
        try:
            data = read_workbook(crypto.decrypt_workbook(path, password))
        except crypto.WrongPassword:
            raise _fail("wrong password — the file was not touched")
    else:
        data = read_workbook(str(path))
    summary["warnings"].extend(data.warnings)
    _refuse_overfull(data)

    # add new people (declaratively: they land in data.persons, so regeneration
    # creates their sheet, Dashboard row, By-Scrip column — everything)
    if add_persons:
        have = {p.casefold() for p in data.persons}
        added = []
        for name in add_persons:
            name = name.strip()
            if name and name.casefold() not in have and len(data.persons) < 10:
                data.persons.append(name)
                have.add(name.casefold())
                added.append(name)
        summary["persons_added"] = added

    # show/hide asset classes chosen at the console (v1.4) — writes the same
    # Settings the user could edit by hand, then regeneration applies it
    if toggle_classes:
        from .model import ClassSetting
        key_by_label = {c.label.casefold(): c.key for c in ASSET_CLASSES}
        label_by_key = {c.key: c.label for c in ASSET_CLASSES}
        toggled = []
        for name in toggle_classes:
            key = key_by_label.get(name.strip().casefold())
            if not key:
                continue
            s = data.class_settings.setdefault(key, ClassSetting())
            s.enabled = not s.enabled
            toggled.append(f"{label_by_key[key]} → "
                           f"{'shown' if s.enabled else 'hidden'}")
            # v1.4.3 (SPEC §3.14): the choice wins — switched off means
            # hidden and not counted, rows kept. Say so at toggle time; the
            # steady-state reminder is the single not-counted line below.
            if not s.enabled and class_has_data(data, key):
                summary["warnings"].append(
                    f"{label_by_key[key]} is now hidden — its rows are "
                    f"saved but not counted; switch it back on in Settings "
                    f"to include them")
        summary["classes_toggled"] = toggled

    # ---- privacy state resolution (SPEC §3.19) — four legal Mask×Lock
    # combinations, transitions included; the user's Settings Yes/No always
    # round-trips untouched (only RESET, an explicit user action, edits it)
    if reset_privacy and not data.lock_enabled:
        data.privacy_enabled = False
        data.privacy_hash = ""
        summary["warnings"].append(
            "privacy mask RESET — the numbers are visible again; switch "
            "Privacy mask to Yes in Settings to start over")
    wants_privacy = data.privacy_enabled or data.lock_enabled
    if wants_privacy and not data.privacy_hash and password:
        data.privacy_hash = crypto.hash_password(password)   # first enable
    pw_ok = bool(password) and (
        locked_at_rest or crypto.verify_password(password, data.privacy_hash))
    if wants_privacy and not data.privacy_hash:
        summary["warnings"].append(
            "privacy is switched on in Settings but no password is set — "
            "run by double-click once to choose one; nothing was masked or "
            "locked this run")
    mask_active = data.privacy_enabled and bool(data.privacy_hash)
    lock_active = data.lock_enabled and bool(data.privacy_hash)
    if lock_active and not locked_at_rest and not pw_ok:
        # never encrypt on an unconfirmed password — that could lock the
        # user out with a password they mistyped or misremember
        lock_active = False
        summary["warnings"].append(
            "Lock file is switched on but the password wasn't confirmed "
            "this run — run by double-click and type it to lock the file")
    if lock_active and not locked_at_rest:
        summary["warnings"].append(
            "the file is now locked; backups made BEFORE locking are still "
            "readable — delete the backups folder if that matters")
    masked_build = mask_active and not (pw_ok and reveal)
    if password and mask_active and not pw_ok:
        # a wrong password is never silent — say why the numbers stayed hidden
        summary["warnings"].append(
            "the password didn't match — the numbers stay masked (•••)")

    # a Manual_Assets row with a Class the dropdown doesn't know lands in NO
    # class column: it shows in the sheet TOTAL but in no person/family total
    # (the dropdown is non-blocking by design — so say it, never guess)
    for r in data.manual_assets:
        if r.asset_class and r.asset_class not in MANUAL_CLASS_LABELS:
            summary["warnings"].append(
                f"Manual_Assets row '{r.description or r.owner}' has "
                f"unrecognised Class '{r.asset_class}' — it is counted in no "
                f"class total; pick Property / Cash / Insurance / Other")

    backup_path: Path | None = None
    if do_backup:
        backup_path = make_backup(
            path, unmasked=(mask_active and not data.masked_at_rest
                            and not locked_at_rest))
        summary["backup"] = str(backup_path)

    # ---- restructures (SPEC §6.15): curated file + Manual rows; the one
    # corporate-action category with no free feed. Runs BEFORE pricing so
    # consumed ISINs route to their successor and demerger children get
    # priced in the same run. A malformed curated file fails loudly.
    if restructures is None:
        try:
            restructures = load_restructures()
        except OSError:
            restructures = []
    else:
        # never mutate the caller's event objects (the Applied stamping
        # below writes into them) — the loaded-from-CSV path is fresh anyway
        from dataclasses import replace as _replace
        restructures = [_replace(c) for c in restructures]
    restructure_events = integrate_restructures(data, restructures, today)

    def _consumed_label(isin: str) -> str | None:
        return consumed_label(isin, restructure_events)
    # demerger child rows are appended AFTER the corporate-actions refresh
    # below: child quantities need the FULL split/bonus history, and a fresh
    # workbook's sheet has none of it yet (§6.15)

    # ---- fetch (graceful per source; the _step lines keep the console alive
    # during the slow network minute) ----
    if price_data is None:
        _step("Fetching share & bond prices (BSE + NSE)…")
        try:
            price_data = bhav_mod.fetch(today=today)
        except Exception as e:  # noqa: BLE001 — any fetch failure degrades
            summary["warnings"].append(f"price fetch failed, keeping old prices: {e}")
    if amfi_data is None:
        _step("Fetching mutual-fund NAVs (AMFI)…")
        try:
            # fetch() itself distrusts a suspiciously-empty feed (v1.6.2,
            # AMFI_MIN_SCHEMES) and raises — landing in this same warning
            amfi_data = amfi_mod.fetch()
        except Exception as e:  # noqa: BLE001
            summary["warnings"].append(f"AMFI fetch failed, keeping old NAVs: {e}")

    stamp = today.strftime("%d-%m-%Y")

    # ---- equity prices + stock master + trading status (SPEC §6.5) ----
    if price_data:
        isin_by_name = {name: isin for _s, name, isin in data.masters.stock_rows}
        trade_date = price_data.trade_date or today
        # absence only means anything when BOTH exchanges answered (SPEC §6.5);
        # on a single-source day, unquoted rows carry their status forward
        dual_source = len(getattr(price_data, "sources", ())) >= 2
        nse_only = getattr(price_data, "nse_only", set())
        matched = 0
        nse_only_hits = 0
        suspended = 0
        name_by_isin_px = {isin: name for _s, name, isin in data.masters.stock_rows}
        for row in data.equity:
            isin = row.isin_override or isin_by_name.get(row.scrip, "")
            if not isin:
                continue
            # a consumed ISIN routes to its survivor (SPEC §6.15): the row
            # prices as the new security while the user's cells stay put
            routed = (resolve_isin(isin, restructure_events, today)
                      if restructure_events else isin)
            quote = price_data.prices.get(routed)
            if routed != isin:
                label = _consumed_label(isin) or "Merged"
                _st, last = data.masters.stock_status.get(isin, ("", None))
                data.masters.stock_status[isin] = (label, last or row.close_date)
                if not row.flag:
                    row.flag = (f"MERGED→{name_by_isin_px.get(routed, routed)}"
                                if label == "Merged" else
                                f"ISIN→{routed}")
            if quote:
                row.close = quote["close"]
                if quote["prev"]:
                    row.prev_close = quote["prev"]
                row.close_date = trade_date
                if routed == isin:
                    data.masters.stock_status[isin] = ("Active", trade_date)
                matched += 1
                nse_only_hits += routed in nse_only
            elif dual_source and routed == isin:
                # absent from both exchanges: escalate by how long it has been
                # (consumed ISINs are exempt — their absence is expected)
                prev_status, last = data.masters.stock_status.get(isin, ("", None))
                last = last or row.close_date
                if prev_status == "Delisted":
                    continue
                if last and (today - last).days > 180:
                    data.masters.stock_status[isin] = ("Delisted", last)
                elif last and (today - last).days > 21:
                    data.masters.stock_status[isin] = ("Suspended", last)
                    suspended += 1
                elif last:
                    data.masters.stock_status[isin] = (prev_status or "Active", last)
        summary["suspended"] = suspended
        summary["nse_only_matched"] = nse_only_hits
        data.masters.stock_rows, added = _merge_stock_master(
            data.masters.stock_rows, price_data.master_rows)
        data.masters.stock_refreshed = stamp
        summary["equity_matched"] = matched
        summary["equity_total"] = sum(1 for r in data.equity if r.scrip or r.isin_override)
        summary["stocks_added"] = added
        summary["price_source"] = f"{price_data.source} {trade_date.strftime('%d-%m-%Y')}"

        # bonds trade on the exchanges too — refresh when the ISIN is quoted
        bonds_matched = 0
        for b in data.bonds:
            quote = price_data.prices.get(b.isin)
            if quote:
                b.cur_price = quote["close"]
                bonds_matched += 1
        summary["bonds_matched"] = bonds_matched

    # ---- fund NAVs + MF master ----
    if amfi_data:
        isin_by_scheme = {scheme: isin for _f, scheme, isin in data.masters.mf_rows}
        matched = 0
        referenced: set[str] = set()
        for rows in (data.mutual_funds, data.sip):
            for row in rows:
                isin = row.isin_override or isin_by_scheme.get(row.scheme, "")
                if isin:
                    referenced.add(isin)
        for row in data.mutual_funds:
            isin = row.isin_override or isin_by_scheme.get(row.scheme, "")
            nav = amfi_data.nav_by_isin.get(isin)
            if nav:
                row.current_nav = nav
                matched += 1
        data.masters.mf_rows = _replace_mf_master(
            data.masters.mf_rows, amfi_data.master_rows, referenced)
        data.masters.mf_refreshed = stamp
        summary["mf_matched"] = matched
        summary["mf_total"] = len(data.mutual_funds)

    # ---- corporate actions (SPEC §6.7): fetch NSE+BSE, keep manual, factor rows,
    # and warn about any holding NEITHER exchange could verify — never skip silently
    isin_by_name = {name: isin for _s, name, isin in data.masters.stock_rows}
    symbol_by_isin = {isin: sym for sym, _n, isin in data.masters.stock_rows}
    name_by_isin = {isin: name for _s, name, isin in data.masters.stock_rows}
    held_isins = {row.isin_override or isin_by_name.get(row.scrip, "")
                  for row in data.equity}
    held_isins.discard("")
    # merged holdings live on as their successor (SPEC §6.15): query the
    # successor's feeds too, so ITS dividends and later splits arrive
    routed_isins = ({resolve_isin(i, data.corporate_actions, today)
                     for i in held_isins} if restructure_events else set())
    query_isins = held_isins | routed_isins
    div_skipped = 0
    ca_injected = ca_data is not None      # injected data is trusted as-is
    ca_checked: set[str] = set()           # ISINs verified by the live fetch
    preserve_isins: set[str] = set()       # feed failed → keep their old rows
    if ca_data is None:
        symbols = {symbol_by_isin[i]: i for i in query_isins if symbol_by_isin.get(i)}
        bse_codes = {code: isin
                     for isin, code in getattr(price_data, "codes_by_isin", {}).items()
                     if isin in query_isins} if price_data else {}
        try:
            checked: set[str] = set()
            if symbols or bse_codes:
                _step("Checking corporate actions & dividends (NSE + BSE)…")
                fy_start = date(today.year if today.month >= 4 else today.year - 1,
                                4, 1)
                ca_data, checked, fetched_divs, div_skipped = ca_mod.fetch(
                    symbols, bse_codes, div_skip_since=fy_start)
                if div_data is None:
                    div_data = fetched_divs
            else:
                ca_data = []
                if div_data is None:
                    div_data = []
            ca_checked = checked
            # a restructured (merged/renamed) ISIN is expected to be absent
            # from the CA feeds — the curated file IS its verification
            consumed = {i for i in held_isins if _consumed_label(i)}
            unchecked = query_isins - checked - consumed
            preserve_isins = unchecked
            summary["ca_unverified"] = sorted(
                name_by_isin.get(i, i) for i in unchecked)
            if unchecked:
                summary["warnings"].append(
                    "corporate actions could NOT be verified for: "
                    + ", ".join(summary["ca_unverified"])
                    + " — their existing rows are kept, but NEW splits/bonuses "
                      "may be missing; add any you know of as Manual rows on "
                      "the Corporate_Actions sheet")
        except Exception as e:  # noqa: BLE001
            summary["warnings"].append(
                f"corporate-actions fetch failed, keeping existing rows: {e}")
    else:
        summary["ca_unverified"] = []
    if ca_data is not None:
        manual = [a for a in data.corporate_actions if a.source != "Auto"]
        manual_keys = {(a.isin, a.type, a.ex_date) for a in manual}
        # a one-symbol feed failure must not revert that stock's already-
        # applied rows: Auto rows of unverified ISINs survive the rebuild
        kept_auto = [a for a in data.corporate_actions
                     if a.source == "Auto" and a.isin in preserve_isins]
        auto = kept_auto + [a for a in ca_data
                            if (a.isin, a.type, a.ex_date) not in manual_keys]
        auto.sort(key=lambda a: (a.symbol, a.ex_date or today))
        # Manual/Curated rows lead: they carry user data and the demerger
        # Applied stamps, so they must never fall past the sheet's last row
        ca_cap = M.CA_LAST_ROW - M.FIRST_DATA_ROW + 1
        overflow = len(manual) + len(auto) - ca_cap
        if overflow > 0:
            if len(manual) > ca_cap:
                # Manual/Curated rows can never be sacrificed: they carry
                # user data and the demerger Applied stamps (a truncated
                # stamp would re-apply the event next run and duplicate its
                # child Equity rows) — refuse, like _refuse_overfull
                raise _fail(
                    f"Corporate_Actions holds {len(manual)} Manual rows "
                    f"but the sheet can only save {ca_cap} - the update "
                    "stopped so nothing is lost. Move old Manual rows to "
                    "another file and run again.")
            oldest = sorted(auto, key=lambda a: a.ex_date or today)[:overflow]
            drop_ids = {id(a) for a in oldest}
            auto = [a for a in auto if id(a) not in drop_ids]
            summary["warnings"].append(
                f"Corporate_Actions sheet is full — dropped the {overflow} "
                f"oldest Auto row(s); adjusted quantities of very old lots "
                f"may be affected")
        data.corporate_actions = manual + auto
        summary["ca_rows"] = len(data.corporate_actions)

    # ---- demerger children (SPEC §6.15) — see compute/restructures.py for
    # the mechanism (append-once, atomic per event, verification/capacity
    # gates that defer WITHOUT stamping so the event retries next run)
    children_added, rs_warnings = apply_demergers(
        data, restructure_events, ca_checked=ca_checked,
        ca_trusted=ca_injected, price_data=price_data, today=today)
    summary["warnings"].extend(rs_warnings)
    summary["restructure_children"] = children_added

    adjusted = 0
    for row in data.equity:
        isin = row.isin_override or isin_by_name.get(row.scrip, "")
        # chained: merger ratios fold in, and later splits on the successor
        # keep applying to the old-ISIN row (SPEC §6.15)
        f = chained_adjustment_factor(isin, row.cost_date, today,
                                      data.corporate_actions)
        row.ca_factor = f if abs(f - 1.0) > 1e-9 else None
        adjusted += row.ca_factor is not None
        cf = cost_adjustment_factor(isin, row.cost_date, today,
                                    data.corporate_actions)
        row.cost_factor = cf if abs(cf - 1.0) > 1e-9 else None
    summary["ca_adjusted_rows"] = adjusted

    # ---- dividends (SPEC §6.12): rebuild current-FY Auto rows from the feed,
    # freeze prior FYs, let Manual rows override the same (isin, type, ex-date)
    fy_now = fy_label(today)
    for d in data.dividends:                     # backfill FY typed-less rows
        if not d.fy and d.ex_date:
            d.fy = fy_label(d.ex_date)
    if div_data is not None:
        manual = [d for d in data.dividends if d.source != "Auto"]
        # keyed WITHOUT div_type (matching dedupe_dividends): the exchanges
        # word the same event differently, and a Manual correction must
        # override the Auto row however the feed typed it
        manual_keys = {(d.isin, d.ex_date) for d in manual}
        frozen_auto = [d for d in data.dividends
                       if d.source == "Auto" and d.fy != fy_now]
        # a feed failure for one symbol must not delete its current-FY rows
        kept = [d for d in data.dividends
                if d.source == "Auto" and d.fy == fy_now
                and d.isin in preserve_isins]
        kept_keys = {(d.isin, d.ex_date) for d in kept}
        fresh: list[DividendRow] = []
        for ev in div_data:
            if (not ev.ex_date or fy_label(ev.ex_date) != fy_now
                    or (ev.isin, ev.ex_date) in manual_keys
                    or (ev.isin, ev.ex_date) in kept_keys):
                continue
            for owner in sorted({r.owner for r in data.equity if r.owner}):
                qty = _dividend_qty(owner, ev.isin, ev.ex_date,
                                    data.equity, isin_by_name,
                                    data.corporate_actions)
                if qty <= 0:
                    continue
                fresh.append(DividendRow(
                    fy=fy_now, owner=owner,
                    scrip=name_by_isin.get(ev.isin, ev.scrip or ev.isin),
                    isin=ev.isin, div_type=ev.div_type, ex_date=ev.ex_date,
                    rate=ev.rate, qty=round(qty, 3), source="Auto",
                    details=ev.details))
        fresh.sort(key=lambda d: (d.ex_date or today, d.scrip, d.owner))
        rows = frozen_auto + kept + fresh + manual
        # never let Manual/current-FY rows fall past the sheet's last row:
        # on overflow the OLDEST frozen prior-FY Auto rows give way, loudly
        div_cap = M.DIV_LAST_ROW - M.FIRST_DATA_ROW + 1
        if len(rows) > div_cap:
            overflow = len(rows) - div_cap
            if len(manual) > div_cap:
                # Manual rows are user data — refuse rather than truncate
                raise _fail(
                    f"Dividends holds {len(manual)} Manual rows but the "
                    f"sheet can only save {div_cap} - the update stopped "
                    "so nothing is lost. Move old Manual rows to another "
                    "file and run again.")
            # oldest Auto rows give way — prior-FY first (they sort oldest),
            # current-FY ones only when even that isn't enough
            oldest = sorted(frozen_auto + kept + fresh,
                            key=lambda d: d.ex_date or today)[:overflow]
            drop_ids = {id(d) for d in oldest}
            rows = [d for d in rows if id(d) not in drop_ids]
            summary["warnings"].append(
                f"Dividends sheet is full — dropped the {overflow} oldest "
                f"Auto row(s); copy old years elsewhere if you want to "
                f"keep the full record")
        data.dividends = rows
        summary["dividend_rows"] = len(fresh)
    if div_skipped:
        summary["warnings"].append(
            f"{div_skipped} dividend announcement(s) could not be parsed "
            "(e.g. percent-of-face-value wording) — add them as Manual rows "
            "on the Dividends sheet if they are yours")
    summary["dividends_fy_total"] = round(sum(
        (d.rate or 0) * (d.qty or 0)
        for d in data.dividends if d.fy == fy_now), 2)
    orphan_div = sum(1 for d in data.dividends
                     if d.fy == fy_now and not d.owner
                     and (d.rate or 0) and (d.qty or 0))
    if orphan_div:
        summary["warnings"].append(
            f"{orphan_div} dividend row(s) this year have no Owner - they "
            "count in the family total but in no one's per-person figure "
            "(add the Owner on the Dividends sheet)")

    # ---- FMV 31-01-2018 fallback for unknown old costs (SPEC §6.6) ----
    fmv_filled = 0
    try:
        fmv_by_isin, fmv_by_symbol = load_fmv()
    except OSError:
        fmv_by_isin, fmv_by_symbol = {}, {}
    if fmv_by_isin:
        isin_by_name = {name: isin for _s, name, isin in data.masters.stock_rows}
        symbol_by_isin = {isin: sym for sym, _n, isin in data.masters.stock_rows}
        cutoff = date(2018, 2, 1)
        for row in data.equity:
            if row.avg_cost is not None or not row.qty:
                continue
            if not row.cost_date or row.cost_date >= cutoff:
                continue
            isin = row.isin_override or isin_by_name.get(row.scrip, "")
            # the 2018 ISIN may differ from today's (post-split reissues) —
            # fall back to the exchange symbol
            fmv = fmv_by_isin.get(isin) or fmv_by_symbol.get(symbol_by_isin.get(isin, ""))
            if fmv:
                row.avg_cost = fmv
                row.fmv_used = True
                fmv_filled += 1
    summary["fmv_filled"] = fmv_filled

    # ---- capital gains (v1.6, SPEC §6.16): computed ONCE here — the console
    # line reads it and the build below reuses it (capgains=), so the two can
    # never differ — and only when the user is actually using the feature ----
    capgains_rep = None
    if data.equity_sells or data.show_capital_gains:
        from .compute.capital_gains import capital_gains_report
        capgains_rep = rep = capital_gains_report(data, today)
        summary["warnings"].extend(rep.warnings)
        now = next((s for s in rep.summaries if s.fy == rep.fy_now), None)
        bits = []
        if now:
            bits.append(f"FY {now.fy}: STCG ₹{now.stcg:,.0f} · "
                        f"LTCG ₹{now.ltcg:,.0f}")
        if now and now.spec_gain:
            bits.append(f"intraday ₹{now.spec_gain:,.0f} (speculative, "
                        f"at your slab)")
        if now and now.st_sheltered:
            # without these two bits the raw figures and the tax numbers
            # printed beside them look mutually impossible in a harvest year
            bits.append(f"debt losses ₹{now.st_sheltered:,.0f} set off vs "
                        f"STCG (Sec 70)")
        if now and now.st_setoff:
            bits.append(f"losses ₹{now.st_setoff:,.0f} set off vs LTCG "
                        f"(Sec 70)")
        if rep.headroom_now is not None:
            bits.append(f"₹{rep.headroom_now:,.0f} LTCG still tax-free "
                        f"this year")
        if bits:
            summary["capgains"] = " — ".join(bits)

    # v1.6.2: a row whose units can't be known is left out of the return
    # figure entirely (cashflows) — that must never be silent. The capital-
    # gains engine warns about the same rows; this covers the default case
    # where that engine is off.
    if capgains_rep is None:
        from .compute.cashflows import _sip_units
        for r in data.sip:
            if (r.owner and r.scheme and r.txn_date and r.amount
                    and r.txn_date <= today and not _sip_units(r)):
                summary["warnings"].append(
                    f"MF_SIP: a {r.scheme} row on {r.txn_date:%d-%m-%Y} "
                    "has no NAV or units - left out of the return figure")

    # ---- gold & silver (SPEC §5.7): SGBs from the bhavcopy; physical metal
    # from IBJA, falling back to the bhavcopy-implied median, else kept as-is
    if data.bullion:
        from .fetch import bullion as bullion_mod
        rates = bullion_rates
        rate_src = "injected"
        if rates is None:
            _step("Fetching today's gold & silver rate (IBJA)…")
            rates = bullion_mod.fetch_ibja()
            rate_src = "IBJA"
            if not rates:
                try:
                    rates = bullion_mod.derive_from_bhavcopy(price_data)
                    rate_src = "market-implied (SGB/ETF closes)"
                except OSError:
                    rates = {}
        sgb_priced = metal_rated = metal_rows = 0
        for r in data.bullion:
            if r.metal_type == "SGB":
                if r.isin and price_data:
                    quote = price_data.prices.get(r.isin)
                    if quote:
                        r.rate_auto = quote["close"]
                        sgb_priced += 1
            elif r.metal_type in ("Gold", "Silver"):
                metal_rows += 1
                rate = rates.get(r.metal_type.lower())
                if rate:
                    r.rate_auto = rate
                    metal_rated += 1
        # the as-of date belongs to the METAL benchmark rate only — SGB
        # closes carry their own dates, and stamping on an SGB-only day
        # would hide how stale the gold/silver rate really is
        if metal_rated:
            data.bullion_rate_asof = today
        if metal_rows and not metal_rated:
            summary["warnings"].append(
                "gold/silver rate unavailable (IBJA and market-implied both "
                "failed) — kept the previous rates; the Rate override column "
                "always wins")
        summary["bullion"] = (f"{sgb_priced} SGB(s) priced, "
                              f"{metal_rated}/{metal_rows} metal row(s) rated"
                              + (f" ({rate_src})" if metal_rated else ""))

    # ---- NPS (SPEC §5.6): daily NAVs + scheme master, keyed by scheme code
    if data.nps and nps_data is None:
        from .fetch import nps as nps_mod
        _step("Fetching NPS NAVs (NPS Trust)…")
        try:
            nps_data = nps_mod.fetch()
        except Exception as e:  # noqa: BLE001
            summary["warnings"].append(
                f"NPS NAV fetch failed, keeping old NAVs: {e}")
    if nps_data:
        known = {code for code, _n, _p in data.masters.nps_rows}
        merged = list(data.masters.nps_rows)
        merged.extend(row for row in nps_data.master_rows
                      if row[0] not in known)
        merged.sort(key=lambda r: r[1].casefold())      # dropdown sort rule
        data.masters.nps_rows = merged
        data.masters.nps_refreshed = stamp
        code_by_scheme = {name: code
                          for code, name, _p in data.masters.nps_rows}
        matched = 0
        for r in data.nps:
            code = r.scheme_code_override or code_by_scheme.get(r.scheme, "")
            nav = nps_data.nav_by_code.get(code)
            if nav:
                r.current_nav = nav
                matched += 1
        summary["nps_matched"] = matched
        summary["nps_total"] = len(data.nps)

    # ---- EPF: auto-fill blank rates from the bundled EPFO table (SPEC §3.17)
    if data.epf:
        try:
            from .model import current_epf_rate
            epf_rate = current_epf_rate()
            for r in data.epf:
                if r.rate is None:
                    r.rate = epf_rate
        except OSError:
            pass

    # ---- PPF: exact balance/interest/XIRR for ledgered accounts (SPEC §6.10) ----
    rates = load_ppf_rates()
    led = ppf_ledger_by_account(data, today)
    ledgered = 0
    for row in data.ppf:
        if row.rate is None:
            row.rate = current_rate(rates)          # auto-fill blank rate
        deposits = led.get((row.owner, row.account_no))
        if deposits:
            bal, interest = ppf_value(deposits, rates, today)
            row.balance_today = round(bal, 2)
            row.interest_earned = round(interest, 2)
            row.xirr = xirr(ppf_cashflows(deposits, bal, today))
            ledgered += 1
        else:
            row.balance_today = None                 # generate → Balance today = D
            row.interest_earned = None
            row.xirr = None
    summary["ppf_ledgered"] = ledgered

    # ---- XIRR + FY-end estimate (always recomputed from current values) ----
    data.xirr = compute_all_xirr(data, today)
    summary["portfolio_xirr"] = data.xirr.portfolio
    data.fy_expected = fy_expected_by_person(data, today)
    summary["fy_expected_total"] = round(sum(data.fy_expected.values()), 2)

    # ---- net-worth history: one snapshot per day (SPEC §6.11) ----
    snap = net_worth_snapshot(data, today)
    # v1.4.3 (SPEC §3.14): a switched-off class is not counted anywhere.
    # One line (mirrored by the Dashboard notice) keeps hidden money on the
    # user's radar — with its value, measured before zeroing the snapshot.
    off_data = off_with_data_classes(data)
    if off_data:
        parts = [f"{c.label} ₹{v:,.0f}" if (v := getattr(snap, c.key, 0.0))
                 else c.label for c in off_data]
        summary["warnings"].append(
            "hidden and not counted (your Settings choice): "
            + " · ".join(parts) + " — switch on in Settings to include")
    for cls in ASSET_CLASSES:
        if not effective_enabled(data, cls):
            setattr(snap, cls.key, 0.0)
    data.history = upsert_snapshot(data.history, snap,
                                   keep=M.HISTORY_LAST_ROW - M.FIRST_DATA_ROW + 1)
    summary["net_worth"] = round(snap.total, 2)
    summary["history_points"] = len(data.history)

    # ---- regenerate, atomically ----
    _regenerate_atomic(path, data, masked_build=masked_build,
                       lock_active=lock_active, password=password,
                       today=today, capgains=capgains_rep)
    if masked_build or lock_active:
        # v1.6.2: this run's own pre-run copy survives ONE cycle as the
        # rollback of last resort; the next masked run removes it
        gone = _purge_unmasked_backups(path, keep=backup_path)
        if gone:
            summary["warnings"].append(
                f"removed {gone} older readable backup(s) now that the file "
                f"is masked again (this run's own backup is kept until the "
                f"next update)")
    summary["privacy"] = ("locked + masked" if lock_active and masked_build
                          else "locked" if lock_active
                          else "masked" if masked_build
                          else "open (viewing)" if mask_active else "")
    summary["workbook"] = str(path)
    return summary


def _regenerate_atomic(path: Path, data, *, masked_build: bool,
                       lock_active: bool, password: str | None,
                       today: date | None = None, capgains=None) -> None:
    """The one regenerate-and-replace implementation, shared by run() and
    relock() so the two can never drift (SPEC §3.19: on the Lock path the
    plain workbook exists only in memory — the disk only ever sees
    ciphertext, and encrypt_workbook self-verifies before anything is
    replaced)."""
    tmp = path.with_name(path.stem + ".new" + path.suffix)
    try:
        if lock_active:
            buf = io.BytesIO()
            build_workbook(data, buf, masked=masked_build, today=today,
                           capgains=capgains)
            tmp.write_bytes(crypto.encrypt_workbook(buf.getvalue(), password))
        else:
            build_workbook(data, str(tmp), masked=masked_build, today=today,
                           capgains=capgains)
        os.replace(tmp, path)
    except BaseException as e:
        # v1.6.2: never leave a half-written .new file behind (it breaks
        # the no-argument workbook auto-detection) — and the cleanup itself
        # must not raise (on Windows the OPEN file may be the tmp)
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        if isinstance(e, PermissionError):
            # the pre-run open check can't catch a file opened DURING the
            # multi-minute fetch — fail in plain words, not a WinError
            raise _fail(f"{path.name} (or its temporary copy) is open in "
                        "another program - close Excel and run the update "
                        "again (your file was not changed)")
        raise


def relock(path: Path, password: str | None = None) -> dict:
    """Offline "put the curtain/safe back" (SPEC §3.19, `--lock`): read →
    regenerate masked/encrypted. No fetching, no computing — seconds, works
    without internet. Used after a viewing run left the file open."""
    ensure_closed(path)
    locked_at_rest = crypto.is_encrypted(path)
    if locked_at_rest and not password:
        raise _fail(f"{path.name} is locked — the password is needed even "
                    "to re-lock it")
    if locked_at_rest:
        try:
            data = read_workbook(crypto.decrypt_workbook(path, password))
        except crypto.WrongPassword:
            raise _fail("wrong password — the file was not touched")
    else:
        data = read_workbook(str(path))
    if not (data.privacy_enabled or data.lock_enabled) or not data.privacy_hash:
        raise _fail("privacy is not switched on for this workbook — "
                    "nothing to lock (see the Settings tab)")
    lock_active = data.lock_enabled and (
        locked_at_rest
        or (bool(password) and crypto.verify_password(password, data.privacy_hash)))
    if data.lock_enabled and not lock_active:
        raise _fail("type your password to lock the file (run by "
                    "double-click, choose --lock)")
    masked_build = data.privacy_enabled
    # "no computing" means reproducing the LAST update's view, not today's:
    # the newest History snapshot carries that run's date, so date-derived
    # content (Capital Gains terms/FY, the FY labels) can't shift under a
    # relock that promised to change nothing. Fresh file with no history →
    # date.today() fallback inside build_workbook.
    for w in data.warnings:
        print(f"⚠  {w}")
    _refuse_overfull(data)
    # v1.6.2: every rewrite path backs up first — relock was the one that
    # didn't, making a read-back edge case unrecoverable
    relock_backup = make_backup(
        path, unmasked=(data.privacy_enabled and not data.masked_at_rest
                        and not locked_at_rest))
    asof = max((s.snap_date for s in data.history if s.snap_date),
               default=None)
    _regenerate_atomic(path, data, masked_build=masked_build,
                       lock_active=lock_active, password=password, today=asof)
    gone = _purge_unmasked_backups(path, keep=relock_backup)
    return {"privacy": ("locked + masked" if lock_active and masked_build
                        else "locked" if lock_active else "masked"),
            "purged_backups": gone, "backup": str(relock_backup),
            "warnings": list(data.warnings), "workbook": str(path)}


def _row(icon: str, label: str, text: str) -> None:
    _say(f"  {icon} {_c('36', f'{label:<11}')} {text}")


def _print_summary(s: dict) -> None:
    _say(_c("36", "  " + "─" * 60))
    if "price_source" in s:
        nse_only = f", {s['nse_only_matched']} NSE-only" if s.get("nse_only_matched") else ""
        _row("📈", "Prices", f"{s['equity_matched']}/{s['equity_total']} stocks matched "
             f"({s['price_source']}{nse_only}) · {s['bonds_matched']} bond(s) priced")
    if "mf_matched" in s:
        _row("📊", "Fund NAVs", f"{s['mf_matched']}/{s['mf_total']} funds matched (AMFI)")
    if s.get("bullion"):
        _row("🪙", "Gold/Silver", s["bullion"])
    if "nps_matched" in s:
        _row("🏛️", "NPS NAVs", f"{s['nps_matched']}/{s['nps_total']} scheme(s) matched (NPS Trust)")
    if s.get("ca_rows") is not None:
        coverage = ("all held stocks verified" if not s.get("ca_unverified")
                    else _c("33", f"{len(s['ca_unverified'])} stock(s) UNVERIFIED"))
        _row("🔀", "Corp acts", f"{s['ca_rows']} action(s) on file · "
             f"{s['ca_adjusted_rows']} holding(s) adjusted · {coverage}")
    if s.get("restructure_children"):
        _row("🧬", "Restructure", f"{s['restructure_children']} new holding row(s) "
             f"appended from a demerger (cost & dates inherited)")
    if s.get("dividend_rows") is not None:
        _row("💸", "Dividends", f"₹{s.get('dividends_fy_total', 0):,.0f} declared this "
             f"financial year ({s['dividend_rows']} row(s) refreshed)")
    if s.get("fmv_filled"):
        _row("🕰️", "FMV filled", f"{s['fmv_filled']} old row(s) got the "
             f"31-01-2018 grandfathering cost")
    if s.get("suspended"):
        _row("⚠️", "Suspended", _c("33", f"{s['suspended']} held scrip(s) not "
                                         f"traded for 21+ days (amber)"))
    if s.get("ppf_ledgered"):
        _row("🏦", "PPF", f"{s['ppf_ledgered']} account(s) computed from the "
             f"deposit ledger (exact interest)")
    x = s.get("portfolio_xirr")
    if x is not None:
        colour = "32" if x >= 0 else "31"
        _row("🎯", "Return", f"portfolio XIRR {_c(colour, f'{x:.2%}')} a year")
    else:
        _row("🎯", "Return", "not enough dated cashflows yet")
    if s.get("fy_expected_total"):
        _row("🔮", "FY-end est", f"₹{s['fy_expected_total']:,.0f} (family total)")
    if s.get("persons_added"):
        _row("👥", "Added", f"sheet(s) for {', '.join(s['persons_added'])}")
    if s.get("classes_toggled"):
        _row("🗂️", "Classes", "; ".join(s["classes_toggled"]))
    if s.get("capgains"):
        _row("🧾", "Cap. gains", s["capgains"] + " (Capital Gains tab)")
    if s.get("privacy"):
        _row("🔒", "Privacy", {
            "masked": "numbers masked (•••)",
            "locked": "file locked (needs your password to open)",
            "locked + masked": "file locked + numbers masked (•••)",
            "open (viewing)": "numbers visible this one time — the next "
                              "update masks them again (or run --lock now)",
        }.get(s["privacy"], s["privacy"]))
    if s.get("backup"):
        _row("💾", "Backup", s["backup"])
    for w in s["warnings"]:
        _say(_c("33", f"  ⚠️  {w}"))
    _say(_c("36", "  " + "─" * 60))
    if s.get("net_worth") is not None:
        _say("  " + _c("1;32", f"🏆 Family net worth: ₹{s['net_worth']:,.0f}")
             + f"   ({s.get('history_points', 0)} day(s) of history — "
               f"watch the Dashboard trend grow)")
    _say(f"  ✅ Updated {s['workbook']}")


RELEASES_API = "https://api.github.com/repos/jay-parikh/NetWorth/releases/latest"
RELEASES_PAGE = "https://github.com/jay-parikh/NetWorth/releases"


def _parse_version(v: str) -> tuple | None:
    """Parse 'v1.2.0' / '1.1.0rc3' → a sortable tuple; a final release sorts
    above any pre-release of the same base (…,1,0) > (…,0,N)."""
    import re
    m = re.match(r"v?(\d+)\.(\d+)\.(\d+)(?:[.\-]?(?:rc|a|b|alpha|beta|pre)\.?(\d+))?",
                 v.strip())
    if not m:
        return None
    major, minor, patch, pre = m.group(1), m.group(2), m.group(3), m.group(4)
    if pre is None:
        return (int(major), int(minor), int(patch), 1, 0)
    return (int(major), int(minor), int(patch), 0, int(pre))


def _latest_release(session=None, timeout: int = 8):
    """→ (parsed_version, tag, url) of the latest GitHub release, or None
    when unreachable / no releases. Never raises."""
    try:
        import requests
        sess = session or requests.Session()
        resp = sess.get(RELEASES_API, timeout=timeout,
                        headers={"Accept": "application/vnd.github+json"})
        if resp.status_code != 200:
            return None
        data = resp.json()
        tag = data.get("tag_name", "") or ""
        parsed = _parse_version(tag)
        if parsed is None:
            return None
        return parsed, tag, data.get("html_url") or RELEASES_PAGE
    except Exception:  # noqa: BLE001 — offline / rate-limited / private repo
        return None


def check_for_update(current: str, session=None, timeout: int = 8) -> str | None:
    """Return a one-line hint if a newer GitHub release exists, else None.
    Never raises — a nicety that must never disturb the update."""
    cur = _parse_version(current)
    if cur is None:
        return None
    latest = _latest_release(session, timeout)
    if latest and latest[0] > cur:
        return f"Update available: {latest[1]} (you have v{current}) — {latest[2]}"
    return None


def version_line(current: str, session=None, timeout: int = 8) -> str:
    """Always says SOMETHING about versions (v1.4): silence used to look
    broken when the user was simply up to date."""
    latest = _latest_release(session, timeout)
    cur = _parse_version(current)
    if latest and cur and latest[0] > cur:
        return f"Update available: {latest[1]} (you have v{current}) — {latest[2]}"
    if latest:
        return f"Version    : v{current} — you're on the latest release"
    return (f"Version    : v{current} (couldn't reach GitHub to check for a "
            f"newer one)")


def peek_persons(path) -> list[str]:
    """Read just the current people from the Dashboard, cheaply (read-only).
    Accepts a path or a decrypted BytesIO (Locked workbooks)."""
    from openpyxl import load_workbook
    if hasattr(path, "seek"):
        path.seek(0)
    wb = load_workbook(path, read_only=True)
    try:
        dash = wb["Dashboard"]
        return [str(dash.cell(r, 1).value).strip()
                for r in range(6, 16)
                if dash.cell(r, 1).value not in (None, "")]
    finally:
        wb.close()


def peek_class_details(path: Path) -> list[tuple[str, bool, bool]]:
    """(label, Settings-enabled, holds-rows) per asset class, cheaply
    (openpyxl read-only). Missing Settings sheet (pre-v1.3 workbook) →
    registry defaults. The holds-rows probe scans each class's holdings
    sheet (owner column; Class column filter for Manual_Assets) — a light
    mirror of class_has_data, close enough for display; run() re-checks
    with the real reader."""
    from openpyxl import load_workbook
    if hasattr(path, "seek"):
        path.seek(0)
    enabled = {c.label: c.default_enabled for c in ASSET_CLASSES}
    has_rows = {c.key: False for c in ASSET_CLASSES}
    wb = load_workbook(path, read_only=True)
    try:
        if "Settings" in wb.sheetnames:
            st = wb["Settings"]
            for row in st.iter_rows(min_row=4, max_row=20, max_col=2):
                label = str(row[0].value or "").strip()
                if label == "Real Estate":            # pre-v1.4.3 label
                    label = "Property"
                if label in enabled and row[1].value is not None:
                    enabled[label] = M.parse_yes_no(row[1].value,
                                                    enabled[label])
        for cls in ASSET_CLASSES:
            sheet = cls.owner_col.split("!")[0]
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            want = cls.class_filter[1].casefold() if cls.class_filter else None
            for row in ws.iter_rows(min_row=4, max_row=min(ws.max_row, 600),
                                    max_col=2, values_only=True):
                owner = str(row[0] or "").strip()
                if not owner:
                    continue
                if want is not None and (
                        str(row[1] or "").strip().casefold() != want):
                    continue
                has_rows[cls.key] = True
                break
    finally:
        wb.close()
    return [(c.label, enabled[c.label], has_rows[c.key])
            for c in ASSET_CLASSES]


def peek_class_states(path: Path) -> list[tuple[str, bool]]:
    """(label, visible?) per asset class. Since v1.4.3 the Settings choice
    IS the state (§3.14) — off means hidden, rows or not."""
    return [(label, on)
            for label, on, _rows in peek_class_details(path)]


def prompt_class_toggle(path: Path) -> list[str]:
    """Interactively show/hide asset classes — easier than editing the
    Settings sheet by hand. Returns the labels to flip."""
    details = peek_class_details(path)
    print("\nYour asset classes — the workbook shows only what's on:")
    for i, (label, on, rows) in enumerate(details, 1):
        state = ("shown" if on else
                 "hidden — holds rows (not counted)" if rows else
                 "hidden")
        print(f"  {i:2}. {label:14} [{state}]")
    print("Show or hide something? Type its number(s), e.g. 7 or 7 9 — "
          "or just press Enter to skip.")
    try:
        raw = input("  Toggle (blank to continue): ").strip()
    except EOFError:
        return []
    toggles: list[str] = []
    for tok in raw.replace(",", " ").split():
        if tok.isdigit() and 1 <= int(tok) <= len(details):
            label, on, rows = details[int(tok) - 1]
            toggles.append(label)
            if on and rows:
                print(f"  ✓ {label} will be hidden — its rows are saved "
                      f"but won't be counted until you show it again")
            elif on:
                print(f"  ✓ {label} will be hidden")
            elif rows:
                print(f"  ✓ {label} will be shown — its saved rows count "
                      f"again")
            else:
                print(f"  ✓ {label} will be shown")
    return toggles


def peek_privacy(source) -> tuple[bool, bool, str]:
    """(mask wanted, lock wanted, stored password fingerprint or "") — cheap
    read-only peek so main() knows which privacy prompt to show (SPEC §3.19);
    the fingerprint lets the mask prompt verify a password on the spot."""
    from openpyxl import load_workbook
    if hasattr(source, "seek"):
        source.seek(0)
    mask = lock = False
    stored_hash = ""
    wb = load_workbook(source, read_only=True)
    try:
        if "Settings" in wb.sheetnames:
            st = wb["Settings"]
            for row in st.iter_rows(min_row=M.SETTINGS_REF_ROW,
                                    max_row=M.SETTINGS_LOCK_ROW + 1,
                                    max_col=2):
                label = str(row[0].value or "").strip()
                val = M.parse_yes_no(row[1].value, False)
                if label == "Privacy mask":
                    mask = val
                elif label.startswith("Lock file"):
                    lock = val
        if "NW_Privacy" in wb.defined_names:
            stored_hash = (wb.defined_names["NW_Privacy"].value
                           or "").strip().strip('"')
    finally:
        wb.close()
    return mask, lock, stored_hash


def _prompt_set_password(for_lock: bool) -> str | None:
    """First-enable flow: choose the privacy password (never echoed, never
    stored — only a fingerprint). Returns None if the user backs out."""
    import getpass
    print("\n🔒 Privacy is switched on — let's set your password.")
    print("   It is never stored anywhere; don't lose it.")
    if for_lock:
        print("   ⚠️  You are LOCKING the file: a forgotten password cannot "
              "be recovered by anyone — write it down somewhere safe.")
    for _ in range(3):
        try:
            p1 = getpass.getpass("  New password (min 4 characters): ")
            if not p1:
                return None
            if len(p1) < 4:
                print("  Too short — 4 characters minimum.")
                continue
            p2 = getpass.getpass("  Type it once more: ")
        except (EOFError, OSError):
            return None
        if p1 == p2:
            return p1
        print("  They don't match — try again.")
    return None


def _prompt_unlock(path: Path) -> str | None:
    """Password prompt for a Locked (encrypted) workbook — 3 tries, each
    verified by a trial decryption. None = give up, file untouched."""
    import getpass
    print(f"\n🔒 {path.name} is locked.")
    for attempt in range(3):
        try:
            pw = getpass.getpass("  Password: ")
        except (EOFError, OSError):
            return None
        if not pw:
            return None
        try:
            crypto.decrypt_workbook(path, pw)
            return pw
        except crypto.WrongPassword:
            print("  Wrong password." + (" Try again." if attempt < 2 else ""))
    return None


def _prompt_mask_password(stored_hash: str) -> tuple[str | None, bool]:
    """Mask-only prompt: (password, reset). One plain question first — Enter
    (the common case) keeps the mask and asks nothing more. Only a "y" leads
    to the password, which is checked on the spot (3 tries) so a typo is
    never silent; RESET at the password prompt is the forgot-it escape."""
    import getpass
    print("\n🔒 Your numbers are masked (•••).")
    try:
        ans = input("  See them this time? Type y — or just press Enter "
                    "to leave them masked: ").strip().casefold()
    except (EOFError, OSError):
        return None, False
    if ans != "y":
        return None, False
    for attempt in range(3):
        try:
            pw = getpass.getpass("  Your password (forgot it? type RESET): ")
        except (EOFError, OSError):
            return None, False
        if not pw:
            return None, False
        if pw.strip() == "RESET":
            print("\n  RESET turns the mask off completely — anyone who "
                  "opens the file will see every number.")
            try:
                sure = input("  Type YES to confirm, or press Enter to "
                             "keep the mask: ").strip()
            except (EOFError, OSError):
                return None, False
            return None, sure == "YES"
        if crypto.verify_password(pw, stored_hash):
            return pw, False
        print("  That password doesn't match."
              + (" Try again." if attempt < 2
                 else " The numbers stay masked this time."))
    return None, False


def prompt_new_persons(existing: list[str],
                       pre_added: list[str] | None = None) -> list[str]:
    """Interactively collect new person names (double-click / Terminal only).
    Adding a person here is optional — you can also just type a name in a yellow
    Dashboard cell. `pre_added` = names already queued via --add-person, so
    the tab prediction and duplicate check see the same final list run() will.
    """
    pre = list(pre_added or [])
    print("\nPeople currently tracked: " + (", ".join(existing) or "(none)"))
    slots = 10 - len(existing) - len(pre)
    if slots <= 0:
        print("(the Dashboard already lists the maximum of 10 people)")
        return []
    print("Add a new person's sheet? Type a name and press Enter, "
          "or just press Enter to skip.")
    added: list[str] = []
    seen = {p.casefold() for p in existing + pre}
    while len(added) < slots:
        try:
            name = input("  New person (blank to continue): ").strip()
        except EOFError:
            break
        if not name:
            break
        if name.casefold() in seen:
            print(f"  '{name}' is already tracked.")
            continue
        seen.add(name.casefold())
        added.append(name)
        # predict from the SAME mapping the build uses — seeding with raw
        # names would mispredict when an earlier tab was already adjusted
        tab = M.person_tab_map(list(existing) + pre + added)[name]
        note = f" (their tab will be called '{tab}')" if tab != name else ""
        print(f"  ✓ will add a sheet for {name}{note} "
              f"(then record holdings with '{name}' in the Owner columns)")
    return added


def _pause() -> None:
    """--pause keeps a double-click console window open; the packaged entry
    always passes it, so a scheduled/headless run (no stdin) must simply
    proceed instead of dying on EOFError after the work is done."""
    try:
        input("\nPress Enter to close...")
    except (EOFError, OSError):
        pass


def _use_os_trust_store() -> None:
    """Validate TLS against the OS certificate store (like browsers and the
    legacy PowerShell did) so corporate/AV proxies don't break the fetch."""
    try:
        import truststore
        truststore.inject_into_ssl()
    except Exception:  # noqa: BLE001 — certifi fallback is fine
        pass


def main(argv: list[str] | None = None) -> int:
    _use_os_trust_store()
    parser = argparse.ArgumentParser(
        description="Refresh prices, NAVs and XIRR in a tracker workbook.")
    parser.add_argument("workbook", nargs="?", help="path to the .xlsx (default: "
                        f"{M.TEMPLATE_FILENAME} next to the current directory)")
    parser.add_argument("--pause", action="store_true",
                        help="wait for Enter before exiting (double-click launchers)")
    parser.add_argument("--no-update-check", action="store_true",
                        help="don't check GitHub for a newer version")
    parser.add_argument("--no-prompt", action="store_true",
                        help="never ask interactive questions (e.g. add a person)")
    parser.add_argument("--add-person", action="append", metavar="NAME", default=[],
                        help="add a person's sheet without prompting (repeatable)")
    parser.add_argument("--lock", action="store_true",
                        help="just put the privacy mask/lock back "
                             "(no fetching; works offline)")
    args = parser.parse_args(argv)

    code = 0
    _banner(f"NetWorth v{__version__} — updating your family portfolio")
    try:
        path = locate_workbook(args.workbook)
        # interactive only when attached to a real console — never hangs a
        # headless/scheduled run
        interactive = sys.stdin.isatty() and not args.no_prompt

        # privacy front door (SPEC §3.19): a Locked file is ciphertext, so
        # the password comes before anything — even the peeks read a
        # decrypted in-memory copy; nothing on disk changes without it
        password: str | None = None
        peek_src = path
        locked = crypto.is_encrypted(path)
        if locked:
            if not interactive:
                raise _fail(f"{path.name} is locked — a scheduled/headless "
                            "run cannot update it; run it by double-click")
            password = _prompt_unlock(path)
            if password is None:
                raise _fail("no password — the file was not touched")
            peek_src = crypto.decrypt_workbook(path, password)

        if args.lock:
            res = relock(path, password=password)
            extra = (f" · removed {res['purged_backups']} readable backup(s)"
                     if res["purged_backups"] else "")
            _row("🔒", "Privacy", res["privacy"] + extra)
            _say(_c("1;32", f"\n  ✅ {Path(res['workbook']).name} is put away"))
        else:
            new_persons = list(args.add_person)
            toggles: list[str] = []
            reveal = reset = False
            if interactive:
                try:
                    new_persons += prompt_new_persons(
                        peek_persons(peek_src), new_persons)
                    toggles = prompt_class_toggle(peek_src)
                except Exception:  # noqa: BLE001 — prompting must never break the run
                    pass
                try:
                    mask_on, lock_on, stored_hash = peek_privacy(peek_src)
                    if (mask_on or lock_on) and not stored_hash:
                        password = _prompt_set_password(lock_on) or password
                    elif lock_on and not locked:
                        # turning the Lock on: confirm the password BEFORE
                        # anything gets encrypted (no lockout by typo)
                        import getpass
                        pw = getpass.getpass(
                            "\n🔒 Confirm your password to LOCK the file: ")
                        password = pw or password
                    elif locked and mask_on:
                        print("\n🔒 Your numbers are masked (•••).")
                        ans = input("  See them this time? Type y — or just "
                                    "press Enter to leave them masked: "
                                    ).strip().casefold()
                        reveal = M.parse_yes_no(ans, False)
                    elif mask_on:
                        pw, reset = _prompt_mask_password(stored_hash)
                        if pw:
                            password, reveal = pw, True
                except Exception:  # noqa: BLE001
                    pass
            summary = run(path, add_persons=new_persons,
                          toggle_classes=toggles, password=password,
                          reveal=reveal, reset_privacy=reset)
            _print_summary(summary)
    except SystemExit as e:
        code = e.code if isinstance(e.code, int) else 1
    except Exception as e:  # noqa: BLE001 — last-resort user-facing message
        print(f"ERROR: {e}", file=sys.stderr)
        code = 1

    if not (args.no_update_check or os.environ.get("NETWORTH_NO_UPDATE_CHECK")):
        line = version_line(__version__)
        _say("  " + (_c("1;33", line) if "Update available" in line
                     else _c("2", line)))

    if args.pause:
        _pause()
    return code


if __name__ == "__main__":
    raise SystemExit(main())
