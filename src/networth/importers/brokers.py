"""Broker equity CSV parsers — registry + generic fallback (SPEC §6.17/§6.18).

Generic by design (Jay, 2026-07-18): exact header signatures are shipped
for brokers we have verified exports from, and everything else goes
through a fuzzy header matcher that looks for the CONCEPTS a tradebook or
holdings file must carry. Matching is on headers only; every VALUE still
has to survive the hardened parsers (common.py) — an unrecognisable file
or line fails politely, naming what was seen, and never guesses.

Both entry points return an ImportBatch; nothing here touches the data
model — merge.py owns that.
"""

from __future__ import annotations

import csv
import io
from datetime import date

from .common import (ImportBatch, ImportedHolding, ImportedTrade,
                     parse_date_any, parse_inr)

# ---- header concepts (casefolded, matched on cleaned header cells) ---------

_SYMBOL_HEADERS = {"symbol", "scrip", "scrip name", "stock", "stock name",
                   "company", "company name", "security", "security name",
                   "instrument", "name of the security", "name"}
_ISIN_HEADERS = {"isin", "isin code", "isin no", "isin no."}
_DATE_HEADERS = {"trade date", "date", "order date", "execution date",
                 "order execution time", "txn date", "transaction date",
                 "deal date"}
_SIDE_HEADERS = {"trade type", "type", "side", "buy/sell", "buy / sell",
                 "transaction type", "b/s", "action"}
_QTY_HEADERS = {"quantity", "qty", "units", "shares", "no. of shares",
                "quantity available", "total quantity", "net qty"}
_PRICE_HEADERS = {"price", "trade price", "rate", "market price",
                  "price per share", "trade rate"}
_AVG_HEADERS = {"average", "average price", "avg price", "avg. price", "avg cost",
                "avg. cost", "buy average", "buy avg", "buy avg.",
                "average cost", "avg trading price", "purchase price",
                "average buy price", "average buying price",
                "avg buy price", "buy average price", "avg rate",
                "avg. rate", "cost rate"}
# traditional back-offices (MoneyMaker etc.) label the avg-cost column just
# "Rate" — but "Rate"/"Net Rate" are ALSO common market-price headers, so
# these only count as avg-cost when NO explicit avg header exists in the
# row (strict-before-loose; a "Rate" column must never shadow "Avg. Cost")
_AVG_LOOSE_HEADERS = {"rate", "net rate", "holding rate"}
_AVG_ANY_HEADERS = _AVG_HEADERS | _AVG_LOOSE_HEADERS
# ledger-style transaction registers (traditional back-offices: one row
# carries buy AND/OR sell columns instead of a buy/sell side column)
_BUYQTY_HEADERS = {"buy qty", "buy quantity", "purchase qty", "bought qty"}
_BUYRATE_HEADERS = {"buy rate", "buy price", "purchase rate",
                    "bought rate"}
_SELLQTY_HEADERS = {"sell qty", "sell quantity", "sale qty", "sold qty"}
_SELLRATE_HEADERS = {"sell rate", "sell price", "sale rate", "sold rate"}
_ACCOUNT_HEADERS = {"client id", "client code", "account", "account id",
                    "ucc", "trading code"}

_BUY_WORDS = {"buy", "b", "purchase", "bought", "credit"}
_SELL_WORDS = {"sell", "s", "sale", "sold", "debit"}

# exact signatures for verified export layouts — matched BEFORE the fuzzy
# fallback so a known broker never depends on heuristics. Keyed by a
# distinctive subset of its headers (all must be present).
SIGNATURES: list[tuple[str, str, frozenset[str]]] = [
    ("Zerodha tradebook", "trades",
     frozenset({"symbol", "isin", "trade_date", "trade_type", "quantity",
                "price"})),
    ("Zerodha holdings", "holdings",
     frozenset({"symbol", "isin", "quantity available", "average price"})),
]


def _clean(h) -> str:
    # trailing periods vary by export vintage ("Qty." vs "Qty") — drop them
    return (str(h or "").strip().strip('"').casefold()
            .replace("_", " ").rstrip("."))


def _find(headers: list[str], concepts: set[str]) -> int | None:
    for i, h in enumerate(headers):
        if h in concepts:
            return i
    return None


def sniff_csv(data) -> tuple[str, str] | None:
    """(source label, kind) when this looks like an equity export we can
    read, else None. Header-based only — values are judged later.
    Accepts CSV text or an already-extracted row list (XLSX path)."""
    headers = _header_row(data)
    if headers is None:
        return None
    hset = set(headers)
    for label, kind, sig in SIGNATURES:
        if {s.replace("_", " ") for s in sig} <= hset:
            return label, kind
    ident = (_find(headers, _ISIN_HEADERS) is not None
             or _find(headers, _SYMBOL_HEADERS) is not None)
    if (ident and _find(headers, _DATE_HEADERS) is not None
            and _find(headers, _SIDE_HEADERS) is not None
            and _find(headers, _QTY_HEADERS) is not None
            and _find(headers, _PRICE_HEADERS) is not None):
        return "broker tradebook", "trades"
    if (ident and _find(headers, _DATE_HEADERS) is not None
            and (_find(headers, _BUYQTY_HEADERS) is not None
                 or _find(headers, _SELLQTY_HEADERS) is not None)):
        return "broker transaction register", "trades_split"
    if (ident and _find(headers, _QTY_HEADERS) is not None
            and _find(headers, _AVG_ANY_HEADERS) is not None
            and _find(headers, _DATE_HEADERS) is None
            and _find(headers, _SIDE_HEADERS) is None
            and _find(headers, _BUYQTY_HEADERS) is None
            and _find(headers, _SELLQTY_HEADERS) is None):
        return "broker holdings", "holdings"
    return None


def _rows(text: str) -> list[list[str]]:
    return [row for row in csv.reader(io.StringIO(text)) if any(
        str(c).strip() for c in row)]


def _header_row(data) -> list[str] | None:
    """The first row that looks like a header line (some exports carry a
    title banner or summary block above it — the wild ones put it at row
    6). Returns cleaned cells. Accepts CSV text or a row list."""
    try:
        rows = data if isinstance(data, list) else _rows(data)
    except csv.Error:
        return None
    for row in rows[:10]:
        cleaned = [_clean(c) for c in row]
        hset = set(cleaned)
        if (hset & (_SYMBOL_HEADERS | _ISIN_HEADERS)) and (
                hset & (_QTY_HEADERS | _PRICE_HEADERS | _AVG_ANY_HEADERS
                        | _BUYQTY_HEADERS | _SELLQTY_HEADERS)):
            return cleaned
    return None


def parse_equity_csv(data, today: date, path: str = "") -> ImportBatch:
    """Parse a tradebook or holdings export into an ImportBatch. Accepts
    CSV text or a row list (the XLSX reader produces one).

    Row-level failures are collected as warnings with the line content;
    merge.py refuses the affected ISIN whole (atomicity), so a half-read
    file can never place a wrong lot on the sheet.
    """
    # CSV text is parsed into rows exactly ONCE; the sniff, the header
    # hunt and the data walk all share the same list
    rows = data if isinstance(data, list) else _rows(data)
    sniffed = sniff_csv(rows)
    if sniffed is None:
        headers = _header_row(rows)
        seen = ", ".join(headers[:8]) if headers else "no header row found"
        raise ValueError(
            "couldn't recognise this file's columns - a tradebook needs "
            "symbol/ISIN, date, buy/sell, quantity and price; a holdings "
            f"file needs symbol/ISIN, quantity and average price (saw: "
            f"{seen})")
    label, kind = sniffed
    batch = ImportBatch(source=label, path=path)
    headers = _header_row(rows)
    start = next(i for i, row in enumerate(rows)
                 if [_clean(c) for c in row] == headers) + 1

    def col(*concept_sets):
        # sets are tried in order: an explicit header (e.g. "avg. cost")
        # always beats a loose one ("rate") anywhere in the row
        for concepts in concept_sets:
            i = _find(headers, concepts)
            if i is not None:
                return lambda row, i=i: (row[i].strip()
                                         if i < len(row) else "")
        return lambda row: ""

    def bad_line(n, row):
        batch.warnings.append(
            f"{label} line {n}: couldn't read it reliably "
            f"({', '.join(str(c) for c in row[:6])})")

    get_sym, get_isin = col(_SYMBOL_HEADERS), col(_ISIN_HEADERS)
    get_qty, get_acct = col(_QTY_HEADERS), col(_ACCOUNT_HEADERS)
    accounts: set[tuple[str, str]] = set()

    if kind == "trades":
        get_date, get_side = col(_DATE_HEADERS), col(_SIDE_HEADERS)
        get_price = col(_PRICE_HEADERS)
        raw: list[ImportedTrade] = []
        for n, row in enumerate(rows[start:], start=start + 1):
            side_txt = get_side(row).strip().casefold()
            side = ("BUY" if side_txt in _BUY_WORDS else
                    "SELL" if side_txt in _SELL_WORDS else "")
            t = ImportedTrade(
                account=get_acct(row), isin=get_isin(row).upper(),
                symbol=get_sym(row),
                trade_date=parse_date_any(get_date(row).split(" ")[0], today),
                qty=parse_inr(get_qty(row)), price=parse_inr(get_price(row)),
                side=side)
            if (t.trade_date is None or not side or not t.qty or t.qty <= 0
                    or not t.price or t.price <= 0
                    or not (t.isin or t.symbol)):
                bad_line(n, row)
                # keep a marker so merge refuses this ISIN/symbol whole
                t.side = "BAD"
                raw.append(t)
                continue
            raw.append(t)
            accounts.add((t.account, ""))
        batch.trades = _collapse(raw)
    elif kind == "trades_split":
        # ledger-style register: one row can carry a buy AND a sell in
        # separate columns (traditional back-offices — MoneyMaker etc.)
        get_date = col(_DATE_HEADERS)
        get_bq, get_br = col(_BUYQTY_HEADERS), col(_BUYRATE_HEADERS)
        get_sq, get_sr = col(_SELLQTY_HEADERS), col(_SELLRATE_HEADERS)
        raw = []
        for n, row in enumerate(rows[start:], start=start + 1):
            base = dict(account=get_acct(row), isin=get_isin(row).upper(),
                        symbol=get_sym(row),
                        trade_date=parse_date_any(
                            get_date(row).split(" ")[0], today))
            legs = []
            bq, sq = parse_inr(get_bq(row)), parse_inr(get_sq(row))
            if bq:
                legs.append(("BUY", bq, parse_inr(get_br(row))))
            if sq:
                legs.append(("SELL", sq, parse_inr(get_sr(row))))
            if not legs:
                continue                     # a totals/empty row — no legs
            for side, qty, price in legs:
                t = ImportedTrade(side=side, qty=qty, price=price, **base)
                if (t.trade_date is None or qty <= 0 or not price
                        or price <= 0 or not (t.isin or t.symbol)):
                    bad_line(n, row)
                    t.side = "BAD"
                raw.append(t)
                if t.side != "BAD":
                    accounts.add((t.account, ""))
        batch.trades = _collapse(raw)
    else:
        get_avg = col(_AVG_HEADERS, _AVG_LOOSE_HEADERS)
        for n, row in enumerate(rows[start:], start=start + 1):
            avg = parse_inr(get_avg(row))
            if not avg or avg <= 0:
                avg = None       # 0 = the broker doesn't KNOW the cost
            h = ImportedHolding(  # (physical→demat) — blank, never ₹0
                account=get_acct(row), isin=get_isin(row).upper(),
                name=get_sym(row), qty=parse_inr(get_qty(row)),
                avg_cost=avg)
            if not (h.isin or h.name) or not h.qty or h.qty <= 0:
                bad_line(n, row)
                continue
            batch.holdings.append(h)
            accounts.add((h.account, ""))
    batch.accounts = sorted(accounts)
    return batch


def _collapse(trades: list[ImportedTrade]) -> list[ImportedTrade]:
    """Multiple exchange fills of one order → one trade per
    (account, isin/symbol, date, side) at the weighted average price.
    BAD markers pass through untouched — they must reach the merge gates."""
    groups: dict[tuple, list[ImportedTrade]] = {}
    out: list[ImportedTrade] = []
    for t in trades:
        if t.side == "BAD":
            out.append(t)
            continue
        groups.setdefault(
            (t.account, t.isin or t.symbol, t.trade_date, t.side),
            []).append(t)
    for (_a, _k, _d, _s), fills in sorted(
            groups.items(), key=lambda kv: (kv[0][2], kv[0][1], kv[0][3])):
        qty = sum(f.qty for f in fills)
        price = sum(f.qty * f.price for f in fills) / qty
        first = fills[0]
        out.append(ImportedTrade(
            account=first.account, isin=first.isin, symbol=first.symbol,
            trade_date=first.trade_date, qty=round(qty, 3),
            price=round(price, 4), side=first.side))
    return out


def rows_from_xlsx(path, max_rows: int = 5000) -> list[list[str]]:
    """First sheet of an Excel export as text rows for the same pipeline
    (traditional back-offices export .xlsx directly). Read-only; dates
    become dd-mm-yyyy strings so the hardened date parser judges them."""
    import warnings
    from datetime import datetime as _dt

    from openpyxl import load_workbook
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        out = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            cells = []
            for c in row:
                if c is None:
                    cells.append("")
                elif isinstance(c, _dt):
                    cells.append(c.strftime("%d-%m-%Y"))
                else:
                    cells.append(str(c))
            if any(x.strip() for x in cells):
                out.append(cells)
        return out
    finally:
        wb.close()
