"""Round-trip reader — reconstructs PortfolioData from an existing workbook.

openpyxl is used strictly READ-ONLY (saving through it destroys charts and
comments). Formula cells come back as "=..." strings, which is exactly how we
tell a generator-written lookup apart from a user's manual override (SPEC §7:
only structured inputs and updater-written values round-trip).

Robustness rules: the header row is located by its known header names within
rows 1–5, and data rows are scanned to the sheet's end, so users inserting,
deleting or sorting rows never breaks the read.
"""

from __future__ import annotations

from datetime import date, datetime

from openpyxl import load_workbook

from .model import (
    ASSET_CLASSES, MANUAL_CLASS_LABELS, SETTINGS_LOCK_ROW, BondRow,
    BullionRow, ClassSetting,
    ClassXirr, CorporateAction, DividendRow, EPFRow, EquityRow, EquitySellRow,
    FDRow,
    HistorySnapshot, ImportMapRow, ImportedFileRow, ManualAssetRow, Masters,
    MFRow, NPSRow, PPFLedgerRow,
    PPFRow, PortfolioData, ScripRef, SIPRow, TaxRule, parse_yes_no,
    person_tab_map,
)

# the Class dropdown is non-blocking, so users can type any casing — Excel's
# SUMIFS matches case-insensitively and the Python side must agree, so typed
# variants are canonicalised on read ("property" → "Property"). "Real Estate"
# was the pre-v1.4.3 label for the Property class; old rows read seamlessly.
_CANON_CLASS = {label.casefold(): label for label in MANUAL_CLASS_LABELS}
_CANON_CLASS["real estate"] = "Property"
# same idiom for the Gold_Silver Type column (v1.6.2)
_CANON_METAL = {"gold": "Gold", "silver": "Silver", "sgb": "SGB"}


def _as_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _as_float(v) -> float | None:
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _num(ws, r: int, c: int, label: str, warnings: list) -> float | None:
    """A make-or-break numeric input cell (v1.6.2): text (or a typed
    formula — these are pure input columns, the generator never writes
    formulas here) where a number belongs used to silently drop the whole
    holding from totals, XIRR and the permanent History snapshot — now it
    says so. The sheet name comes from the worksheet itself."""
    v = ws.cell(r, c).value
    f = _as_float(v)
    if f is None and isinstance(v, str) and v.strip():
        if v.startswith("="):
            warnings.append(
                f"{ws.title} row {r}: a formula in the {label} column "
                "isn't read (and won't survive an update) - type the "
                "plain number")
        else:
            warnings.append(
                f"{ws.title} row {r}: '{v.strip()}' in the {label} column "
                "is not a number - that row is left out of the totals "
                "until it is fixed")
    return f


def _as_str(v) -> str:
    return v.strip() if isinstance(v, str) else ("" if v is None else str(v))


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _manual(v) -> str:
    """A user override in a normally-formula column: plain text, not a formula."""
    return "" if _is_formula(v) or v is None else _as_str(v)


def _header_row(ws, first_header: str) -> int:
    for r in range(1, 6):
        if _as_str(ws.cell(r, 1).value) == first_header:
            return r
    raise ValueError(f"{ws.title}: header row with '{first_header}' not found in rows 1-5")


def _data_rows(ws, header: int):
    for r in range(header + 1, ws.max_row + 1):
        # TOTAL rows carry no Owner/ISIN in column A and 'TOTAL' in column C
        if _as_str(ws.cell(r, 3).value) == "TOTAL":
            continue
        yield r


def read_workbook(source) -> PortfolioData:
    """`source` is a path string or a file-like (the Lock path decrypts to a
    BytesIO so plaintext never touches disk — SPEC §3.19)."""
    wb = load_workbook(source, read_only=False, data_only=False)
    data = PortfolioData()

    # privacy fingerprint + at-rest state (SPEC §3.19) — stored as constant
    # defined names; openpyxl returns the constant WITH its quotes
    def _named(name: str) -> str:
        if name in wb.defined_names:
            return (wb.defined_names[name].value or "").strip().strip('"')
        return ""
    data.privacy_hash = _named("NW_Privacy")
    data.masked_at_rest = _named("NW_Masked") == "yes"

    dash = wb["Dashboard"]
    data.persons = [
        _as_str(dash.cell(r, 1).value)
        for r in range(6, 16) if _as_str(dash.cell(r, 1).value)
    ]
    # v1.6.2: Excel's SUMIFS matches Owner case-insensitively but the
    # Python joins (FY-expected, dividend estimates) were exact-match — a
    # row owned by "JAY" showed in the sheet totals yet dropped out of the
    # computed columns. Canonicalise typed owners onto the person list, the
    # _CANON_CLASS idiom; unknown owners stay as typed.
    _owner_map = {p.casefold(): p for p in data.persons}

    def _owner(v) -> str:
        s = _as_str(v)
        return _owner_map.get(s.casefold(), s)

    # v1.6.2: names that can't be Excel tab names as typed (too long, a /,
    # a duplicate, a fixed sheet's name…) used to crash the rebuild after
    # the whole fetch; now the tab is quietly adjusted — say so (the same
    # person_tab_map the generator builds from, so this can never drift)
    for _p, _tab in person_tab_map(data.persons).items():
        if _tab != _p:
            data.warnings.append(
                f"'{_p}' can't be a tab name as typed - their tab is "
                f"called '{_tab}'; totals and figures are unaffected")
    infl = _as_float(dash["E3"].value)
    data.inflation_pct = infl if infl is not None else 7
    exp = _as_float(dash["E2"].value)
    data.expected_return_pct = exp if exp is not None else 10
    # the FY-expected column moves with the number of enabled classes —
    # locate it by its "Expected @ ..." header in row 5
    fy_col = None
    for c in range(2, 30):
        if _as_str(dash.cell(5, c).value).startswith("Expected @"):
            fy_col = c
            break
    if fy_col:
        for r in range(6, 16):
            person = _as_str(dash.cell(r, 1).value)
            fy = _as_float(dash.cell(r, fy_col).value)
            if person and fy is not None:
                data.fy_expected[person] = fy

    # class XIRRs live in the allocation table — locate its header row and
    # map rows back by class label (layout is dynamic since R10)
    data.xirr = ClassXirr(portfolio=_as_float(dash["B4"].value))
    key_by_label = {c.label: c.key for c in ASSET_CLASSES}
    key_by_label["Real Estate"] = "real_estate"       # pre-v1.4.3 label
    for r in range(15, 40):
        if _as_str(dash.cell(r, 1).value) == "Asset class":
            for rr in range(r + 1, r + 1 + len(ASSET_CLASSES) + 5):
                key = key_by_label.get(_as_str(dash.cell(rr, 1).value))
                if key is None:
                    break
                setattr(data.xirr, key, _as_float(dash.cell(rr, 3).value))
            break

    if "Settings" in wb.sheetnames:
        st = wb["Settings"]

        def _yes_no(row: int, label: str, default: bool) -> bool:
            """One Yes/No reading for EVERY Settings switch (v1.6.2),
            delegating to model.parse_yes_no — the SAME truth the
            interactive peeks use, so a prompt can never disagree with the
            build ('Y' under Privacy mask used to silently mean OFF here
            and on for the peek). Unrecognised non-blank text warns and
            falls back to the row's default."""
            raw = _as_str(st.cell(row, 2).value)
            recognised = (parse_yes_no(raw, True) == parse_yes_no(raw, False))
            if raw and not recognised:
                data.warnings.append(
                    f"Settings: '{raw}' for {label} is not Yes or No - "
                    f"using {'Yes' if default else 'No'}")
            return parse_yes_no(raw, default)

        label_rows = {_as_str(st.cell(r, 1).value): r
                      for r in range(4, SETTINGS_LOCK_ROW + 1)}
        for cls in ASSET_CLASSES:
            r = label_rows.get(cls.label)
            if r is None and cls.key == "real_estate":
                r = label_rows.get("Real Estate")     # pre-v1.4.3 label
            if r is None:
                continue
            data.class_settings[cls.key] = ClassSetting(
                enabled=_yes_no(r, cls.label, cls.default_enabled),
                target_pct=_as_float(st.cell(r, 3).value),
            )
        ref_row = label_rows.get("Reference lists")
        if ref_row:                                   # absent pre-v1.4.3 → No
            data.show_references = _yes_no(ref_row, "Reference lists", False)
        cg_row = label_rows.get("Capital gains report")
        if cg_row:                                    # absent pre-v1.6 → No
            data.show_capital_gains = _yes_no(cg_row, "Capital gains report",
                                              False)
        # privacy switches (SPEC §3.19) — absent pre-v1.5 → both off
        pr = label_rows.get("Privacy mask")
        if pr:
            data.privacy_enabled = _yes_no(pr, "Privacy mask", False)
        lk = next((r for lbl, r in label_rows.items()
                   if lbl.startswith("Lock file")), None)
        if lk:
            data.lock_enabled = _yes_no(lk, "Lock file", False)
        tol_row = next((r for r in range(16, 25)
                        if _as_str(st.cell(r, 1).value).startswith("Drift tolerance")),
                       None)
        if tol_row:
            tol = _as_float(st.cell(tol_row, 2).value)
            if tol is not None:
                data.drift_tolerance_pct = tol
    # no Settings sheet (pre-v1.3 workbook) → registry defaults already set

    ws = wb["Equity"]
    h = _header_row(ws, "Owner")
    for r in _data_rows(ws, h):
        owner = _owner(ws.cell(r, 1).value)
        scrip = _manual(ws.cell(r, 3).value)
        if not owner and not scrip:
            continue
        # R = Flags: "FMV", a restructure flag, or both joined with " | "
        flag_parts = [p.strip()
                      for p in _as_str(ws.cell(r, 18).value).split("|")]
        data.equity.append(EquityRow(
            owner=owner, scrip=scrip,
            qty=_num(ws, r, 4, "Quantity", data.warnings),
            avg_cost=_num(ws, r, 5, "Avg. cost", data.warnings),
            close=_as_float(ws.cell(r, 6).value),
            prev_close=_as_float(ws.cell(r, 7).value),
            close_date=_as_date(ws.cell(r, 8).value),
            cost_date=_as_date(ws.cell(r, 13).value),
            isin_override=_manual(ws.cell(r, 2).value),
            fmv_used="FMV" in flag_parts,
            flag=" | ".join(p for p in flag_parts if p and p != "FMV"),
            ca_factor=_as_float(ws.cell(r, 19).value),            # S = Adj factor
            cost_factor=_as_float(ws.cell(r, 20).value),          # T = Cost factor
        ))

    ws = wb["MutualFunds"]
    h = _header_row(ws, "Owner")
    for r in _data_rows(ws, h):
        owner = _owner(ws.cell(r, 1).value)
        scheme = _manual(ws.cell(r, 3).value)
        if not owner and not scheme:
            continue
        data.mutual_funds.append(MFRow(
            owner=owner, scheme=scheme,
            current_nav=_as_float(ws.cell(r, 7).value),
            xirr=_as_float(ws.cell(r, 12).value),
            fund_house_override=_manual(ws.cell(r, 2).value),
            isin_override=_manual(ws.cell(r, 4).value),
            tax_type=_as_str(ws.cell(r, 13).value),   # v1.6 col M, blank pre-v1.6
        ))

    ws = wb["MF_SIP"]
    h = _header_row(ws, "Owner")
    for r in _data_rows(ws, h):
        owner = _owner(ws.cell(r, 1).value)
        scheme = _manual(ws.cell(r, 3).value)
        if not owner and not scheme:
            continue
        units = ws.cell(r, 8).value
        data.sip.append(SIPRow(
            owner=owner, scheme=scheme,
            txn_date=_as_date(ws.cell(r, 5).value),
            amount=_num(ws, r, 6, "Amount", data.warnings),
            nav=_as_float(ws.cell(r, 7).value),
            units_override=None if _is_formula(units) else _as_float(units),
            fund_house_override=_manual(ws.cell(r, 2).value),
            isin_override=_manual(ws.cell(r, 4).value),
        ))

    ws = wb["FixedDeposits"]
    h = _header_row(ws, "Owner")
    for r in _data_rows(ws, h):
        owner = _owner(ws.cell(r, 1).value)
        if not owner:
            continue
        data.fixed_deposits.append(FDRow(
            owner=owner,
            bank=_as_str(ws.cell(r, 2).value),
            fd_no=_as_str(ws.cell(r, 3).value),
            principal=_num(ws, r, 4, "Principal", data.warnings),
            rate=_as_float(ws.cell(r, 5).value),
            start=_as_date(ws.cell(r, 6).value),
            maturity=_as_date(ws.cell(r, 7).value),
            comp_per_year=(int(v) if (v := _as_float(ws.cell(r, 8).value)) else None),
        ))

    ws = wb["PPF"]
    h = _header_row(ws, "Owner")
    for r in _data_rows(ws, h):
        owner = _owner(ws.cell(r, 1).value)
        if not owner:
            continue
        data.ppf.append(PPFRow(
            owner=owner,
            institution=_as_str(ws.cell(r, 2).value),
            account_no=_as_str(ws.cell(r, 3).value),
            balance=_num(ws, r, 4, "Balance", data.warnings),
            as_on=_as_date(ws.cell(r, 5).value),
            rate=_as_float(ws.cell(r, 6).value),
            notes=_as_str(ws.cell(r, 7).value),
            balance_today=_as_float(ws.cell(r, 8).value),    # H (formula → None)
            interest_earned=_as_float(ws.cell(r, 9).value),  # I
            xirr=_as_float(ws.cell(r, 10).value),            # J
        ))

    if "PPF_Ledger" in wb.sheetnames:
        ws = wb["PPF_Ledger"]
        h = _header_row(ws, "Owner")
        for r in range(h + 1, ws.max_row + 1):
            owner = _owner(ws.cell(r, 1).value)
            acct = _as_str(ws.cell(r, 2).value)
            if not owner and not acct:
                continue
            data.ppf_ledger.append(PPFLedgerRow(
                owner=owner, account_no=acct,
                txn_date=_as_date(ws.cell(r, 3).value),
                amount=_num(ws, r, 4, "Amount", data.warnings),
            ))

    if "EPF" in wb.sheetnames:
        ws = wb["EPF"]
        h = _header_row(ws, "Owner")
        for r in _data_rows(ws, h):
            owner = _owner(ws.cell(r, 1).value)
            if not owner:
                continue
            data.epf.append(EPFRow(
                owner=owner,
                establishment=_as_str(ws.cell(r, 2).value),
                member_id=_as_str(ws.cell(r, 3).value),
                balance=_num(ws, r, 4, "Balance", data.warnings),
                as_on=_as_date(ws.cell(r, 5).value),
                rate=_as_float(ws.cell(r, 6).value),
                notes=_as_str(ws.cell(r, 7).value),
            ))

    if "Gold_Silver" in wb.sheetnames:
        ws = wb["Gold_Silver"]
        h = _header_row(ws, "Owner")
        data.bullion_rate_asof = _as_date(ws["I2"].value)
        for r in _data_rows(ws, h):
            owner = _owner(ws.cell(r, 1).value)
            if not owner:
                continue
            data.bullion.append(BullionRow(
                owner=owner,
                # v1.6.2: canonicalised like Manual_Assets classes — a typed
                # "gold" used to miss the rate refresh with no warning
                metal_type=_CANON_METAL.get(
                    _as_str(ws.cell(r, 2).value).casefold(),
                    _as_str(ws.cell(r, 2).value)),
                description=_as_str(ws.cell(r, 3).value),
                isin=_as_str(ws.cell(r, 4).value),
                qty=_num(ws, r, 5, "Qty", data.warnings),
                purity=_as_float(ws.cell(r, 6).value),
                buy_price=_as_float(ws.cell(r, 7).value),
                buy_date=_as_date(ws.cell(r, 8).value),
                rate_auto=_as_float(ws.cell(r, 9).value),
                rate_override=_as_float(ws.cell(r, 10).value),
                maturity=_as_date(ws.cell(r, 14).value),
            ))

    if "NPS" in wb.sheetnames:
        ws = wb["NPS"]
        h = _header_row(ws, "Owner")
        for r in _data_rows(ws, h):
            owner = _owner(ws.cell(r, 1).value)
            scheme = _manual(ws.cell(r, 3).value)
            if not owner and not scheme:
                continue
            data.nps.append(NPSRow(
                owner=owner,
                pran=_as_str(ws.cell(r, 2).value),
                scheme=scheme,
                units=_num(ws, r, 5, "Units", data.warnings),
                current_nav=_as_float(ws.cell(r, 6).value),
                total_contributed=_as_float(ws.cell(r, 8).value),
                first_contribution=_as_date(ws.cell(r, 9).value),
                xirr=_as_float(ws.cell(r, 10).value),
                scheme_code_override=_manual(ws.cell(r, 4).value),
            ))

    if "Manual_Assets" in wb.sheetnames:
        ws = wb["Manual_Assets"]
        h = _header_row(ws, "Owner")
        for r in _data_rows(ws, h):
            owner = _owner(ws.cell(r, 1).value)
            if not owner:
                continue
            raw_class = _as_str(ws.cell(r, 2).value)
            data.manual_assets.append(ManualAssetRow(
                owner=owner,
                asset_class=_CANON_CLASS.get(raw_class.casefold(), raw_class),
                description=_as_str(ws.cell(r, 3).value),
                institution=_as_str(ws.cell(r, 4).value),
                invested=_as_float(ws.cell(r, 5).value),
                cost_date=_as_date(ws.cell(r, 6).value),
                value=_num(ws, r, 7, "Value today", data.warnings),
                as_on=_as_date(ws.cell(r, 8).value),
                notes=_as_str(ws.cell(r, 10).value),
            ))

    ws = wb["Bonds"]
    h = _header_row(ws, "Owner")
    for r in _data_rows(ws, h):
        owner = _owner(ws.cell(r, 1).value)
        if not owner:
            continue
        data.bonds.append(BondRow(
            owner=owner,
            issuer=_as_str(ws.cell(r, 2).value),
            isin=_as_str(ws.cell(r, 3).value),
            qty=_num(ws, r, 4, "Qty", data.warnings),
            face=_as_float(ws.cell(r, 5).value),
            buy_price=_as_float(ws.cell(r, 6).value),
            cur_price=_as_float(ws.cell(r, 7).value),
            coupon=_as_float(ws.cell(r, 8).value),
            maturity=_as_date(ws.cell(r, 9).value),
            buy_date=_as_date(ws.cell(r, 13).value),
        ))

    ws = wb["By Scrip"]
    h = _header_row(ws, "ISIN")
    for r in _data_rows(ws, h):
        isin = _as_str(ws.cell(r, 1).value)
        if not isin:
            continue
        data.by_scrip.append(ScripRef(isin=isin, name=_as_str(ws.cell(r, 2).value)))

    if "Corporate_Actions" in wb.sheetnames:
        ws = wb["Corporate_Actions"]
        h = _header_row(ws, "Symbol")
        for r in _data_rows(ws, h):
            symbol = _as_str(ws.cell(r, 1).value)
            isin = _as_str(ws.cell(r, 2).value)
            if not symbol and not isin:
                continue
            data.corporate_actions.append(CorporateAction(
                symbol=symbol, isin=isin,
                type=_as_str(ws.cell(r, 3).value).upper(),
                ex_date=_as_date(ws.cell(r, 4).value),
                ratio_from=_as_float(ws.cell(r, 5).value),
                ratio_to=_as_float(ws.cell(r, 6).value),
                source=_as_str(ws.cell(r, 8).value) or "Manual",
                details=_as_str(ws.cell(r, 9).value),
                new_isin=_as_str(ws.cell(r, 10).value),
                cost_pct=_as_float(ws.cell(r, 11).value),
                applied=_as_date(ws.cell(r, 12).value),
            ))

    if "Dividends" in wb.sheetnames:
        ws = wb["Dividends"]
        h = _header_row(ws, "FY")
        for r in _data_rows(ws, h):
            scrip = _as_str(ws.cell(r, 3).value)
            isin = _as_str(ws.cell(r, 4).value)
            if not scrip and not isin:
                continue                     # skips the by-month block too
            data.dividends.append(DividendRow(
                fy=_as_str(ws.cell(r, 1).value),
                owner=_as_str(ws.cell(r, 2).value),
                scrip=scrip, isin=isin,
                div_type=_as_str(ws.cell(r, 5).value).capitalize(),
                ex_date=_as_date(ws.cell(r, 6).value),
                rate=_as_float(ws.cell(r, 7).value),
                qty=_as_float(ws.cell(r, 8).value),
                source=_as_str(ws.cell(r, 10).value) or "Manual",
                details=_as_str(ws.cell(r, 11).value),
            ))

    if "Equity_Sells" in wb.sheetnames:            # absent pre-v1.6 → no-op
        ws = wb["Equity_Sells"]
        h = _header_row(ws, "Owner")
        for r in _data_rows(ws, h):
            owner = _owner(ws.cell(r, 1).value)
            scrip = _manual(ws.cell(r, 3).value)
            if not owner and not scrip:
                continue
            data.equity_sells.append(EquitySellRow(
                owner=owner, scrip=scrip,
                isin_override=_manual(ws.cell(r, 2).value),
                qty=_as_float(ws.cell(r, 4).value),
                buy_date=_as_date(ws.cell(r, 5).value),
                buy_price=_as_float(ws.cell(r, 6).value),
                sell_date=_as_date(ws.cell(r, 7).value),
                sell_price=_as_float(ws.cell(r, 8).value),
                notes=_as_str(ws.cell(r, 11).value),   # cols 9/10 are formulas
            ))

    if "Tax_Rules" in wb.sheetnames:               # absent pre-v1.6 → bundled
        ws = wb["Tax_Rules"]
        h = _header_row(ws, "Asset")
        for r in _data_rows(ws, h):
            asset = _as_str(ws.cell(r, 1).value)
            if not asset:
                continue
            lt = _as_float(ws.cell(r, 3).value)
            data.tax_rules.append(TaxRule(
                asset=asset,
                effective_from=_as_date(ws.cell(r, 2).value),
                lt_days=int(lt) if lt else 365,
                stcg_pct=_as_float(ws.cell(r, 4).value),
                ltcg_pct=_as_float(ws.cell(r, 5).value),
                ltcg_exempt=_as_float(ws.cell(r, 6).value) or 0.0,
                notes=_as_str(ws.cell(r, 7).value),
            ))

    if "Import_Map" in wb.sheetnames:              # absent pre-v1.7 → no-op
        ws = wb["Import_Map"]
        h = _header_row(ws, "Source")
        for r in _data_rows(ws, h):
            account = _as_str(ws.cell(r, 2).value)
            if account:
                data.import_map.append(ImportMapRow(
                    source=_as_str(ws.cell(r, 1).value),
                    account=account,
                    name_hint=_as_str(ws.cell(r, 3).value),
                    owner=_owner(ws.cell(r, 4).value)))
            fname = _as_str(ws.cell(r, 6).value)
            if fname:
                data.imported_files.append(ImportedFileRow(
                    file=fname,
                    fingerprint=_as_str(ws.cell(r, 7).value),
                    imported_on=_as_date(ws.cell(r, 8).value),
                    decision=_as_str(ws.cell(r, 9).value)))

    def master_rows(sheet: str, key_col: int = 3) -> list[tuple[str, str, str]]:
        """Master triples, kept only when the KEY column is non-empty.
        MF/Stock masters key on col 3 (ISIN); NPS_Master keys on col 1
        (Scheme Code) — its col 3 is the PFM name, and a blank PFM must
        never drop a scheme from the master."""
        m = wb[sheet]
        out = []
        for r in range(4, m.max_row + 1):
            if not _as_str(m.cell(r, key_col).value):
                continue
            out.append((_as_str(m.cell(r, 1).value),
                        _as_str(m.cell(r, 2).value),
                        _as_str(m.cell(r, 3).value)))
        return out

    if "History" in wb.sheetnames:
        ws = wb["History"]
        h = _header_row(ws, "Date")
        # columns are label-keyed (SPEC §6.11): a class column may be absent
        # (never enabled) or in any position; unknown labels are ignored
        key_by_label = {c.label: c.key for c in ASSET_CLASSES}
        key_by_label["Real Estate"] = "real_estate"   # pre-v1.4.3 label
        col_key: dict[int, str] = {}
        for c in range(2, 40):
            key = key_by_label.get(_as_str(ws.cell(h, c).value))
            if key:
                col_key[c] = key
        for r in range(h + 1, ws.max_row + 1):
            d = _as_date(ws.cell(r, 1).value)
            if d is None:
                continue
            snap = HistorySnapshot(snap_date=d)
            for c, key in col_key.items():
                setattr(snap, key, _as_float(ws.cell(r, c).value) or 0.0)
            data.history.append(snap)

    sm = wb["Stock_Master"]
    stock_status: dict[str, tuple[str, object]] = {}
    for r in range(4, sm.max_row + 1):
        isin = _as_str(sm.cell(r, 3).value)
        st = _as_str(sm.cell(r, 4).value)
        if isin and st:
            stock_status[isin] = (st, _as_date(sm.cell(r, 5).value))

    # date sanity (v1.6.2): a serial-0 typo reads as 1899-12-30 and a
    # passbook "as-on" of 1899 compounds ₹1L into crores — warn, never
    # silently rewrite (the user fixes the cell; the warning is the guard)
    _floor = date(1980, 1, 1)
    for _attr, _field, _sheet, _label in (
            ("equity", "cost_date", "Equity", "Buy date"),
            ("sip", "txn_date", "MF_SIP", "date"),
            ("fixed_deposits", "start", "FixedDeposits", "Start date"),
            ("ppf", "as_on", "PPF", "As-on date"),
            ("ppf_ledger", "txn_date", "PPF_Ledger", "date"),
            ("epf", "as_on", "EPF", "As-on date"),
            ("bonds", "buy_date", "Bonds", "Buy date"),
            ("bullion", "buy_date", "Gold_Silver", "Buy date"),
            ("nps", "first_contribution", "NPS", "First contribution"),
            ("manual_assets", "cost_date", "Manual_Assets", "Buy date"),
            ("equity_sells", "buy_date", "Equity_Sells", "Buy date"),
            ("equity_sells", "sell_date", "Equity_Sells", "Sell date")):
        for _row in getattr(data, _attr):
            _d = getattr(_row, _field, None)
            if _d and _d < _floor:
                data.warnings.append(
                    f"{_sheet}: a {_label} of {_d:%d-%m-%Y} looks wrong - "
                    "figures using it will be nonsense until it is fixed")

    data.masters = Masters(
        mf_rows=master_rows("MF_Master"),
        stock_rows=master_rows("Stock_Master"),
        nps_rows=(master_rows("NPS_Master", key_col=1)
                  if "NPS_Master" in wb.sheetnames else []),
        mf_refreshed=_as_str(wb["MF_Master"]["E2"].value),
        stock_refreshed=_as_str(wb["Stock_Master"]["E2"].value),
        nps_refreshed=(_as_str(wb["NPS_Master"]["E2"].value)
                       if "NPS_Master" in wb.sheetnames else ""),
        stock_status=stock_status,
    )
    wb.close()
    return data
