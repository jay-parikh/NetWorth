"""v1.6: the capital-gains engine + Capital Gains / Equity_Sells sheets.

Covers the two silent-corruption risks called out in the plan (sell-time
units; FMV normalised across post-2018 corporate actions), the 2024-07-23
mid-FY rate switch, MF FIFO matching, the ₹1.25L shared exemption bucket,
self-explanatory UX (banners + glosses), visibility switch, and the masked
build on the new sheets.
"""

from datetime import date

import pytest
from openpyxl import load_workbook

from networth.compute.capital_gains import capital_gains_report
from networth.generate import build_workbook
from networth.model import (CorporateAction, EquitySellRow, MFRow,
                            PortfolioData, SIPRow, TaxRule, load_tax_rules,
                            tax_rule_for)
from networth.sample_data import sample_portfolio

TODAY = date(2026, 7, 15)

RULES = [
    TaxRule("equity", date(2018, 4, 1), 365, 15, 10, 100000),
    TaxRule("equity", date(2024, 7, 23), 365, 20, 12.5, 125000),
    TaxRule("mf_equity", date(2018, 4, 1), 365, 15, 10, 100000),
    TaxRule("mf_equity", date(2024, 7, 23), 365, 20, 12.5, 125000),
    TaxRule("mf_debt", date(2024, 7, 23), 730, None, 12.5, 0),
]
FMV = ({"INE000A01010": 500.0}, {})


def _sell(**kw) -> PortfolioData:
    d = PortfolioData(persons=["Amit"])
    d.equity_sells = [EquitySellRow(owner="Amit", scrip="ACME LTD.",
                                    isin_override="INE000A01010", **kw)]
    return d


def _one(d, **kw):
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV, **kw)
    assert len(rep.realised) == 1, rep.warnings
    return rep.realised[0], rep


# ---- tax-rule resolution ---------------------------------------------------

def test_bundled_rules_load_and_mid_fy_switch():
    rules = load_tax_rules()
    assert tax_rule_for(rules, "equity", date(2024, 7, 22)).stcg_pct == 15
    assert tax_rule_for(rules, "equity", date(2024, 7, 23)).stcg_pct == 20
    assert tax_rule_for(rules, "equity", date(2017, 6, 1)) is None
    assert tax_rule_for(rules, "mf_debt", date(2025, 1, 1)).lt_days == 730


# ---- grandfathering goldens (buy 2016, FMV 500) ----------------------------

@pytest.mark.parametrize("cost,sale,taxable_cost,note_bit", [
    (100.0, 400.0, 400.0, "grandfathered"),   # min(FMV,sale)=400 > cost
    (100.0, 600.0, 500.0, "grandfathered"),   # FMV caps at 500
    (700.0, 600.0, 700.0, "grandfathered"),   # actual cost stays the highest
    (None, 400.0, 400.0, "buy price blank"),  # blank cost → min(FMV, sale)
])
def test_grandfathering_goldens(cost, sale, taxable_cost, note_bit):
    row, _ = _one(_sell(qty=10, buy_date=date(2016, 6, 1), buy_price=cost,
                        sell_date=date(2026, 5, 1), sell_price=sale))
    assert row.taxable_cost == pytest.approx(10 * taxable_cost)
    assert row.gain == pytest.approx(10 * (sale - taxable_cost))
    assert note_bit in row.note and row.term == "Long-term"


def test_fmv_normalised_across_post_2018_split():
    d = _sell(qty=10, buy_date=date(2016, 6, 1), buy_price=None,
              sell_date=date(2026, 5, 1), sell_price=400.0)
    # 1:2 split in 2020 (face 10 → 5) → today's shares are half-value units:
    # the 2018 per-share FMV halves too
    d.corporate_actions = [CorporateAction(
        symbol="ACME", isin="INE000A01010", type="SPLIT",
        ex_date=date(2020, 6, 1), ratio_from=10, ratio_to=5)]
    row, _ = _one(d)
    assert row.taxable_cost == pytest.approx(10 * 250.0)   # 500 / 2


def test_incomplete_and_impossible_sells_warn_not_guess():
    d = _sell(qty=10, buy_date=date(2020, 1, 1), buy_price=None,
              sell_date=date(2026, 1, 1), sell_price=100.0)  # post-2018 blank
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=({}, {}))
    assert not rep.realised and any("buy price" in w for w in rep.warnings)
    d = _sell(qty=10, buy_date=date(2026, 2, 1), buy_price=50.0,
              sell_date=date(2026, 1, 1), sell_price=100.0)  # sold before buy
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.realised and any("sold before it was bought" in w
                                    for w in rep.warnings)


# ---- MF FIFO ----------------------------------------------------------------

def _mf(sip_rows, tax_type="Equity") -> PortfolioData:
    d = PortfolioData(persons=["Amit"])
    d.mutual_funds = [MFRow(owner="Amit", scheme="FUND X", current_nav=25.0,
                            tax_type=tax_type)]
    d.sip = sip_rows
    return d


def test_mf_fifo_partial_and_multi_lot():
    d = _mf([
        SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0),   # 100 u @10
        SIPRow("Amit", "FUND X", date(2024, 6, 10), 1500, 15.0),   # 100 u @15
        SIPRow("Amit", "FUND X", date(2026, 1, 10), -3000, 20.0),  # sell 150 u
    ])
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert len(rep.realised) == 2                    # 100 from lot1, 50 lot2
    a, b = rep.realised
    assert (a.qty, a.taxable_cost, a.gain) == (100, 1000, pytest.approx(1000))
    assert (b.qty, b.taxable_cost) == (50, pytest.approx(750))
    assert a.term == "Long-term" and b.term == "Long-term"
    # remaining 50 units of lot2 are the open (unrealised) position
    assert len(rep.unrealised) == 1
    assert rep.unrealised[0].qty == pytest.approx(50)
    assert rep.unrealised[0].value_today == pytest.approx(50 * 25.0)


def test_mf_oversell_and_navless_warn():
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 1, 10), -5000, 20.0)])
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert any("more units redeemed" in w for w in rep.warnings)
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 1, 10), -500, None)])
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.realised and any("no NAV" in w for w in rep.warnings)


def test_debt_lots_split_by_purchase_date():
    d = _mf([
        SIPRow("Amit", "FUND X", date(2022, 1, 10), 1000, 10.0),  # old regime
        SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0),  # slab
        SIPRow("Amit", "FUND X", date(2026, 1, 10), -4000, 20.0),
    ], tax_type="Debt")
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert [r.bucket for r in rep.realised] == ["mf_debt", "slab"]
    assert rep.realised[1].term == "At your slab"
    s = rep.summaries[0]
    assert s.debt_gain == pytest.approx(1000) and s.slab_gain == pytest.approx(1000)


def test_blank_tax_type_assumed_equity_with_note():
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 1, 10), -2000, 20.0)],
            tax_type="")
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert rep.realised[0].bucket == "mf_equity"
    assert "assumed Equity" in rep.realised[0].note


# ---- FY summary / exemption bucket -----------------------------------------

def test_shared_exemption_bucket_and_headroom():
    d = _mf([SIPRow("Amit", "FUND X", date(2023, 1, 10), 20000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 10), -60000, 30.0)])
    d.equity_sells = [EquitySellRow(
        owner="Amit", scrip="ACME LTD.", isin_override="INE000A01010",
        qty=100, buy_date=date(2024, 1, 1), buy_price=100.0,
        sell_date=date(2026, 6, 1), sell_price=1100.0)]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    # equity LTCG 100k + mf_equity LTCG 40k share ONE ₹1.25L allowance
    assert s.ltcg == pytest.approx(140000)
    assert s.exemption_used == 125000 and s.headroom == 0
    assert s.tax_ltcg == pytest.approx(15000 * 0.125)
    assert rep.headroom_now == 0


def test_stcg_taxed_per_sale_date_across_mid_fy_switch():
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        EquitySellRow(owner="A", scrip="ACME LTD.",
                      isin_override="INE000A01010", qty=10,
                      buy_date=date(2024, 3, 1), buy_price=100.0,
                      sell_date=date(2024, 7, 1), sell_price=200.0),   # 15%
        EquitySellRow(owner="A", scrip="ACME LTD.",
                      isin_override="INE000A01010", qty=10,
                      buy_date=date(2024, 3, 1), buy_price=100.0,
                      sell_date=date(2024, 8, 1), sell_price=200.0),   # 20%
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.fy == "2024-25"
    assert s.tax_stcg == pytest.approx(1000 * 0.15 + 1000 * 0.20)


def test_lt_on_date_for_sell_planning():
    d = PortfolioData(persons=["Amit"])
    from networth.model import EquityRow
    d.equity = [EquityRow(owner="Amit", scrip="ACME LTD.", qty=10,
                          avg_cost=100, close=150,
                          cost_date=date(2026, 1, 10))]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=({}, {}))
    u = rep.unrealised[0]
    assert u.term == "Short-term"
    assert u.lt_on == date(2027, 1, 11)             # cost_date + 366

# ---- the sheets -------------------------------------------------------------

def test_sheets_default_hidden_and_switch_shows(tmp_path):
    path = tmp_path / "wb.xlsx"
    d = sample_portfolio()
    build_workbook(d, str(path), today=TODAY)
    wb = load_workbook(path)
    assert wb["Equity_Sells"].sheet_state == "hidden"
    assert wb["Capital Gains"].sheet_state == "hidden"
    assert wb["Tax_Rules"].sheet_state == "hidden"
    assert wb["Settings"]["A17"].value == "Capital gains report"
    d.show_capital_gains = True
    build_workbook(d, str(path), today=TODAY)
    wb = load_workbook(path)
    assert wb["Equity_Sells"].sheet_state == "visible"
    assert wb["Capital Gains"].sheet_state == "visible"
    assert wb["Tax_Rules"].sheet_state == "visible"


def test_sheet_ux_teaches_itself(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path), today=TODAY)
    wb = load_workbook(path)
    es = wb["Equity_Sells"]
    assert "reduce the Quantity on the Equity tab" in es["A2"].value
    # worked sample rows present, incl. the blank-buy-price grandfathering demo
    assert es["A4"].value == "Amit" and es["F5"].value is None
    # every jargon header carries a gloss comment
    for cell in ("B3", "D3", "E3", "F3", "I3", "J3"):
        assert es[cell].comment, f"Equity_Sells {cell} lacks a gloss"
    cg = wb["Capital Gains"]
    assert "for planning, not for filing" in cg["A2"].value
    assert "LTCG still tax-free this year" in cg["A3"].value
    assert cg["A3"].comment and "LTCG = " in cg["A3"].comment.text
    assert cg["B6"].comment and cg["B6"].comment.text.startswith("STCG = ")
    # the grandfathering demo row's note is on the sheet, in plain words
    notes = [cg.cell(r, 13).value for r in range(1, 30)
             if cg.cell(r, 13).value]
    assert any("31-Jan-2018" in n for n in notes)
    mf = wb["MutualFunds"]
    assert mf["M3"].value == "Tax type" and mf["M3"].comment


def test_masked_build_covers_new_sheets(tmp_path):
    import re
    import zipfile

    path = tmp_path / "wb.xlsx"
    d = sample_portfolio()
    d.show_capital_gains = True
    d.privacy_enabled = True
    from networth import crypto
    d.privacy_hash = crypto.hash_password("pw")
    build_workbook(d, str(path), masked=True, today=TODAY)
    with zipfile.ZipFile(path) as z:
        charts = [n for n in z.namelist() if re.match(r"xl/charts/chart\d", n)]
        assert not charts                       # masked build: charts stay 0
    wb = load_workbook(path)
    for name in ("Equity_Sells", "Capital Gains"):
        ws = wb[name]
        assert ws.protection.sheet, f"{name} not protected under the mask"
    # no ₹ amount composed into a text cell (the @ mask section shows text
    # verbatim, so "₹1,20,000" inside a sentence would leak through the mask)
    for name in ("Equity_Sells", "Capital Gains"):
        for row in wb[name].iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    assert not re.search(r"₹\s?[\d,]+", c.value), \
                        f"{name}: amount leaked in text cell: {c.value!r}"


def test_roundtrip_with_sells(tmp_path):
    from dataclasses import asdict

    from networth.reader import read_workbook
    path = tmp_path / "wb.xlsx"
    d = sample_portfolio()
    build_workbook(d, str(path), today=TODAY)
    back = read_workbook(str(path))
    assert [asdict(s) for s in back.equity_sells] == \
        [asdict(s) for s in d.equity_sells]
    assert [m.tax_type for m in back.mutual_funds] == ["Equity", "Equity"]


# ---- review fixes (2026-07-18) ----------------------------------------------

def test_bad_tax_rules_csv_warns_instead_of_crashing(tmp_path, monkeypatch):
    bad = tmp_path / "tax_rules_in.csv"
    bad.write_text("asset,effective_from,lt_days,stcg_pct,ltcg_pct,"
                   "ltcg_exempt_inr,notes\nequity,not-a-date,365,15,10,1,x\n")
    with pytest.raises(ValueError):
        load_tax_rules(tmp_path)
    # the engine must degrade to a warning, never crash the build
    # (effective_tax_rules resolves the loader in the model namespace)
    def boom():
        raise ValueError("tax_rules_in.csv: bad row")
    monkeypatch.setattr("networth.model.load_tax_rules", boom)
    d = _sell(qty=10, buy_date=date(2024, 1, 1), buy_price=100.0,
              sell_date=date(2026, 1, 1), sell_price=200.0)
    rep = capital_gains_report(d, TODAY, fmv=FMV)
    assert any("bad row" in w for w in rep.warnings)
    assert rep.realised                       # gains still listed
    assert rep.summaries[0].tax_stcg is None  # tax columns left blank


def test_engine_skipped_when_feature_unused(tmp_path, monkeypatch):
    d = sample_portfolio()
    d.equity_sells = []                       # no sales, switch off (default)

    def boom(*a, **kw):
        raise AssertionError("engine must not run for non-users")
    monkeypatch.setattr(
        "networth.compute.capital_gains.capital_gains_report", boom)
    path = tmp_path / "wb.xlsx"
    build_workbook(d, str(path), today=TODAY)  # must not raise
    wb = load_workbook(path)
    assert "appear here after an update" in wb["Capital Gains"]["A3"].value


def test_tax_type_conflict_warns_first_nonblank_wins():
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 1, 10), -2000, 20.0)])
    d.mutual_funds = [
        MFRow(owner="Amit", scheme="FUND X", current_nav=25.0,
              tax_type="Debt"),
        MFRow(owner="Priya", scheme="FUND X", current_nav=25.0,
              tax_type="Equity"),
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert any("marked both Equity and Debt" in w for w in rep.warnings)
    # first non-blank row (Debt) wins; the 2024 lot lands in the slab bucket
    assert rep.realised[0].bucket == "slab"


def test_full_redemption_paise_rounding_no_phantom_or_oversell():
    # amounts are paise-rounded, units NAV-derived: a legit full exit must
    # not warn "more units redeemed" nor leave a dust lot in "still hold"
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 5000, 7.0),
             SIPRow("Amit", "FUND X", date(2026, 1, 10), -5001, 7.001)])
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not any("more units redeemed" in w for w in rep.warnings)
    assert not rep.unrealised
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 5000, 7.0),
             SIPRow("Amit", "FUND X", date(2026, 1, 10), -4999.9, 7.0)])
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.unrealised                 # ₹0.10 residue is dust


def test_zero_prices_are_real_data():
    # worthless-share write-off: typed sell price 0 = a real capital loss
    row, rep = _one(_sell(qty=10, buy_date=date(2024, 1, 1), buy_price=100.0,
                          sell_date=date(2026, 1, 1), sell_price=0.0))
    assert row.gain == pytest.approx(-1000)
    assert not rep.warnings
    # bonus shares: typed buy price 0 = a real (zero) cost, full gain taxable
    row, _ = _one(_sell(qty=10, buy_date=date(2024, 1, 1), buy_price=0.0,
                        sell_date=date(2026, 1, 1), sell_price=50.0))
    assert row.gain == pytest.approx(500)


def test_stcg_tax_netted_against_losses():
    d = PortfolioData(persons=["A"])
    mk = lambda buy, sell, sd: EquitySellRow(
        owner="A", scrip="ACME LTD.", isin_override="INE000A01010", qty=10,
        buy_date=date(2025, 1, 1), buy_price=buy, sell_date=sd,
        sell_price=sell)
    d.equity_sells = [mk(100.0, 600.0, date(2025, 6, 1)),    # +5000 ST
                      mk(600.0, 300.0, date(2025, 7, 1))]    # -3000 ST
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.stcg == pytest.approx(2000)                     # netted display
    assert s.tax_stcg == pytest.approx(2000 * 0.20)          # tax matches it
    assert s.st_setoff == 0                # losses < gains: nothing spills over


def test_pre_2024_debt_sale_term_uses_old_1095_days():
    d = _mf([SIPRow("Amit", "FUND X", date(2022, 1, 10), 1000, 10.0),
             SIPRow("Amit", "FUND X", date(2024, 1, 10), -2000, 20.0)],
            tax_type="Debt")
    rep = capital_gains_report(d, TODAY, rules=load_tax_rules(), fmv=FMV)
    assert rep.realised[0].bucket == "mf_debt"
    assert rep.realised[0].term == "Short-term"   # 730 days < old 1095


def test_same_day_sale_is_speculative_income_not_cg():
    # intraday round trip → its own bucket, never mixed into STCG/LTCG
    d = _sell(qty=10, buy_date=date(2026, 5, 1), buy_price=50.0,
              sell_date=date(2026, 5, 1), sell_price=60.0)
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.warnings
    row = rep.realised[0]
    assert (row.bucket, row.term) == ("speculative", "Intraday")
    assert row.gain == pytest.approx(100)
    s = rep.summaries[0]
    assert s.spec_gain == pytest.approx(100)
    assert s.stcg == 0 and s.ltcg == 0          # not capital gains
    assert s.tax_stcg == pytest.approx(0)       # no tax computed on it
    # ...but an intraday trade with no buy price can't be computed
    d = _sell(qty=10, buy_date=date(2026, 5, 1), buy_price=None,
              sell_date=date(2026, 5, 1), sell_price=60.0)
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.realised
    assert any("intraday" in w and "buy price" in w for w in rep.warnings)
    # ...and it stays OUT of XIRR (speculative income, not investment return)
    from networth.compute.cashflows import equity_flows
    d = _sell(qty=10, buy_date=date(2026, 5, 1), buy_price=50.0,
              sell_date=date(2026, 5, 1), sell_price=60.0)
    assert equity_flows(d, TODAY) == []


def test_future_sales_warn():
    d = _sell(qty=10, buy_date=date(2026, 1, 1), buy_price=50.0,
              sell_date=date(2027, 1, 1), sell_price=60.0)
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.realised
    assert any("in the future" in w for w in rep.warnings)


def test_negative_qty_or_price_warns_everywhere():
    from networth.compute.cashflows import equity_flows
    d = _sell(qty=-10, buy_date=date(2024, 1, 1), buy_price=50.0,
              sell_date=date(2026, 1, 1), sell_price=60.0)
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.realised
    assert any("negative quantity or price" in w for w in rep.warnings)
    assert equity_flows(d, TODAY) == []
    d = _sell(qty=10, buy_date=date(2024, 1, 1), buy_price=-50.0,
              sell_date=date(2026, 1, 1), sell_price=60.0)
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert not rep.realised and equity_flows(d, TODAY) == []


# ---- Sec 70(2): excess ST loss sets off same-FY LTCG (v1.6.1) ---------------

def _mk_sell(qty, buy, sell, bd, sd):
    return EquitySellRow(owner="A", scrip="ACME LTD.",
                         isin_override="INE000A01010", qty=qty,
                         buy_date=bd, buy_price=buy, sell_date=sd,
                         sell_price=sell)


def test_st_loss_excess_sets_off_ltcg_before_exemption():
    # a loss-harvesting year: net ST loss 2L + LT gain 3L, all in fy_now
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        _mk_sell(100, 3000.0, 1000.0, date(2026, 1, 10), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 4000.0, date(2024, 4, 1), date(2026, 5, 1)),
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.stcg == pytest.approx(-200000)
    assert s.ltcg == pytest.approx(300000)      # RAW — set-off never folded in
    assert s.st_setoff == pytest.approx(200000)
    # exemption applies to the post-set-off 1L, so tax is zero
    assert s.exemption_used == pytest.approx(100000)
    assert s.headroom == pytest.approx(25000)
    assert s.tax_stcg == pytest.approx(0)
    assert s.tax_ltcg == pytest.approx(0)
    assert rep.headroom_now == pytest.approx(25000)


def test_st_setoff_capped_at_ltcg_and_never_carried_forward():
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        # fy_now: ST loss 4L swamps the 1L LT gain
        _mk_sell(100, 5000.0, 1000.0, date(2026, 1, 10), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 2000.0, date(2024, 4, 1), date(2026, 5, 1)),
        # the PREVIOUS FY had a plain 2L LT gain — must stay untouched
        _mk_sell(100, 1000.0, 3000.0, date(2023, 1, 1), date(2025, 6, 1)),
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    now, prev = rep.summaries[0], rep.summaries[1]
    assert now.st_setoff == pytest.approx(100000)   # capped at the LTCG
    assert now.exemption_used == pytest.approx(0, abs=1e-9)
    assert now.tax_ltcg == pytest.approx(0, abs=1e-9)
    # unused 3L of loss is NOT added to headroom (conservative), and
    # carry-forward to other FYs is not modelled
    assert now.headroom == pytest.approx(125000)
    assert prev.fy == "2025-26"
    assert prev.st_setoff == pytest.approx(0, abs=1e-9)
    assert prev.exemption_used == pytest.approx(125000)
    assert prev.tax_ltcg == pytest.approx(75000 * 0.125)


def test_setoff_era_gated_no_rule_no_setoff():
    # pre-2018 §10(38) years: LTCG was exempt, an ST loss could only carry
    # forward — a rule-less FY must show NO set-off, like its blank taxes
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        _mk_sell(10, 600.0, 500.0, date(2017, 5, 1), date(2017, 6, 1)),
        _mk_sell(10, 100.0, 600.0, date(2015, 1, 1), date(2017, 6, 1)),
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.fy == "2017-18" and s.ltcg > 0 and s.stcg < 0
    assert s.st_setoff == 0
    assert s.tax_ltcg is None and s.exemption == 0


def test_st_rows_taxed_at_their_own_assets_rate():
    # Tax_Rules lets equity and mf_equity diverge — an MF short-term gain
    # must then use mf_equity's rate, not equity's
    diverged = RULES + [TaxRule("mf_equity", date(2025, 4, 1), 365, 25, 12.5,
                                125000)]
    d = _mf([SIPRow("Amit", "FUND X", date(2026, 1, 10), 10000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 1), -15000, 15.0)])
    d.equity_sells = [_mk_sell(100, 100.0, 150.0,
                               date(2026, 1, 10), date(2026, 5, 1))]
    d.equity_sells[0].owner = "Amit"
    rep = capital_gains_report(d, TODAY, rules=diverged, fmv=FMV)
    s = rep.summaries[0]
    assert s.stcg == pytest.approx(10000)         # 5k equity + 5k MF
    assert s.tax_stcg == pytest.approx(5000 * 0.20 + 5000 * 0.25)


def test_setoff_dust_clamped_to_keep_blank_when_zero():
    # a sub-half-paisa ST "loss" is float noise, not a set-off
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        _mk_sell(1, 1000.004, 1000.0, date(2026, 1, 10), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 4000.0, date(2024, 4, 1), date(2026, 5, 1)),
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.stcg < 0                              # a real (tiny) net loss
    assert s.st_setoff == 0.0                      # clamped: column stays blank


def test_setoff_column_on_sheet_and_masked_presence(tmp_path):
    # the sheet-side contract: header present, L filled in a harvest year,
    # blank otherwise — and in a MASKED build the cell exists on EVERY row
    # so its presence can't leak a loss-harvest year through the mask
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        _mk_sell(100, 3000.0, 1000.0, date(2026, 1, 10), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 4000.0, date(2024, 4, 1), date(2026, 5, 1)),
        # previous FY: plain LT gain, no set-off → its L cell stays blank
        _mk_sell(100, 1000.0, 3000.0, date(2023, 1, 1), date(2025, 6, 1)),
    ]
    out = tmp_path / "setoff.xlsx"
    build_workbook(d, out, today=TODAY)
    ws = load_workbook(out)["Capital Gains"]
    assert ws["L6"].value == "Losses used vs LTCG ₹"      # v1.6.2 header
    assert ws["L7"].value == pytest.approx(200000)   # newest FY first
    assert ws["L8"].value is None                    # normal year: blank
    m = tmp_path / "setoff-masked.xlsx"
    build_workbook(d, m, masked=True, today=TODAY)
    wm = load_workbook(m)["Capital Gains"]
    assert wm["L7"].value == pytest.approx(200000)
    assert wm["L8"].value == pytest.approx(0)        # written, masked, no leak


def test_debt_st_loss_reduces_equity_stcg_tax():
    # Sec 70(2) across buckets (v1.6.2): a slab-bucket (post-2023 debt)
    # loss wipes the tax on an equal equity ST gain; raw columns stay raw
    d = _mf([SIPRow("Amit", "FUND X", date(2025, 6, 1), 200000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 1), -100000, 5.0)],
            tax_type="Debt")
    d.equity_sells = [_mk_sell(100, 1000.0, 2000.0,
                               date(2026, 1, 10), date(2026, 5, 1))]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.stcg == pytest.approx(100000)        # raw, untouched
    assert s.slab_gain == pytest.approx(-100000)  # raw, untouched
    assert s.tax_stcg == pytest.approx(0)         # loss absorbed the tax
    assert s.st_setoff == 0                       # nothing left for LTCG


def test_debt_st_loss_excess_spills_to_ltcg_before_exemption():
    d = _mf([SIPRow("Amit", "FUND X", date(2025, 6, 1), 400000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 1), -200000, 5.0)],
            tax_type="Debt")                      # slab ST loss 2L
    d.equity_sells = [
        _mk_sell(50, 1000.0, 2000.0, date(2026, 1, 10), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 4000.0, date(2024, 4, 1), date(2026, 5, 1)),
    ]                                             # eq ST +0.5L, eq LT +3L
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.tax_stcg == pytest.approx(0)         # 0.5L gain fully sheltered
    assert s.st_setoff == pytest.approx(150000)   # the excess, vs LTCG
    assert s.exemption_used == pytest.approx(125000)
    assert s.headroom == pytest.approx(0)
    assert s.tax_ltcg == pytest.approx(25000 * 0.125)


def test_debt_lt_loss_reduces_ltcg_but_never_stcg():
    # Sec 70(3): an LT loss only nets against LT gains — cross-asset
    d = _mf([SIPRow("Amit", "FUND X", date(2022, 1, 10), 200000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 1), -100000, 5.0)],
            tax_type="Debt")                      # mf_debt LT loss 1L
    d.equity_sells = [
        _mk_sell(100, 1000.0, 3000.0, date(2024, 4, 1), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 2000.0, date(2026, 1, 10), date(2026, 5, 1)),
    ]                                             # eq LT +2L, eq ST +1L
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert rep.realised[2].term == "Long-term"    # the debt redemption row
    assert s.debt_gain == pytest.approx(-100000)  # raw
    assert s.st_setoff == pytest.approx(100000)   # LT loss applied vs LTCG
    assert s.exemption_used == pytest.approx(100000)
    assert s.tax_ltcg == pytest.approx(0)         # 1L eff < 1.25L allowance
    assert s.tax_stcg == pytest.approx(100000 * 0.20)   # NEVER touched


def test_slab_st_gain_absorbs_equity_st_loss_before_ltcg_spill():
    # the whole short-term head nets first (v1.6.2): an equity ST loss is
    # consumed by a slab-bucket ST gain, so NOTHING spills to LTCG
    d = _mf([SIPRow("Amit", "FUND X", date(2025, 6, 1), 100000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 1), -300000, 30.0)],
            tax_type="Debt")                      # slab ST GAIN +2L
    d.equity_sells = [
        _mk_sell(100, 2000.0, 1000.0, date(2026, 1, 10), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 4000.0, date(2024, 4, 1), date(2026, 5, 1)),
    ]                                             # eq ST -1L, eq LT +3L
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.slab_gain == pytest.approx(200000)   # raw, untouched
    assert s.st_setoff == 0                       # loss absorbed in ST head
    assert s.exemption_used == pytest.approx(125000)
    assert s.tax_ltcg == pytest.approx(175000 * 0.125)


def test_equity_loss_never_changes_debt_display_sums():
    d = _mf([SIPRow("Amit", "FUND X", date(2025, 6, 1), 100000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 5, 1), -300000, 30.0)],
            tax_type="Debt")                      # slab GAIN 2L
    d.equity_sells = [_mk_sell(100, 2000.0, 1000.0,
                               date(2026, 1, 10), date(2026, 5, 1))]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.slab_gain == pytest.approx(200000)   # no tax computed → losses
    assert s.stcg == pytest.approx(-100000)       # are never applied to it
    assert s.st_setoff == 0 and s.tax_stcg == pytest.approx(0)


def test_speculative_loss_never_feeds_the_setoff():
    # Sec 73: an intraday LOSS stays in its own bucket — it must not reduce
    # capital gains
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        _mk_sell(100, 2000.0, 1000.0, date(2026, 5, 1), date(2026, 5, 1)),
        _mk_sell(100, 1000.0, 4000.0, date(2024, 4, 1), date(2026, 5, 1)),
    ]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    s = rep.summaries[0]
    assert s.spec_gain == pytest.approx(-100000)
    assert s.st_setoff == 0
    assert s.exemption_used == pytest.approx(125000)
    assert s.tax_ltcg == pytest.approx(175000 * 0.125)


def test_tax_rules_range_and_duplicate_validation():
    from networth.model import effective_tax_rules
    bad_rate = TaxRule("equity", date(2027, 4, 1), 365, 150, 10, 100000)
    neg_allow = TaxRule("equity", date(2027, 4, 1), 365, 15, 10, -5)
    rules, invalid, warnings = effective_tax_rules([bad_rate, neg_allow])
    assert bad_rate in invalid and neg_allow in invalid
    assert sum("can't be right" in w for w in warnings) == 2
    assert all(t.effective_from != date(2027, 4, 1) for t in rules)
    dup = [TaxRule("equity", date(2027, 4, 1), 365, 20, 12.5, 125000),
           TaxRule("equity", date(2027, 4, 1), 365, 25, 15, 150000)]
    rules, _inv, warnings = effective_tax_rules(dup)
    assert any("two rows for equity" in w for w in warnings)
    assert tax_rule_for(rules, "equity",
                        date(2027, 5, 1)).stcg_pct == 25  # lower row wins


def test_unrealised_mf_rows_carry_caveat_notes():
    d = _mf([SIPRow("Amit", "FUND X", date(2024, 1, 10), 1000, 10.0)],
            tax_type="")
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert rep.unrealised
    assert "assumed Equity" in rep.unrealised[0].note


def test_equity_sells_overflow_warns():
    d = PortfolioData(persons=["A"])
    d.equity_sells = [
        EquitySellRow(owner="A", scrip=f"S{i}", qty=1,
                      buy_date=date(2024, 1, 1), buy_price=1.0,
                      sell_date=date(2025, 1, 1), sell_price=2.0)
        for i in range(201)]
    rep = capital_gains_report(d, TODAY, rules=RULES, fmv=FMV)
    assert any("DROPPED from the saved file" in w for w in rep.warnings)


def test_fy_labels_follow_build_today_not_wall_clock(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path), today=date(2025, 6, 1))
    wb = load_workbook(path)
    assert wb["Amit"]["A4"].value == "Dividends FY 2025-26"
    assert '"2025-26"' in wb["Amit"]["B4"].value
    assert wb["Dashboard"]["A17"].value == "Dividends FY 2025-26"


# ---- Tax_Rules in the workbook (v1.6, §3.22) --------------------------------

def test_workbook_rule_overrides_bundled_exemption():
    # Govt raises the allowance to ₹2L: user edits the 2024 equity row in
    # Excel — no app release needed
    d = _sell(qty=10, buy_date=date(2024, 1, 1), buy_price=100.0,
              sell_date=date(2026, 5, 1), sell_price=200.0)   # current FY
    d.tax_rules = [TaxRule("equity", date(2024, 7, 23), 365, 20, 12.5,
                           200000)]
    rep = capital_gains_report(d, TODAY, fmv=FMV)     # rules from the data
    assert rep.summaries[0].exemption == 200000
    assert rep.headroom_now == pytest.approx(200000 - 1000)


def test_user_added_future_rule_wins_from_its_date():
    from networth.model import effective_tax_rules, tax_rule_for
    user = [TaxRule("equity", date(2027, 4, 1), 365, 25, 15, 300000,
                    notes="Budget 2027")]
    rules, invalid, warnings = effective_tax_rules(user)
    assert not invalid and not warnings
    assert tax_rule_for(rules, "equity", date(2027, 5, 1)).stcg_pct == 25
    # sales before the new date keep the old regime
    assert tax_rule_for(rules, "equity", date(2027, 3, 1)).stcg_pct == 20
    # bundled rows the workbook didn't touch are still there
    assert tax_rule_for(rules, "mf_debt", date(2024, 1, 1)).lt_days == 1095


def test_invalid_workbook_rule_warns_and_is_kept_on_sheet(tmp_path):
    from networth.model import effective_tax_rules
    bad = TaxRule("shares", date(2027, 4, 1), 365, 25, 15, 0)
    rules, invalid, warnings = effective_tax_rules([bad])
    assert bad in invalid and any("Tax_Rules" in w for w in warnings)
    assert all(t.asset != "shares" for t in rules)    # never computed with
    # ...and the writer keeps it visible for fixing instead of dropping it
    d = sample_portfolio()
    d.tax_rules = load_tax_rules() + [bad]
    path = tmp_path / "wb.xlsx"
    build_workbook(d, str(path), today=TODAY)
    wb = load_workbook(path)
    col_a = [wb["Tax_Rules"].cell(r, 1).value for r in range(4, 34)]
    assert "shares" in col_a


def test_tax_rules_sheet_roundtrips_and_prefills_bundled(tmp_path):
    from dataclasses import asdict

    from networth.reader import read_workbook
    path = tmp_path / "wb.xlsx"
    d = sample_portfolio()
    build_workbook(d, str(path), today=TODAY)
    back = read_workbook(str(path))
    assert [asdict(t) for t in back.tax_rules] == \
        [asdict(t) for t in load_tax_rules()]
    # an edit survives the next regeneration (the workbook is the truth)
    back.tax_rules[0].ltcg_exempt = 999999.0
    build_workbook(back, str(path), today=TODAY)
    again = read_workbook(str(path))
    assert any(t.ltcg_exempt == 999999.0 for t in again.tax_rules)
