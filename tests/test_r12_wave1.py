"""R12: Manual_Assets (RE/Cash/Insurance/Other) + EPF (SPEC §3.17/§3.18)."""

from dataclasses import asdict
from datetime import date

import pytest
from openpyxl import load_workbook

from networth.compute.cashflows import manual_asset_flows
from networth.compute.snapshot import net_worth_snapshot
from networth.compute.xirr import xirr
from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.model import (ClassSetting, EPFRow, ManualAssetRow,
                            current_epf_rate, load_epf_rates)
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import run

TODAY = date(2026, 7, 15)


def _wave1_data():
    data = sample_portfolio()
    data.class_settings["real_estate"] = ClassSetting(enabled=True)
    data.class_settings["cash"] = ClassSetting(enabled=True)
    data.class_settings["epf"] = ClassSetting(enabled=True)
    data.manual_assets = [
        ManualAssetRow(owner="Amit", asset_class="Real Estate",
                       description="2BHK Baner", invested=4500000,
                       cost_date=date(2016, 7, 15), value=9000000,
                       as_on=date(2026, 7, 1)),
        ManualAssetRow(owner="Priya", asset_class="Cash",
                       description="HDFC savings", institution="HDFC Bank",
                       value=250000, as_on=date(2026, 7, 10)),
    ]
    data.epf = [EPFRow(owner="Amit", establishment="AcmeCorp / UAN 100200",
                       member_id="MH/12345/678", balance=1500000,
                       as_on=date(2026, 3, 31), rate=8.25)]
    return data


def test_epf_rates_table():
    rates = load_epf_rates()
    assert rates[0][0] <= 2015 and rates[-1][0] >= 2024
    assert current_epf_rate(rates) == rates[-1][1]
    assert dict(rates)[2021] == 8.1          # the famous 44-year low


def test_re_xirr_matches_hand_checked_two_flow():
    data = _wave1_data()
    flows = manual_asset_flows(data, TODAY, "Real Estate")
    assert flows == [(date(2016, 7, 15), -4500000), (TODAY, 9000000)]
    # 45L → 90L over 10.0 years ≈ 7.18% — hand-checked doubling rate
    assert xirr(flows) == pytest.approx(0.0718, abs=0.0005)
    # cash never produces flows
    assert manual_asset_flows(data, TODAY, "Cash") == []


def test_snapshot_and_dashboard_columns(tmp_path):
    data = _wave1_data()
    snap = net_worth_snapshot(data, TODAY)
    assert snap.real_estate == 9000000
    assert snap.cash == 250000
    # EPF accrued from 31-03-2026 at 8.25% for ~106 days
    assert snap.epf == pytest.approx(1500000 * 1.0825 ** (106 / 365), rel=1e-6)
    assert snap.insurance == 0.0

    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    d = wb["Dashboard"]
    headers = [d.cell(5, c).value for c in range(1, 11)]
    assert headers == ["Person", "Equity", "Mutual Funds", "Fixed Deposits",
                       "PPF", "EPF", "Bonds", "Real Estate", "Cash", "Total"]
    # shared-sheet classes filter by their Class column
    assert ('SUMIFS(Manual_Assets!$G:$G,Manual_Assets!$A:$A,$A6,'
            'Manual_Assets!$B:$B,"Real Estate")') in d["H6"].value
    assert wb["Manual_Assets"].sheet_state == "visible"
    assert wb["EPF"].sheet_state == "visible"
    # Cash has no XIRR cell (has_xirr false) — its allocation XIRR is blank
    alloc = {d.cell(r, 1).value: d.cell(r, 3).value for r in range(20, 29)}
    assert alloc["Cash"] is None


def test_shared_sheet_hides_only_when_all_subclasses_off(tmp_path):
    data = _wave1_data()
    data.manual_assets = [r for r in data.manual_assets
                          if r.asset_class == "Real Estate"]
    data.class_settings["cash"] = ClassSetting(enabled=False)
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    assert wb["Manual_Assets"].sheet_state == "visible"   # RE still on

    data.class_settings["real_estate"] = ClassSetting(enabled=False)
    data.manual_assets = []
    build_workbook(data, str(path))
    wb = load_workbook(path)
    assert wb["Manual_Assets"].sheet_state == "hidden"    # all four off


def test_epf_sheet_formula_and_rate_autofill(tmp_path):
    data = _wave1_data()
    data.epf[0].rate = None                    # blank → updater fills
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    e = wb["EPF"]
    assert "YEARFRAC($E4,TODAY())" in e["H4"].value
    run(path, price_data=PriceData(trade_date=TODAY, source="T"),
        amfi_data=AmfiData(), ca_data=[], div_data=[], today=TODAY)
    back = read_workbook(str(path))
    assert back.epf[0].rate == current_epf_rate()


def test_wave1_round_trip_and_history(tmp_path):
    data = _wave1_data()
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    back = read_workbook(str(path))
    assert asdict(back.manual_assets[0]) == asdict(data.manual_assets[0])
    assert asdict(back.epf[0]) == asdict(data.epf[0])
    # full round-trip identity with the new classes populated
    path2 = tmp_path / "wb2.xlsx"
    build_workbook(back, str(path2))
    assert asdict(read_workbook(str(path2))) == asdict(back)

    # an updater run writes the new classes into History (label-keyed)
    run(path, price_data=PriceData(trade_date=TODAY, source="T"),
        amfi_data=AmfiData(), ca_data=[], div_data=[], today=TODAY)
    wb = load_workbook(path)
    h = wb["History"]
    headers = [h.cell(3, c).value for c in range(1, 12)]
    assert "Real Estate" in headers and "EPF" in headers and "Cash" in headers
    again = read_workbook(str(path))
    assert again.history[-1].real_estate == 9000000
    assert again.history[-1].cash == 250000
