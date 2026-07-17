"""v1.4.3 "calm first open": the Settings choice wins everywhere.

A switched-off class is hidden AND excluded from every displayed number —
Dashboard, allocation, person tabs, Projection, portfolio XIRR, and new
History snapshots — while its rows stay saved. Reference sheets tuck away
behind one Settings switch, and the tab strip is colour-coded.
"""

import re
import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook

from networth.compute.cashflows import compute_all_xirr
from networth.compute.projections import fy_expected_by_person
from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.model import ClassSetting, EPFRow
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import run

TODAY = date(2026, 7, 15)


def _run(path):
    return run(path, price_data=PriceData(trade_date=TODAY, source="T"),
               amfi_data=AmfiData(), ca_data=[], div_data=[], today=TODAY)


def test_portfolio_xirr_combines_only_shown_classes():
    from conftest import classic
    data = classic()                                 # bonds on, holds a row
    x_on = compute_all_xirr(data, TODAY)

    data.class_settings["bonds"] = ClassSetting(enabled=False)
    x_off = compute_all_xirr(data, TODAY)
    assert x_off.bonds is not None                   # per-class still computed
    assert x_off.portfolio != x_on.portfolio         # family figure excludes it

    stripped = classic()
    stripped.bonds = []                              # same flows as "off"
    assert compute_all_xirr(stripped, TODAY).portfolio == pytest.approx(
        x_off.portfolio)


def test_projection_excludes_hidden_classes():
    from conftest import classic
    data = classic()
    data.epf = [EPFRow(owner="Amit", establishment="Acme", balance=1000000,
                       as_on=date(2026, 3, 31), rate=8.25)]
    base = fy_expected_by_person(data, TODAY)        # epf is OFF in classic
    data.class_settings["epf"] = ClassSetting(enabled=True)
    with_epf = fy_expected_by_person(data, TODAY)
    assert with_epf["Amit"] > base["Amit"] + 1000000 * 0.99


def test_history_snapshot_zeroes_hidden_classes(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    _run(path)
    back = read_workbook(str(path))
    last = back.history[-1]
    assert last.equity > 0 and last.bonds > 0        # shown classes counted
    assert last.epf == 0.0 and last.gold_silver == 0.0
    assert last.real_estate == 0.0                   # Property is off
    assert back.epf and back.bullion                 # ...but the rows survive


def test_tab_colours_present(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)

    def sheet_xml(name):
        idx = wb.sheetnames.index(name) + 1
        with zipfile.ZipFile(path) as z:
            return z.read(f"xl/worksheets/sheet{idx}.xml").decode()

    assert re.search(r'tabColor[^/]*1F4E79', sheet_xml("Dashboard"))
    assert re.search(r'tabColor[^/]*4472C4', sheet_xml("Equity"))
    assert re.search(r'tabColor[^/]*BF8F00', sheet_xml("Guide"))
    assert re.search(r'tabColor[^/]*31859C', sheet_xml("Amit"))


def test_old_real_estate_label_still_maps():
    from networth.reader import _CANON_CLASS
    assert _CANON_CLASS["real estate"] == "Property"
    assert _CANON_CLASS["property"] == "Property"


def test_guide_mentions_reference_lists_and_gold_steps():
    from networth.guide_text import GUIDE_ROWS
    text = " ".join(str(part) for row in GUIDE_ROWS for part in row)
    assert "Reference lists" in text
    assert "22K = 0.916" in text                     # the gold how-to
    assert "Your choice wins" in text                # no delete-to-hide talk
