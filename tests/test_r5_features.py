"""R5: red/green formats, bond maturity + coupons, bank master, FY-end value."""

import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook

from networth.compute.cashflows import bond_coupon_flows, coupon_dates
from networth.compute.projections import fy_end, fy_expected_by_person
from networth.compute.xirr import xirr
from networth.generate import build_workbook
from networth.model import BondRow
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio

TODAY = date(2026, 7, 15)


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    # classic five-class layout — these tests assert the R5-era positions
    from conftest import classic
    path = tmp_path_factory.mktemp("wb") / "template.xlsx"
    build_workbook(classic(), str(path))
    return path


@pytest.fixture(scope="module")
def wb(built):
    return load_workbook(built, data_only=False)


# ---- coupons & bond XIRR ----

def test_coupon_dates_annual():
    dates = coupon_dates(date(2029, 1, 15), after=date(2024, 1, 15),
                         before=date(2026, 7, 15))
    assert dates == [date(2025, 1, 15), date(2026, 1, 15)]


def test_bond_xirr_includes_coupons():
    bond = BondRow("Rahul", "NHAI", "INE906B07CB9", qty=50, face=1000,
                   buy_price=1000, cur_price=1015, coupon=8.5,
                   maturity=date(2029, 1, 15), buy_date=date(2024, 1, 15))
    coupons = bond_coupon_flows(bond, TODAY)
    assert len(coupons) == 2 and coupons[0][1] == pytest.approx(50 * 1000 * 0.085)
    flows = [(bond.buy_date, -50 * 1000)] + coupons + [(TODAY, 50 * 1015)]
    r = xirr(flows)
    # ~8.5% coupon + small price gain over 2.5y → high single digits
    assert 0.07 < r < 0.11


# ---- FY-end expected value ----

def test_fy_end_boundary():
    assert fy_end(date(2026, 7, 15)) == date(2027, 3, 31)
    assert fy_end(date(2027, 3, 31)) == date(2027, 3, 31)
    assert fy_end(date(2027, 4, 1)) == date(2028, 3, 31)


def test_fy_expected_values():
    from conftest import classic
    data = classic()
    out = fy_expected_by_person(data, TODAY)
    assert set(out) == {"Amit", "Priya", "Rahul"}
    end = date(2027, 3, 31)
    growth = 1.10 ** ((end - TODAY).days / 365)
    # Rahul: equity grows at expected return; FD accrues to FY end;
    # bond redeems + pays coupons only after FY end → price value + 1 coupon
    eq = (25 * 3620 + 120 * 815 + 30 * 2410) * growth
    n = 4
    fd = 300000 * (1 + 0.068 / n) ** (n * ((end - date(2025, 1, 15)).days / 365))
    bond = 50 * 1015 + 50 * 1000 * 0.085 * 1     # coupon on 15-01-2027 ≤ FY end
    assert out["Rahul"] == pytest.approx(eq + fd + bond, rel=1e-6)


def test_fy_bond_redeemed_within_fy():
    data = sample_portfolio()
    data.bonds[0].maturity = date(2026, 12, 31)
    out = fy_expected_by_person(data, TODAY)
    # redemption at face (50*1000) replaces price value (50*1015)
    data2 = sample_portfolio()
    delta = out["Rahul"] - fy_expected_by_person(data2, TODAY)["Rahul"]
    # face redemption vs price+coupon path: -750 price diff, coupon timing shifts
    assert delta != 0


# ---- workbook structure ----

def test_bank_master_and_dropdown(built, wb):
    assert "Bank_Master" in wb.sheetnames
    bm = wb["Bank_Master"]
    assert bm["A3"].value == "Bank Name"
    names = [bm.cell(r, 1).value for r in range(4, 30)]
    assert "HDFC Bank" in names
    assert names == sorted(names, key=lambda s: s.casefold())
    assert "Bank_NameList" in set(wb.defined_names)
    with zipfile.ZipFile(built) as z:
        idx = wb.sheetnames.index("FixedDeposits") + 1
        fd = z.read(f"xl/worksheets/sheet{idx}.xml").decode()
    assert "OFFSET(Bank_Master!$A$3" in fd and 'sqref="B4:B53"' in fd


def test_bond_maturity_columns(wb):
    b = wb["Bonds"]
    assert b["O3"].value == "Maturity Value"
    assert b["O4"].value == '=IF(OR($D4="",$E4=""),"",$D4*$E4)'
    assert "YEARFRAC(TODAY(),$I4)" in b["P4"].value
    assert b["O55"].value == "=SUM(O4:O53)"


def test_redgreen_conditional_formats(built, wb):
    with zipfile.ZipFile(built) as z:
        def sheet_xml(name):
            return z.read(f"xl/worksheets/sheet{wb.sheetnames.index(name) + 1}.xml").decode()
        eq = sheet_xml("Equity")
        dash = sheet_xml("Dashboard")
        amit = sheet_xml("Amit")
    assert 'sqref="K4:L253"' in eq and "conditionalFormatting" in eq
    assert 'sqref="C20:C24"' in dash
    assert 'sqref="F16:G55"' in amit


def test_dashboard_fy_column(wb):
    d = wb["Dashboard"]
    assert d["E2"].value == 10
    assert str(d["H5"].value).startswith("Expected @ 31-Mar-")
    assert d["H16"].value == '=IF(SUM(H6:H15)=0,"",SUM(H6:H15))'


def test_fy_roundtrip(tmp_path):
    data = sample_portfolio()
    data.expected_return_pct = 12
    data.fy_expected = {"Amit": 1234567.89, "Priya": 100.5}
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    back = read_workbook(str(path))
    assert back.expected_return_pct == 12
    assert back.fy_expected == {"Amit": 1234567.89, "Priya": 100.5}
