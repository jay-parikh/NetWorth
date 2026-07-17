"""v1.5.1 keep + gloss: terms stay, hover comments explain them.

Two invariants, both from Jay's review:
1. Domain terms are NEVER renamed or softened — headers and banners keep
   compute / NAV / ISIN / XIRR etc. verbatim.
2. Each jargon-bearing header carries a plain-language comment that leads
   with the term; obscure phrases (IBJA, LTCG/FMV/grandfathering, Alt+Down)
   no longer appear in any user-facing text.
"""

from openpyxl import load_workbook

from networth.generate import build_workbook
from networth.sample_data import sample_portfolio


def _wb(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    return load_workbook(str(path))


def _comment(wb, sheet, cell):
    c = wb[sheet][cell].comment
    return c.text if c else ""


def test_terms_stay_and_glosses_lead_with_the_term(tmp_path):
    wb = _wb(tmp_path)
    # headers unchanged (keep) …
    assert wb["Equity"]["N3"].value == "XIRR"
    assert wb["Equity"]["B3"].value == "ISIN"
    assert wb["MutualFunds"]["G3"].value == "Current NAV"
    assert wb["FixedDeposits"]["H3"].value == "Comp./yr"
    assert wb["NPS"]["B3"].value == "PRAN"
    assert wb["Dashboard"]["A4"].value == "Portfolio XIRR"
    # … and each gloss comment leads WITH the term (gloss)
    assert _comment(wb, "Equity", "N3").startswith("XIRR = ")
    assert _comment(wb, "Equity", "B3").startswith("ISIN = ")
    assert "NAV = " in _comment(wb, "MutualFunds", "G3")
    assert "times a year" in _comment(wb, "FixedDeposits", "H3")
    assert "NPS account number" in _comment(wb, "NPS", "B3")
    assert "UAN" in _comment(wb, "EPF", "B3")
    assert "Sovereign Gold Bond" in _comment(wb, "Gold_Silver", "B3")
    assert "ex-date decides" in _comment(wb, "Dividends", "F3")
    assert "Corpus = your total money" in _comment(wb, "Projection", "B3")
    assert "inflation" in _comment(wb, "Dashboard", "D4")
    assert "Drift = " in _comment(wb, "Dashboard", "F19")
    assert _comment(wb, "Amit", "C5")            # '# holdings' gloss


def test_obscure_phrases_are_gone_from_user_text(tmp_path):
    wb = _wb(tmp_path)
    blobs = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=4):
            blobs += [str(c.value) for c in row if isinstance(c.value, str)]
        blobs += [c.comment.text for r in ws.iter_rows(max_row=4)
                  for c in r if c.comment]
    text = " ".join(blobs)
    for banished in ("Alt+Down", "IBJA", "grandfathering", "LTCG",
                     "Yellow-ish", "monthly-minimum-balance",
                     "market-implied"):
        assert banished not in text, banished
    # banners kept their precise terms
    assert "compute themselves" in text                    # FD banner
    assert "stay out of XIRR" in text                      # Bonds banner
    assert "the ISIN fills itself" in text                 # Equity banner