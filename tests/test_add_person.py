"""Runtime 'add a person' — declarative append that regenerates their sheet."""

from datetime import date

import pytest
from openpyxl import load_workbook

from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import peek_persons, run

TODAY = date(2026, 7, 15)


def _sources():
    return dict(price_data=PriceData(trade_date=TODAY, source="T"),
                amfi_data=AmfiData(), ca_data=[], today=TODAY)


def test_peek_persons_reads_dashboard(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    assert peek_persons(path) == ["Amit", "Priya", "Rahul"]


def test_add_person_creates_sheet_and_dashboard_row(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    summary = run(path, add_persons=["Neha"], **_sources())
    assert summary["persons_added"] == ["Neha"]

    wb = load_workbook(path)
    assert "Neha" in wb.sheetnames                       # sheet generated
    assert wb["Neha"]["B2"].value == "Neha"
    dash_names = [wb["Dashboard"].cell(r, 1).value for r in range(6, 16)]
    assert "Neha" in dash_names                           # Dashboard row filled
    byscrip_hdr = [c.value for c in wb["By Scrip"][3]]
    assert "Neha" in byscrip_hdr                          # her per-person column
    back = read_workbook(str(path))
    assert back.persons == ["Amit", "Priya", "Rahul", "Neha"]


def test_add_person_dedupes_and_is_idempotent(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    s1 = run(path, add_persons=["Priya", "priya", "Neha"], **_sources())
    assert s1["persons_added"] == ["Neha"]               # existing/dupes ignored
    s2 = run(path, add_persons=["Neha"], **_sources())
    assert s2["persons_added"] == []                      # already there
    assert read_workbook(str(path)).persons.count("Neha") == 1


def test_prompt_collects_dedupes_and_stops_on_blank(monkeypatch):
    from networth.update import prompt_new_persons
    answers = iter(["Neha", "priya", "Neha", "Vikram", ""])
    monkeypatch.setattr("builtins.input", lambda *a: next(answers))
    assert prompt_new_persons(["Amit", "Priya", "Rahul"]) == ["Neha", "Vikram"]


def test_prompt_skips_when_full(monkeypatch):
    from networth.update import prompt_new_persons
    monkeypatch.setattr("builtins.input", lambda *a: "ShouldNotAsk")
    assert prompt_new_persons([f"P{i}" for i in range(10)]) == []


def test_add_person_respects_ten_cap(tmp_path):
    path = tmp_path / "wb.xlsx"
    data = sample_portfolio()
    data.persons = [f"P{i}" for i in range(9)]            # 9 existing → 1 slot
    build_workbook(data, str(path))
    summary = run(path, add_persons=["A", "B", "C"], **_sources())
    assert summary["persons_added"] == ["A"]
    assert len(read_workbook(str(path)).persons) == 10
