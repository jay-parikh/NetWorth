"""v1.6.2 "Sturdy" — the stability batch.

Covers the whole-product sweep verdicts (2026-07-18): silent row loss,
backup gaps, empty-feed distrust, person-name crashes, XIRR extremes,
future-dated rows, text-in-number cells, Yes/No asymmetry, case joins and
the friendly late-failure path.
"""

import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook

from networth import crypto
from networth.compute.capital_gains import capital_gains_report
from networth.compute.cashflows import (flat_accrual, mf_flows_by_fund,
                                        ppf_flows)
from networth.compute.xirr import xirr
from networth.generate import build_workbook
from networth.model import (EquityRow, PPFLedgerRow, PPFRow, PortfolioData,
                            SIPRow, person_sheet_name)
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio

TODAY = date(2026, 7, 15)


# ---- capacity: warn before rows are lost (A2) ------------------------------

def test_raised_equity_cap_round_trips(tmp_path):
    # v1.6.2 raised the tight budgets (Equity 250 data rows) — a full sheet
    # must survive the build→read round trip losslessly
    d = sample_portfolio()
    d.equity = [EquityRow(owner="Amit", scrip=f"S{i}", qty=1, avg_cost=10,
                          cost_date=date(2024, 1, 1))
                for i in range(250)]
    p = tmp_path / "big.xlsx"
    build_workbook(d, str(p))
    back = read_workbook(str(p))
    assert len(back.equity) == 250
    assert not any("Equity holds" in w for w in back.warnings)


def test_update_refuses_overfull_sheet_before_touching_anything(tmp_path):
    # openpyxl-saved copies lose charts/comments, which the READER never
    # needs — a legitimate way to fabricate user-grown sheets in tests
    from networth.update import _refuse_overfull
    d = sample_portfolio()
    p = tmp_path / "wb.xlsx"
    build_workbook(d, str(p))
    wb = load_workbook(p)
    ws = wb["Tax_Rules"]                                # smallest cap: 30
    for i in range(35):
        ws.cell(4 + i, 1, "equity")
        ws.cell(4 + i, 2, date(2030 + i, 4, 1))
        ws.cell(4 + i, 3, 365)
    p2 = tmp_path / "grown.xlsx"
    wb.save(p2)
    back = read_workbook(str(p2))
    with pytest.raises(SystemExit):     # regeneration would truncate — the
        _refuse_overfull(back)          # update refuses, nothing is lost
    _refuse_overfull(sample_portfolio())                # within budget: fine


# ---- backups (A3 + A4) ------------------------------------------------------

def test_purge_keeps_exactly_this_runs_backup(tmp_path):
    from networth.update import _purge_unmasked_backups
    path = tmp_path / "wb.xlsx"
    path.write_bytes(b"x")
    bdir = tmp_path / "backups"
    bdir.mkdir()
    old = bdir / "wb.unmasked-backup-20260101-000000.xlsx"
    new = bdir / "wb.unmasked-backup-20260718-120000.xlsx"
    old.write_bytes(b"old")
    new.write_bytes(b"new")
    gone = _purge_unmasked_backups(path, keep=new)
    assert gone == 1 and not old.exists() and new.exists()


def test_relock_backs_up_first(tmp_path):
    from networth.update import relock
    d = sample_portfolio()
    d.privacy_enabled = True
    d.privacy_hash = crypto.hash_password("pw")
    path = tmp_path / "wb.xlsx"
    build_workbook(d, str(path))                        # at rest: readable
    res = relock(path)
    assert res["privacy"] == "masked"
    backups = list((tmp_path / "backups").glob("wb.*backup-*"))
    assert backups, "relock must back up before rewriting"


# ---- distrust empty feeds (A5) ---------------------------------------------

class _Resp:
    def __init__(self, payload, text=""):
        self._p = payload
        self.text = text

    def raise_for_status(self):
        pass

    def json(self):
        return self._p


class _Sess:
    def __init__(self, payload, text=""):
        self._p = payload
        self._t = text

    def get(self, *a, **k):
        return _Resp(self._p, self._t)


def test_ca_structurally_empty_body_is_not_checked():
    from networth.fetch import corporate_actions as ca
    # dict WITHOUT the expected key = changed/maintenance response → counts
    # as a failure; with every query failing, the total-failure escalation
    # fires and the updater keeps ALL existing Auto rows
    with pytest.raises(RuntimeError):
        ca.fetch({"ACME": "INE000A01010"},
                 session=_Sess({"message": "maintenance"}))
    # a VALID empty list still counts as checked (genuinely no actions)
    _actions, checked, _divs, _ = ca.fetch(
        {"ACME": "INE000A01010"}, session=_Sess({"data": []}))
    assert checked == {"INE000A01010"}


def test_amfi_tiny_result_keeps_master(tmp_path, monkeypatch):
    from networth.fetch import amfi
    # the floor lives in fetch() itself: a 200-OK page with 2 schemes raises
    with pytest.raises(ValueError, match="almost nothing"):
        amfi.fetch(session=_Sess(None, text="header\nrow"))
    # and run() turns that raise into keep-old-NAVs, master untouched
    import networth.update as U
    d = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(d, str(path))
    monkeypatch.setattr(U.bhav_mod, "fetch", lambda *a, **k: None)

    def _tiny(*a, **k):
        raise ValueError("AMFI returned almost nothing - keeping the "
                         "existing fund list and NAVs")
    monkeypatch.setattr(U.amfi_mod, "fetch", _tiny)
    monkeypatch.setattr(U.ca_mod, "fetch",
                        lambda *a, **k: ([], set(), [], 0))
    before = len(read_workbook(str(path)).masters.mf_rows)
    s = U.run(path, today=TODAY)
    assert any("almost nothing" in w for w in s["warnings"])
    assert len(read_workbook(str(path)).masters.mf_rows) == before


# ---- person names never crash (A6) -----------------------------------------

def test_person_sheet_name_rules():
    assert person_sheet_name("Jayeshkumar Ramanbhai Patel (HUF)", set()) == \
        "Jayeshkumar Ramanbhai Patel (HU"
    assert person_sheet_name("Jay/HUF", set()) == "Jay-HUF"
    assert person_sheet_name("asha", {"Asha"}) == "asha-2"
    assert person_sheet_name("Equity", set()) == "Equity-2"
    assert person_sheet_name("D'Souza", set()) == "D'Souza"


def test_weird_person_names_build_and_warn(tmp_path):
    d = sample_portfolio()
    d.persons = ["D'Souza", "Jay/HUF",
                 "Jayeshkumar Ramanbhai Patel (HUF)"]
    p = tmp_path / "wb.xlsx"
    build_workbook(d, str(p))                           # must not raise
    wb = load_workbook(p)
    assert "D'Souza" in wb.sheetnames and "Jay-HUF" in wb.sheetnames
    # the teal person strip must follow the ADJUSTED tab name (§3.1)
    tc = wb["Jay-HUF"].sheet_properties.tabColor
    assert tc is not None and "31859C" in str(tc.rgb)
    with zipfile.ZipFile(p) as z:
        charts = [z.read(n).decode() for n in z.namelist()
                  if "charts/chart" in n]
    assert any("D''Souza" in c for c in charts)         # escaped ref
    back = read_workbook(str(p))
    assert any("Jay/HUF" in w and "Jay-HUF" in w for w in back.warnings)


# ---- xirr keeps its never-raises promise (A7) ------------------------------

def test_xirr_extreme_dates_return_none():
    assert xirr([(date(1899, 12, 30), -10000),
                 (date(2026, 7, 15), 250000)]) is None
    assert xirr([(date(2026, 1, 1), -10000),
                 (date(9999, 12, 31), 250000)]) is None


# ---- future-dated rows wait for their date (A8) ----------------------------

def test_future_sip_and_ppf_rows_are_excluded():
    d = PortfolioData(persons=["Amit"])
    d.sip = [SIPRow("Amit", "FUND X", date(2026, 1, 10), 10000, 10.0),
             SIPRow("Amit", "FUND X", date(2027, 1, 10), 10000, 10.0)]
    funds = mf_flows_by_fund(d, TODAY, {("Amit", "FUND X"): 12.0})
    flows = funds[("Amit", "FUND X")]
    assert all(dt <= TODAY for dt, _ in flows)
    assert flows[-1][1] == pytest.approx(1000 * 12.0)   # only real units
    rep = capital_gains_report(d, TODAY, rules=[], fmv=({}, {}))
    assert not rep.realised                              # future lot ignored

    d2 = PortfolioData(persons=["Amit"])
    d2.ppf = [PPFRow(owner="Amit", account_no="A1")]
    d2.ppf_ledger = [
        PPFLedgerRow("Amit", "A1", date(2025, 4, 3), 150000),
        PPFLedgerRow("Amit", "A1", date(2027, 4, 3), 150000)]
    flows = ppf_flows(d2, TODAY)
    assert all(dt <= TODAY for dt, _ in flows)
    assert sum(1 for _, a in flows if a < 0) == 1        # one real deposit


def test_negative_nav_row_is_left_out():
    # a stray minus on the NAV must not flip the units sign silently —
    # the row counts as "no NAV" (same guard as the capital-gains twin)
    d = PortfolioData(persons=["Amit"])
    d.sip = [SIPRow("Amit", "FUND X", date(2026, 1, 10), 10000, 10.0),
             SIPRow("Amit", "FUND X", date(2026, 2, 10), 10000, -152.3)]
    funds = mf_flows_by_fund(d, TODAY, {("Amit", "FUND X"): 12.0})
    flows = funds[("Amit", "FUND X")]
    assert sum(1 for _, a in flows if a < 0) == 1        # one real buy
    assert flows[-1][1] == pytest.approx(1000 * 12.0)    # units unpolluted


def test_yearfrac_clamped_no_discounting():
    assert flat_accrual(100000, 7.1, date(2030, 1, 1),
                        date(2026, 1, 1)) == pytest.approx(100000)


# ---- text in a number cell warns (W1) --------------------------------------

def test_text_in_qty_cell_warns(tmp_path):
    d = sample_portfolio()
    p = tmp_path / "wb.xlsx"
    build_workbook(d, str(p))
    wb = load_workbook(p)
    wb["Equity"].cell(4, 4, "10 shares")
    p2 = tmp_path / "typed.xlsx"
    wb.save(p2)
    back = read_workbook(str(p2))
    assert any("'10 shares'" in w and "Quantity" in w for w in back.warnings)


def test_implausible_date_warns(tmp_path):
    d = sample_portfolio()
    d.equity[0].cost_date = date(1975, 1, 1)
    p = tmp_path / "wb.xlsx"
    build_workbook(d, str(p))
    back = read_workbook(str(p))
    assert any("looks wrong" in w and "1975" in w for w in back.warnings)


# ---- symmetric Yes/No (A11) ------------------------------------------------

def test_yes_no_symmetry_and_garbage_warns(tmp_path):
    d = sample_portfolio()
    p = tmp_path / "wb.xlsx"
    build_workbook(d, str(p))
    wb = load_workbook(p)
    st = wb["Settings"]
    rows = {st.cell(r, 1).value: r for r in range(4, 25)}
    st.cell(rows["Privacy mask"], 2, "Y")               # used to mean OFF
    st.cell(rows["Equity"], 2, "off")                   # used to mean ON
    st.cell(rows["Bonds"], 2, "TRUE")                   # Excel bool survives
    st.cell(rows["PPF"], 2, "maybe")                    # garbage → default
    p2 = tmp_path / "edited.xlsx"
    wb.save(p2)
    back = read_workbook(str(p2))
    assert back.privacy_enabled is True
    assert back.class_settings["equity"].enabled is False   # off means off
    assert back.class_settings["bonds"].enabled is True
    assert any("'maybe'" in w and "PPF" in w for w in back.warnings)
    assert back.class_settings["ppf"].enabled is True       # registry default


# ---- case-insensitive joins, like Excel (A12) ------------------------------

def test_owner_and_metal_case_canonicalised(tmp_path):
    d = sample_portfolio()
    d.equity[0].owner = "AMIT"
    d.bullion[0].metal_type = "gold"    # sample ships bullion rows; if that
    p = tmp_path / "wb.xlsx"            # ever changes this must fail loudly
    build_workbook(d, str(p))
    back = read_workbook(str(p))
    assert back.equity[0].owner == "Amit"
    assert back.bullion[0].metal_type == "Gold"


# ---- friendly late failure (A13) -------------------------------------------

def test_regenerate_permissionerror_is_friendly(tmp_path, monkeypatch):
    import os
    import networth.update as U
    d = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(d, str(path))

    def _locked(*a, **k):
        raise PermissionError(13, "in use")
    monkeypatch.setattr(os, "replace", _locked)
    with pytest.raises(SystemExit):
        U._regenerate_atomic(path, d, masked_build=False, lock_active=False,
                             password=None)
    assert not path.with_name("wb.new.xlsx").exists()   # tmp cleaned up
