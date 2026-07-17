"""Regression tests for the v1.4.2 polish fixes (2026-07-17 review, items
cut by the main report's cap — numbered 1..8 as presented to Jay; item 9 was
docs-only)."""

from datetime import date

import pytest
from openpyxl import load_workbook

import networth.update as U
from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.model import CorporateAction
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import peek_class_details, peek_class_states, run

TODAY = date(2026, 7, 15)
RELIANCE = "INE002A01018"
CHILD = "INE0CHILD012"
SGB_ISIN = "IN0020200104"


def _prices(**extra):
    p = PriceData(trade_date=TODAY, source="T", sources=["BSE", "NSE"])
    p.prices.update(extra)
    return p


# 1 — Avg cost today reflects a demerger's docked basis ---------------------

def test_avg_cost_today_applies_cost_factor(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)
    assert wb["Equity"]["P4"].value == (
        '=IF(OR($D4="",$E4=""),"",$E4*IF($T4="",1,$T4)/IF($S4="",1,$S4))')


# 2 — SGB-only pricing must not refresh the metal rates-as-on stamp ---------

def test_sgb_only_day_keeps_stale_metal_stamp(tmp_path):
    data = sample_portfolio()                    # asof ships as 2026-07-16
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    summary = run(path, price_data=_prices(**{SGB_ISIN: {"close": 14100.0,
                                                         "prev": 14050.0}}),
                  amfi_data=AmfiData(), ca_data=[], div_data=[],
                  bullion_rates={}, today=TODAY)
    assert any("rate unavailable" in w for w in summary["warnings"])
    back = read_workbook(str(path))
    assert back.bullion_rate_asof == date(2026, 7, 16)   # NOT bumped
    sgb = next(b for b in back.bullion if b.metal_type == "SGB")
    assert sgb.rate_auto == 14100.0                      # SGB still priced
    # and the stale-benchmark amber excludes SGB rows (their closes carry
    # their own dates) — assert on the workbook's conditional-format XML
    import zipfile
    with zipfile.ZipFile(path) as z:
        sheets = [z.read(n).decode() for n in z.namelist()
                  if n.startswith("xl/worksheets/sheet")]
    assert any('$B4&lt;&gt;"SGB"' in s and "$I$2" in s for s in sheets)


# 3 — the Dividends Qty comment routes corrections through Manual -----------

def test_dividend_qty_comment_explains_manual(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)
    comment = wb["Dividends"]["H3"].comment
    assert comment is not None and "Manual" in comment.text


# 4 — the console prompt reports EFFECTIVE visibility -----------------------

def test_peek_states_match_settings_on_shipped_template(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    # v1.4.3: the Settings choice IS the state — core five shown, rest hidden
    states = dict(peek_class_states(path))
    assert states["Equity"] is True and states["Bonds"] is True
    assert states["EPF"] is False and states["Property"] is False
    details = {label: (on, rows) for label, on, rows in peek_class_details(path)}
    assert details["EPF"] == (False, True)       # hidden, sample rows inside
    assert details["Equity"] == (True, True)
    assert details["Property"][1] is True        # Manual_Assets class filter
    assert details["Other"][1] is True


def test_peek_states_on_classic_workbook(tmp_path):
    from conftest import classic
    path = tmp_path / "wb.xlsx"
    build_workbook(classic(), str(path))
    states = dict(peek_class_states(path))
    assert states["EPF"] is False                # off AND empty → hidden
    assert states["Equity"] is True


# 5 — --pause survives a closed stdin ---------------------------------------

def test_pause_swallows_eof(monkeypatch):
    def _boom(*_a, **_k):
        raise EOFError
    monkeypatch.setattr("builtins.input", _boom)
    U._pause()                                   # must not raise


# 6 — NPS_Master keeps schemes whose PFM cell is blank ----------------------

def test_nps_master_row_with_blank_pfm_survives(tmp_path):
    data = sample_portfolio()
    data.masters.nps_rows.append(("SM999999", "ZZZ TEST SCHEME", ""))
    data.masters.nps_rows.sort(key=lambda r: r[1].casefold())
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    back = read_workbook(str(path))
    assert ("SM999999", "ZZZ TEST SCHEME", "") in back.masters.nps_rows


# 7 — run() leaves the caller's restructure objects untouched ---------------

def test_run_does_not_mutate_injected_restructures(tmp_path):
    data = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    events = [
        CorporateAction(symbol="RELIANCE", isin=RELIANCE, type="DEMERGER",
                        ex_date=date(2026, 5, 1), ratio_from=1, ratio_to=1,
                        source="Curated", new_isin=RELIANCE, new_name="PARENT",
                        cost_pct=60),
        CorporateAction(symbol="RELIANCE", isin=RELIANCE, type="DEMERGER",
                        ex_date=date(2026, 5, 1), ratio_from=1, ratio_to=1,
                        source="Curated", new_isin=CHILD, new_name="CHILDCO",
                        new_symbol="CHILDCO", cost_pct=40),
    ]
    run(path, price_data=_prices(), amfi_data=AmfiData(), ca_data=[],
        div_data=[], restructures=events, today=TODAY)
    assert all(ev.applied is None for ev in events)      # caller untouched
    back = read_workbook(str(path))
    assert all(a.applied == TODAY for a in back.corporate_actions
               if a.type == "DEMERGER")                  # workbook stamped


# 8 — a nameless restructure row shows the symbol, not the raw ISIN ---------

def test_child_named_from_symbol_when_name_blank(tmp_path):
    data = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    events = [
        CorporateAction(symbol="RELIANCE", isin=RELIANCE, type="DEMERGER",
                        ex_date=date(2026, 5, 1), ratio_from=1, ratio_to=1,
                        source="Curated", new_isin=RELIANCE, new_name="PARENT",
                        cost_pct=60),
        CorporateAction(symbol="RELIANCE", isin=RELIANCE, type="DEMERGER",
                        ex_date=date(2026, 5, 1), ratio_from=1, ratio_to=1,
                        source="Curated", new_isin=CHILD, new_name="",
                        new_symbol="CHILDCO", cost_pct=40),
    ]
    summary = run(path, price_data=_prices(), amfi_data=AmfiData(),
                  ca_data=[], div_data=[], restructures=events, today=TODAY)
    assert summary["restructure_children"] == 1
    back = read_workbook(str(path))
    child = next(r for r in back.equity if r.isin_override == CHILD)
    assert child.scrip == "CHILDCO"              # not the raw ISIN
    assert ("CHILDCO", "CHILDCO", CHILD) in back.masters.stock_rows
