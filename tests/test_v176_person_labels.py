"""v1.7.6 — Dashboard headings must never be read as family members.

The owner reported "Allocation by asset class / Dividends FY 2026-27 / TOTAL —
above sheets does not update data". They were not sheets that failed to
update: the updater had read those three Dashboard HEADINGS as people and
built a tab for each. The person matrix is the one place the reader trusts
a fixed row window (Dashboard A6:A15), so deleting the unused person rows
— which the docs invite ("delete rows freely") — slides TOTAL up from row
16 into that window.

Reading now stops at the first heading, and a file already carrying the
damage repairs itself on the next update.
"""

from datetime import date

import pytest

from networth import model as M
from networth.generate import build_workbook
from networth.model import is_person_label, persons_from_column
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio

TODAY = date(2026, 8, 10)
REAL = ["Amit", "Priya", "Rahul", "Neha", "Vikram"]
STRAYS = ["TOTAL", "Dividends FY 2026-27", "Allocation by asset class"]


def _col(mapping):
    return lambda r: mapping.get(r, "")


def test_headings_are_not_people():
    for bad in STRAYS + ["Person", "Asset class", "As on", "Family net worth",
                         "Portfolio XIRR", "Equity", "Dashboard", ""]:
        assert not is_person_label(bad), bad
    for good in REAL + ["Amit", "Priya", "Totally Fictional Person",
                        "Dividend Kumar"]:
        assert is_person_label(good), good


def test_canonical_layout_reads_only_people():
    rows = {6 + i: n for i, n in enumerate(REAL)}
    rows[M.DASH_TOTAL_ROW] = "TOTAL"          # outside the window anyway
    assert persons_from_column(_col(rows)) == REAL


def test_deleted_empty_rows_do_not_promote_headings():
    # the user tidied up: TOTAL/headings slid up into A11..A13
    rows = {6 + i: n for i, n in enumerate(REAL)}
    for i, s in enumerate(STRAYS):
        rows[11 + i] = s
    assert persons_from_column(_col(rows)) == REAL


def test_a_gap_between_people_is_tolerated():
    # a blank row is skipped (a heading is what stops the scan, not a blank)
    assert persons_from_column(_col({6: "Amit", 8: "Priya"})) == ["Amit",
                                                                 "Priya"]


def test_ten_people_still_all_read():
    ten = [f"P{i}" for i in range(10)]
    rows = {6 + i: n for i, n in enumerate(ten)}
    assert persons_from_column(_col(rows)) == ten


def test_damaged_workbook_repairs_itself(tmp_path):
    """End to end: a file whose headings became people comes back clean,
    the junk tabs go, the real tabs and rows stay."""
    d = sample_portfolio()
    d.persons = REAL + STRAYS                 # the damage, as the reported file had
    d.equity = [M.EquityRow(owner="Amit", scrip="RELIANCE INDUSTRIES LTD.",
                            qty=10, avg_cost=900.0,
                            cost_date=date(2024, 1, 1))]
    broken = tmp_path / "broken.xlsx"
    build_workbook(d, str(broken), today=TODAY)

    back = read_workbook(str(broken))
    assert back.persons == REAL
    assert any("run into its own headings" in w for w in back.warnings)
    # the holding survives untouched
    assert [(r.owner, r.qty) for r in back.equity] == [("Amit", 10)]

    fixed = tmp_path / "fixed.xlsx"
    build_workbook(back, str(fixed), today=TODAY)
    reread = read_workbook(str(fixed))
    assert reread.persons == REAL
    # second pass is clean: nothing left to warn about
    assert not any("run into its own headings" in w for w in reread.warnings)

    from openpyxl import load_workbook
    wb = load_workbook(str(fixed), read_only=True)
    try:
        tabs = set(wb.sheetnames)
        assert {"TOTAL", "Dividends FY 2026-27",
                "Allocation by asset class"} & tabs == set()
        assert set(REAL) <= tabs
        dash = wb["Dashboard"]
        # TOTAL is back where the layout puts it
        assert str(dash.cell(M.DASH_TOTAL_ROW, 1).value) == "TOTAL"
    finally:
        wb.close()


def test_peek_agrees_with_the_reader(tmp_path):
    """The prompt's list and the reader's list are one rule — the console
    said 'People currently tracked: … TOTAL, Dividends FY …' before this."""
    from networth.update import peek_persons
    d = sample_portfolio()
    d.persons = REAL + STRAYS
    p = tmp_path / "wb.xlsx"
    build_workbook(d, str(p), today=TODAY)
    assert peek_persons(str(p)) == REAL == read_workbook(str(p)).persons
