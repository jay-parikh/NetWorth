"""R11: allocation targets — drift, band, rebalance hint (SPEC §3.3/§6.13)."""

import re
import zipfile

from openpyxl import load_workbook

from networth.generate import build_workbook
from networth.model import ClassSetting
from networth.sample_data import sample_portfolio


def _built(tmp_path, data=None):
    # classic layout: these tests assert the five-class column positions
    from conftest import classic
    path = tmp_path / "wb.xlsx"
    build_workbook(data or classic(), str(path))
    return path


def test_allocation_table_columns(tmp_path):
    wb = load_workbook(_built(tmp_path))
    d = wb["Dashboard"]
    assert [d.cell(19, c).value for c in range(1, 8)] == [
        "Asset class", "Value", "XIRR", "Actual %", "Target %", "Drift",
        "Rebalance hint"]
    # live formulas, family-total-relative, Settings-driven
    assert d["D20"].value == '=IF($G$16=0,"",B20/$G$16)'
    assert d["E20"].value == '=IF(Settings!$C$4="","",Settings!$C$4/100)'
    assert d["F20"].value == '=IF($E20="","",$D20-$E20)'
    g = d["G20"].value
    assert '"On target"' in g and "Settings!$B$18" in g
    assert '"Move ₹"&TEXT(ABS($F20)*$G$16,"#,##0")' in g
    assert '" out"' in g and '" in"' in g


def test_target_formula_uses_registry_settings_row(tmp_path):
    # disable Mutual Funds (emptied) — PPF shifts up a Dashboard row, but its
    # Target must still read the PPF row on Settings (registry row 7)
    from conftest import classic
    data = classic()
    data.mutual_funds = []
    data.sip = []
    data.class_settings["mutual_funds"] = ClassSetting(enabled=False)
    wb = load_workbook(_built(tmp_path, data))
    d = wb["Dashboard"]
    labels = [d.cell(r, 1).value for r in range(20, 24)]
    assert labels == ["Equity", "Fixed Deposits", "PPF", "Bonds"]
    assert "Settings!$C$4" in d["E20"].value      # Equity  -> registry row 4
    assert "Settings!$C$6" in d["E21"].value      # FD      -> registry row 6
    assert "Settings!$C$7" in d["E22"].value      # PPF     -> registry row 7
    assert "Settings!$C$9" in d["E23"].value      # Bonds   -> registry row 9 (EPF sits at 8)


def test_drift_band_and_target_chart_present(tmp_path):
    path = _built(tmp_path)
    wb = load_workbook(path)
    dash_xml = f"xl/worksheets/sheet{wb.sheetnames.index('Dashboard') + 1}.xml"
    with zipfile.ZipFile(path) as z:
        dash = z.read(dash_xml).decode()
        charts = "".join(
            z.read(c).decode() for c in z.namelist()
            if re.fullmatch(r"xl/charts/chart\d+\.xml", c))
    # two formula CFs referencing the Settings tolerance band the Drift column
    assert dash.count("Settings!$B$18/100") >= 2
    assert "Actual vs Target %" in charts


def test_targets_written_from_settings_roundtrip(tmp_path):
    from networth.reader import read_workbook
    data = sample_portfolio()
    data.class_settings["equity"] = ClassSetting(enabled=True, target_pct=55)
    data.class_settings["ppf"] = ClassSetting(enabled=True, target_pct=10)
    path = _built(tmp_path, data)
    back = read_workbook(str(path))
    assert back.class_settings["equity"].target_pct == 55
    assert back.class_settings["ppf"].target_pct == 10
    assert back.class_settings["gold_silver"].target_pct is None
