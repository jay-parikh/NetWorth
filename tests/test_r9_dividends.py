"""R9: Dividends sheet — parser, FY lifecycle, qty estimate (SPEC §3.13/§6.12)."""

from datetime import date

from openpyxl import load_workbook

from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.fetch.corporate_actions import parse_dividend
from networth.generate import build_workbook
from networth.model import CorporateAction, DividendRow, EquityRow, fy_label
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import run

TODAY = date(2026, 7, 15)          # FY 2026-27
RELIANCE = "INE002A01018"


def _sources():
    return (PriceData(trade_date=TODAY, source="TEST", sources=["BSE", "NSE"]),
            AmfiData())


# ------------------------------------------------------------------ parser --

def test_parse_dividend_goldens():
    assert parse_dividend("Interim Dividend - Rs. - 5.5000") == ("Interim", 5.5)
    assert parse_dividend("Dividend - Rs 8 Per Share") == ("Final", 8.0)
    assert parse_dividend("Dividend - Re. 1/- Per Share") == ("Final", 1.0)
    assert parse_dividend("Special Dividend Rs. 12.50 per share") == ("Special", 12.5)
    assert parse_dividend("Final Dividend - ₹2.50") == ("Final", 2.5)
    # percent-of-face wording is skipped (Manual row covers it)
    assert parse_dividend("Final Dividend 250%") is None
    # non-dividend subjects never match
    assert parse_dividend("Bonus 1:1") is None
    assert parse_dividend("Annual General Meeting") is None


def test_fy_label():
    assert fy_label(date(2026, 4, 1)) == "2026-27"
    assert fy_label(date(2026, 7, 15)) == "2026-27"
    assert fy_label(date(2026, 3, 31)) == "2025-26"
    assert fy_label(date(2027, 1, 10)) == "2026-27"


# -------------------------------------------------------------- integration --

def _event(ex, rate=5.5, div_type="Final", isin=RELIANCE):
    return DividendRow(scrip="RELIANCE", isin=isin, div_type=div_type,
                       ex_date=ex, rate=rate, source="Auto",
                       details=f"{div_type} Dividend - Rs {rate} Per Share")


def test_rows_created_per_owner_and_idempotent(tmp_path):
    data = sample_portfolio()
    # RELIANCE held by Amit (50, sample) and now Priya too
    data.equity.append(EquityRow(owner="Priya", scrip="RELIANCE INDUSTRIES LTD.",
                                 qty=30, avg_cost=1200, cost_date=date(2024, 1, 5)))
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))

    prices, amfi = _sources()
    ev = _event(date(2026, 6, 10))
    summary = run(path, price_data=prices, amfi_data=amfi, ca_data=[],
                  div_data=[ev], today=TODAY)
    assert summary["dividend_rows"] == 2

    back = read_workbook(str(path))
    rows = [d for d in back.dividends if d.isin == RELIANCE]
    assert {(d.owner, d.qty) for d in rows} == {("Amit", 50.0), ("Priya", 30.0)}
    assert all(d.fy == "2026-27" and d.rate == 5.5 and d.source == "Auto"
               and d.scrip == "RELIANCE INDUSTRIES LTD." for d in rows)

    # second run with the same feed: same rows, no duplicates
    summary2 = run(path, price_data=_sources()[0], amfi_data=AmfiData(),
                   ca_data=[], div_data=[ev], today=TODAY)
    assert summary2["dividend_rows"] == 2
    back2 = read_workbook(str(path))
    assert len([d for d in back2.dividends if d.isin == RELIANCE]) == 2


def test_qty_is_ca_adjusted_as_of_ex_date(tmp_path):
    data = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))

    split = CorporateAction(symbol="RELIANCE", isin=RELIANCE, type="SPLIT",
                            ex_date=date(2026, 5, 1), ratio_from=10,
                            ratio_to=5, source="Auto", details="Split 10->5")
    before = _event(date(2026, 4, 15), div_type="Interim")   # before the split
    after = _event(date(2026, 6, 1), div_type="Final")       # after the split
    run(path, price_data=_sources()[0], amfi_data=AmfiData(),
        ca_data=[split], div_data=[before, after], today=TODAY)

    back = read_workbook(str(path))
    by_type = {d.div_type: d for d in back.dividends if d.isin == RELIANCE}
    assert by_type["Interim"].qty == 50.0          # pre-split count
    assert by_type["Final"].qty == 100.0           # doubled by the 10:5 split


def test_manual_row_overrides_and_prior_fy_freezes(tmp_path):
    data = sample_portfolio()
    manual = DividendRow(fy="2026-27", owner="Amit",
                         scrip="RELIANCE INDUSTRIES LTD.", isin=RELIANCE,
                         div_type="Final", ex_date=date(2026, 6, 10),
                         rate=6.0, qty=48, source="Manual",
                         details="typed from the bank statement")
    data.dividends.append(manual)
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))

    # the feed reports the same (isin, type, ex-date) with a different rate
    run(path, price_data=_sources()[0], amfi_data=AmfiData(), ca_data=[],
        div_data=[_event(date(2026, 6, 10), rate=5.5)], today=TODAY)

    back = read_workbook(str(path))
    rel = [d for d in back.dividends if d.isin == RELIANCE and d.fy == "2026-27"]
    assert len(rel) == 1 and rel[0].source == "Manual" and rel[0].rate == 6.0
    # the sample's prior-FY Auto row (ITC, 2025-26) froze — not rebuilt, not lost
    itc = [d for d in back.dividends if d.fy == "2025-26"]
    assert len(itc) == 1 and itc[0].scrip == "ITC LTD." and itc[0].qty == 200.0
    # current-FY total counts the manual row
    assert any(d.rate == 6.0 for d in rel)


def test_feed_failure_leaves_sheet_untouched(tmp_path):
    data = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    # div_data None (e.g. CA fetch failed) → keep all rows, incl. current FY
    run(path, price_data=_sources()[0], amfi_data=AmfiData(), ca_data=[],
        div_data=None, today=TODAY)
    back = read_workbook(str(path))
    assert {d.fy for d in back.dividends} == {"2025-26", "2026-27"}


# --------------------------------------------------------------- structure --

def test_sheet_structure_and_dashboard_cell(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)
    ws = wb["Dividends"]
    assert [ws.cell(3, c).value for c in range(1, 6)] == [
        "FY", "Owner", "Scrip", "ISIN", "Type"]
    assert ws["I4"].value.startswith("=IF(OR($G4")
    assert "SUMPRODUCT" in ws["N4"].value and "MONTH" in ws["N4"].value
    assert ws["M4"].value == "Apr" and ws["M15"].value == "Mar"
    dash = wb["Dashboard"]
    assert dash["A17"].value.startswith("Dividends FY")
    assert "SUMIFS(Dividends!" in dash["B17"].value
