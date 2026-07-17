"""Structural checks on the generated workbook (SPEC §3)."""

import re
import zipfile

import pytest
from openpyxl import load_workbook

from networth.generate import build_workbook
from networth.sample_data import sample_portfolio

EXPECTED_SHEETS = [
    "Dashboard", "Projection", "Settings", "Amit", "Priya", "Rahul", "Equity",
    "MutualFunds", "MF_SIP", "MF_Master", "Stock_Master", "Bank_Master",
    "FixedDeposits", "PPF", "PPF_Ledger", "EPF", "Bonds", "Gold_Silver",
    "NPS", "NPS_Master", "Manual_Assets", "By Scrip", "Corporate_Actions",
    "Dividends", "History", "Guide",
]


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    path = tmp_path_factory.mktemp("wb") / "template.xlsx"
    build_workbook(sample_portfolio(), str(path))
    return path


@pytest.fixture(scope="module")
def wb(built):
    return load_workbook(built, data_only=False)


def test_sheet_order(wb):
    assert wb.sheetnames == EXPECTED_SHEETS


def test_defined_names(wb):
    names = set(wb.defined_names)
    assert {"MF_SchemeList", "Stock_NameList"} <= names


def test_charts_present(built):
    with zipfile.ZipFile(built) as z:
        charts = [n for n in z.namelist() if re.fullmatch(r"xl/charts/chart\d+\.xml", n)]
        assert len(charts) == 10
        types = "".join(z.read(c).decode() for c in charts)
    assert types.count("<c:pieChart>") == 4      # dashboard + 3 persons
    # net worth by person + dividends by month + actual-vs-target
    assert types.count("<c:barChart>") == 3
    assert types.count("<c:lineChart>") == 2      # projection + net-worth trend
    assert types.count("<c:areaChart>") == 1      # net worth by class (stacked)


def test_dropdown_tip_within_excel_limit():
    # Excel silently drops a validation whose input message exceeds 255 chars
    from networth.generate import _DROPDOWN_TIP
    assert len(_DROPDOWN_TIP) <= 255


def test_typeahead_validations(built):
    with zipfile.ZipFile(built) as z:
        eq = z.read("xl/worksheets/sheet7.xml").decode()   # Equity is 7th sheet
    assert "OFFSET(Stock_Master!$B$3" in eq
    assert 'sqref="C4:C140"' in eq


def test_comments_survive(built):
    with zipfile.ZipFile(built) as z:
        comment_files = [n for n in z.namelist() if "comments" in n and n.endswith(".xml")]
        text = "".join(z.read(c).decode() for c in comment_files)
    assert "Portfolio equity XIRR" in text
    assert "Units = Amount / NAV on date" in text


def test_dashboard_formulas(wb):
    d = wb["Dashboard"]
    assert d["B3"].value == "=G16"      # 5 shown classes → Total in column G
    assert "SUMIFS(Equity!$I:$I,Equity!$A:$A,$A6)" in d["B6"].value
    assert d["G16"].value == "=SUM(G6:G15)"
    assert d["E4"].value.startswith('=IF(B4="","",(1+B4)/(1+E3/100)-1)')
    assert d["E3"].value == 7          # inflation input
    assert d["B4"].value == pytest.approx(0.0676209694)


def test_equity_sheet(wb):
    e = wb["Equity"]
    assert e["I4"].value == '=IF($D4="","",$D4*IF($S4="",1,$S4)*$F4)'
    assert e["O4"].value == '=IF($D4="","",$D4*IF($S4="",1,$S4))'
    assert e["P4"].value == '=IF(OR($D4="",$E4=""),"",$E4*IF($T4="",1,$T4)/IF($S4="",1,$S4))'
    assert e["O3"].value == "Qty today" and e["P3"].value == "Avg cost today"
    assert "MATCH($C4,Stock_Master!$B:$B,0)" in e["B4"].value
    assert e["N142"].value == pytest.approx(0.0664365522)
    assert e["C142"].value == "TOTAL"
    assert e["I142"].value == "=SUM(I4:I140)"
    assert e["M4"].value.date().isoformat() == "2018-01-31"
    assert e.freeze_panes == "A4"


def test_mf_linkage(wb):
    m = wb["MutualFunds"]
    assert "SUMIFS(MF_SIP!$H:$H,MF_SIP!$A:$A,$A4,MF_SIP!$D:$D,$D4)" in m["E4"].value
    s = wb["MF_SIP"]
    assert s["H4"].value == '=IF(OR($F4="",$G4=""),"",$F4/$G4)'
    assert s["J2"].value == pytest.approx(0.0876706058)


def test_fd_formula(wb):
    f = wb["FixedDeposits"]
    assert "YEARFRAC($F4,MIN(TODAY(),$G4))" in f["I4"].value
    assert f["J55"].value == "=SUM(J4:J53)"


def test_projection_span(wb):
    p = wb["Projection"]
    assert p["A4"].value == "=YEAR(TODAY())+0"
    assert p["B24"].value == "=Dashboard!$B$3*(1+Dashboard!$B$4)^20"


def test_person_sheet_blocks(wb):
    a = wb["Amit"]
    assert a["B2"].value == "Amit"
    # 5 shown classes: summary rows 6-10, total row 11, blocks from 14
    assert a["A11"].value == "Total"
    assert a["A14"].value == "EQUITY"
    assert 'MATCH($B$2&"#"&1,Equity!$Q:$Q,0)' in a["A16"].value
    # person Qty/Avg-cost pull the post-action (demat) view
    assert "INDEX(Equity!$O:$O" in a["C16"].value
    assert "INDEX(Equity!$P:$P" in a["D16"].value
    # stacked blocks: EQ(14) MF(57) FD(80) PPF(98) BONDS(111); hidden
    # classes (EPF etc.) contribute no block at all
    assert a["A111"].value == "BONDS"
    col_a = [a.cell(r, 1).value for r in range(1, 140)]
    assert "EPF" not in col_a and "GOLD & SILVER" not in col_a


def test_masters_loaded(wb):
    assert wb["MF_Master"].max_row > 10000
    assert wb["Stock_Master"].max_row > 4000
    assert wb["MF_Master"]["B3"].value == "Scheme Name"
    # sorted by name — dropdown filter requirement
    names = [wb["Stock_Master"].cell(r, 2).value for r in range(4, 104)]
    assert names == sorted(names, key=lambda s: s.casefold())
