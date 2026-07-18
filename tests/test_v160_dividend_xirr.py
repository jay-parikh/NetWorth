"""v1.6: dividends + recorded sales enter the equity XIRR (SPEC §6.2).

The equity return is now money-weighted for real: the strict two-flow model
per holding gains (a) one inflow per dividend row at its ex-date and (b) a
buy→sell round trip per complete Equity_Sells row. Flow ORDER is part of the
contract — older tests pin flows[0]/[1] as the first row's outflow/inflow, so
everything new is appended after the per-row pairs.
"""

from datetime import date

from networth.compute.cashflows import equity_flows
from networth.compute.xirr import xirr
from networth.model import (DividendRow, EquityRow, EquitySellRow,
                            PortfolioData)

TODAY = date(2026, 7, 15)


def _base() -> PortfolioData:
    d = PortfolioData(persons=["Amit"])
    d.equity = [EquityRow(owner="Amit", scrip="X", qty=100, avg_cost=100,
                          close=120, cost_date=date(2025, 7, 15))]
    return d


def test_dividend_flow_appended_and_golden():
    d = _base()
    d.dividends = [DividendRow(fy="2026-27", owner="Amit", scrip="X",
                               isin="INE000X", ex_date=date(2026, 1, 15),
                               rate=5, qty=100, source="Auto")]
    flows = equity_flows(d, TODAY)
    # per-row pair first (order contract), dividend appended after
    assert flows[0] == (date(2025, 7, 15), -10000.0)
    assert flows[1] == (TODAY, 12000.0)
    assert flows[2] == (date(2026, 1, 15), 500.0)
    # dividends must RAISE the return vs the bare two-flow
    assert xirr(flows) > xirr(flows[:2])


def test_future_and_incomplete_dividends_are_excluded():
    d = _base()
    d.dividends = [
        DividendRow(owner="Amit", ex_date=date(2026, 8, 1), rate=5, qty=100),
        DividendRow(owner="Amit", ex_date=date(2026, 1, 1), rate=None, qty=9),
        DividendRow(owner="Amit", ex_date=date(2026, 1, 1), rate=5, qty=None),
        DividendRow(owner="Amit", ex_date=None, rate=5, qty=100),
    ]
    assert len(equity_flows(d, TODAY)) == 2       # only the holding's pair


def test_sell_round_trip_enters_xirr():
    d = _base()
    d.equity_sells = [EquitySellRow(owner="Amit", scrip="Y", qty=10,
                                    buy_date=date(2024, 1, 10),
                                    buy_price=1400,
                                    sell_date=date(2026, 5, 20),
                                    sell_price=1650)]
    flows = equity_flows(d, TODAY)
    assert (date(2024, 1, 10), -14000.0) in flows
    assert (date(2026, 5, 20), 16500.0) in flows


def test_incomplete_or_impossible_sells_stay_out():
    d = _base()
    d.equity_sells = [
        # blank buy price = grandfathering path → not a complete round trip
        EquitySellRow(owner="A", scrip="Y", qty=10,
                      buy_date=date(2016, 1, 1), buy_price=None,
                      sell_date=date(2026, 1, 1), sell_price=800),
        # future sale
        EquitySellRow(owner="A", scrip="Y", qty=10,
                      buy_date=date(2024, 1, 1), buy_price=100,
                      sell_date=date(2027, 1, 1), sell_price=200),
        # sold before bought
        EquitySellRow(owner="A", scrip="Y", qty=10,
                      buy_date=date(2026, 1, 1), buy_price=100,
                      sell_date=date(2025, 1, 1), sell_price=200),
    ]
    assert len(equity_flows(d, TODAY)) == 2


def test_all_sold_dividends_only_returns_none_not_crash():
    d = PortfolioData(persons=["Amit"])
    d.dividends = [DividendRow(owner="Amit", ex_date=date(2026, 1, 15),
                               rate=5, qty=100)]
    flows = equity_flows(d, TODAY)
    assert len(flows) == 1 and xirr(flows) is None   # same-sign guard


def test_person_sheet_dividend_split(tmp_path):
    from openpyxl import load_workbook

    from networth.generate import build_workbook
    from networth.sample_data import sample_portfolio

    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path), today=TODAY)
    wb = load_workbook(path)
    amit = wb["Amit"]
    assert str(amit["A4"].value).startswith("Dividends FY ")
    f = amit["B4"].value
    assert "SUMIFS(Dividends!$I:$I" in f and "Dividends!$B:$B,$B$2" in f
    # family total on the Dashboard is untouched
    assert "SUMIFS(Dividends!" in wb["Dashboard"]["B17"].value


def test_person_sheet_split_absent_when_equity_off(tmp_path):
    from openpyxl import load_workbook

    from networth.generate import build_workbook
    from networth.model import ClassSetting
    from networth.sample_data import sample_portfolio

    d = sample_portfolio()
    d.class_settings["equity"] = ClassSetting(enabled=False)
    path = tmp_path / "wb.xlsx"
    build_workbook(d, str(path), today=TODAY)
    wb = load_workbook(path)
    assert wb["Amit"]["A4"].value is None


def test_zero_buy_price_sale_counts_in_xirr():
    # bonus/ESOP shares: typed buy price 0 is a real (zero) cost — the sale
    # must enter the return, matching the Capital Gains view (review fix)
    d = _base()
    d.equity_sells = [EquitySellRow(owner="Amit", scrip="X", qty=10,
                                    buy_date=date(2024, 1, 1), buy_price=0.0,
                                    sell_date=date(2025, 1, 1),
                                    sell_price=50.0)]
    flows = equity_flows(d, TODAY)
    assert (date(2024, 1, 1), 0.0) in flows
    assert (date(2025, 1, 1), 500.0) in flows
