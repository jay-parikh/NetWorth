"""v1.4 onboarding polish: console class toggle, version line, live samples."""

from datetime import date

from openpyxl import load_workbook

from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.sample_data import sample_portfolio
from networth.update import peek_class_states, run, version_line

TODAY = date(2026, 7, 16)


def test_every_class_sheet_ships_with_sample_rows():
    data = sample_portfolio()
    assert data.epf and data.bullion and data.nps
    classes = {r.asset_class for r in data.manual_assets}
    assert {"Real Estate", "Cash", "Insurance"} <= classes
    # SGB + physical gold + silver all demonstrated
    assert {b.metal_type for b in data.bullion} == {"SGB", "Gold", "Silver"}
    # a few targets light up the drift view, and they sum to 100
    targets = [s.target_pct for s in data.class_settings.values()
               if s.target_pct is not None]
    assert sum(targets) == 100


def test_console_toggle_round_trips_through_run(tmp_path):
    from conftest import classic
    path = tmp_path / "wb.xlsx"
    build_workbook(classic(), str(path))
    assert dict(peek_class_states(path))["EPF"] is False

    summary = run(path, price_data=PriceData(trade_date=TODAY, source="T"),
                  amfi_data=AmfiData(), ca_data=[], div_data=[],
                  toggle_classes=["EPF", "gold & silver"], today=TODAY)
    assert set(summary["classes_toggled"]) == {"EPF → shown",
                                               "Gold & Silver → shown"}
    wb = load_workbook(path)
    assert wb["EPF"].sheet_state == "visible"
    assert wb["Gold_Silver"].sheet_state == "visible"
    states = dict(peek_class_states(path))
    assert states["EPF"] is True and states["Gold & Silver"] is True

    # and back off again (they hold no rows, so they really hide)
    run(path, price_data=PriceData(trade_date=TODAY, source="T"),
        amfi_data=AmfiData(), ca_data=[], div_data=[],
        toggle_classes=["EPF"], today=TODAY)
    wb = load_workbook(path)
    assert wb["EPF"].sheet_state == "hidden"


def test_version_line_always_says_something():
    class _Resp:
        def __init__(self, status, payload):
            self.status_code = status
            self._p = payload
        def json(self):
            return self._p

    class _Sess:
        def __init__(self, status=200, payload=None):
            self._r = _Resp(status, payload or {})
        def get(self, *a, **k):
            return self._r

    newer = _Sess(200, {"tag_name": "v9.9.9", "html_url": "https://x"})
    assert "Update available" in version_line("1.4.0", session=newer)
    current = _Sess(200, {"tag_name": "v1.4.0"})
    assert "latest release" in version_line("1.4.0", session=current)
    offline = _Sess(404, {})
    assert "couldn't reach GitHub" in version_line("1.4.0", session=offline)
