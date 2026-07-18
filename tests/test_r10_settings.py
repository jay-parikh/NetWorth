"""R10: Settings sheet, class registry, selectable classes (SPEC §2/§3.14)."""

from dataclasses import asdict
from datetime import date

from openpyxl import load_workbook

from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.model import ASSET_CLASSES, ClassSetting, HistorySnapshot
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import run

TODAY = date(2026, 7, 15)


REFERENCE = {"MF_Master", "Stock_Master", "Bank_Master", "NPS_Master",
             "Corporate_Actions"}


def test_default_template_shows_core_only(tmp_path):
    # v1.4.3 calm first open: the classic five show; the other classes ship
    # hidden WITH their sample rows waiting inside; the reference sheets
    # tuck away behind the Settings "Reference lists" switch
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)
    assert "Settings" in wb.sheetnames
    hidden = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}
    assert hidden == ({"EPF", "Gold_Silver", "NPS", "Manual_Assets"}
                      | {"Equity_Sells", "Capital Gains", "Tax_Rules"}   # v1.6
                      | REFERENCE)
    st = wb["Settings"]
    n = len(ASSET_CLASSES)
    labels = [st.cell(r, 1).value for r in range(4, 4 + n)]
    assert labels == [c.label for c in ASSET_CLASSES]
    for i, cls in enumerate(ASSET_CLASSES):
        if cls.default_enabled:
            assert st.cell(4 + i, 2).value == "Yes"
            assert st.cell(4 + i, 4).value == "Shown"
        else:
            assert st.cell(4 + i, 2).value == "No"
            assert st.cell(4 + i, 4).value == "Hidden - has data (not counted)"
    assert st["A16"].value == "Reference lists" and st["B16"].value == "No"
    assert st["A17"].value == "Capital gains report" and st["B17"].value == "No"
    assert st["B19"].value == 5                     # drift tolerance default
    # the Dashboard carries the one-line awareness note for the hidden money
    notice = wb["Dashboard"]["I1"].value or ""
    assert "Hidden, not counted" in notice and "EPF" in notice


def test_reference_lists_switch_shows_masters(tmp_path):
    data = sample_portfolio()
    data.show_references = True
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    for name in REFERENCE:
        assert wb[name].sheet_state == "visible", name
    back = read_workbook(str(path))
    assert back.show_references is True             # round-trips
    data.show_references = False
    build_workbook(data, str(path))
    assert read_workbook(str(path)).show_references is False


def test_registry_defaults_hide_new_classes(tmp_path):
    # without the sample rows/settings, the registry defaults still ship the
    # classic five visible and everything newer hidden
    from conftest import classic
    path = tmp_path / "wb.xlsx"
    build_workbook(classic(), str(path))
    wb = load_workbook(path)
    hidden = {ws.title for ws in wb.worksheets if ws.sheet_state != "visible"}
    assert hidden == ({"EPF", "Manual_Assets", "Gold_Silver", "NPS"}
                      | {"Equity_Sells", "Capital Gains", "Tax_Rules"}   # v1.6
                      | REFERENCE)


def test_disabled_empty_class_hides_sheets_and_columns(tmp_path):
    from conftest import classic
    data = classic()
    data.bonds = []                                  # no data → may hide
    data.class_settings["bonds"] = ClassSetting(enabled=False)
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    assert wb["Bonds"].sheet_state == "hidden"       # hidden, never omitted
    assert wb["Equity"].sheet_state == "visible"
    # Dashboard matrix loses the Bonds column; Total/FY shift left
    d = wb["Dashboard"]
    assert [d.cell(5, c).value for c in range(1, 8)] == [
        "Person", "Equity", "Mutual Funds", "Fixed Deposits", "PPF",
        "Total", d.cell(5, 7).value]
    assert str(d.cell(5, 7).value).startswith("Expected @")
    assert d["B3"].value == "=F16"                   # total col moved G→F
    # allocation table has 4 rows, no Bonds
    alloc = [d.cell(r, 1).value for r in range(20, 25)]
    assert "Bonds" not in alloc and alloc[3] == "PPF" and alloc[4] is None
    # person sheet: 4 summary rows, total at row 10, blocks re-stacked
    a = wb["Amit"]
    assert a["A9"].value == "PPF" and a["A10"].value == "Total"
    assert a["B3"].value == "=B10"
    assert a["A14"].value == "EQUITY"
    st = wb["Settings"]
    assert st.cell(9, 4).value == "Hidden"           # Bonds registry row status


def test_disabled_class_with_data_hides_and_warns(tmp_path):
    data = sample_portfolio()                        # bonds HAS a row
    data.class_settings["bonds"] = ClassSetting(enabled=False)
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    # v1.4.3: the Settings choice wins — hidden even though it holds a row
    assert wb["Bonds"].sheet_state == "hidden"
    assert (wb["Settings"].cell(9, 4).value
            == "Hidden - has data (not counted)")
    # every run carries exactly ONE awareness line, naming the hidden money
    summary = run(path, price_data=PriceData(trade_date=TODAY, source="T"),
                  amfi_data=AmfiData(), ca_data=[], div_data=[], today=TODAY)
    hidden_lines = [w for w in summary["warnings"]
                    if w.startswith("hidden and not counted")]
    assert len(hidden_lines) == 1 and "Bonds" in hidden_lines[0]
    # ...mirrored by the Dashboard notice in the regenerated workbook
    wb = load_workbook(path)
    assert "Bonds" in (wb["Dashboard"]["I1"].value or "")
    # the user's No round-trips unchanged (it is their setting, not ours)
    back = read_workbook(str(path))
    assert back.class_settings["bonds"].enabled is False
    # toggling OFF a class that holds rows says what happens, right then
    summary = run(path, price_data=PriceData(trade_date=TODAY, source="T"),
                  amfi_data=AmfiData(), ca_data=[], div_data=[],
                  toggle_classes=["Equity"], today=TODAY)
    assert any("Equity is now hidden" in w for w in summary["warnings"])


def test_settings_round_trip_any_combination(tmp_path):
    data = sample_portfolio()
    data.bonds = []
    data.fixed_deposits = []
    data.class_settings["bonds"] = ClassSetting(enabled=False, target_pct=10)
    data.class_settings["fixed_deposits"] = ClassSetting(enabled=False)
    data.class_settings["equity"] = ClassSetting(enabled=True, target_pct=55)
    data.drift_tolerance_pct = 7.5
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))

    back = read_workbook(str(path))
    assert back.class_settings == data.class_settings
    assert back.drift_tolerance_pct == 7.5

    # full round-trip identity: rebuild from the read and compare
    path2 = tmp_path / "wb2.xlsx"
    build_workbook(back, str(path2))
    again = read_workbook(str(path2))
    assert asdict(again) == asdict(back)


def test_history_columns_are_label_keyed(tmp_path):
    from conftest import classic
    data = classic()
    data.bonds = []
    data.class_settings["bonds"] = ClassSetting(enabled=False)
    # bonds disabled and empty NOW, but it has past history → column stays
    data.history = [HistorySnapshot(snap_date=date(2026, 7, 1), equity=100.0,
                                    mutual_funds=50.0, fixed_deposits=25.0,
                                    ppf=10.0, bonds=5.0)]
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    h = wb["History"]
    headers = [h.cell(3, c).value for c in range(1, 8)]
    assert headers == ["Date", "Equity", "Mutual Funds", "Fixed Deposits",
                       "PPF", "Bonds", "Total"]
    back = read_workbook(str(path))
    assert back.history[0].bonds == 5.0

    # with no bonds history either, the column disappears and reads as 0
    data.history = [HistorySnapshot(snap_date=date(2026, 7, 1), equity=100.0,
                                    mutual_funds=50.0, fixed_deposits=25.0,
                                    ppf=10.0)]
    build_workbook(data, str(path))
    wb = load_workbook(path)
    h = wb["History"]
    headers = [h.cell(3, c).value for c in range(1, 7)]
    assert headers == ["Date", "Equity", "Mutual Funds", "Fixed Deposits",
                       "PPF", "Total"]
    back = read_workbook(str(path))
    assert back.history[0].bonds == 0.0
    assert back.history[0].equity == 100.0


def test_dashboard_visual_upgrades_present(tmp_path):
    import re
    import zipfile
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)
    dash_xml = f"xl/worksheets/sheet{wb.sheetnames.index('Dashboard') + 1}.xml"
    eq_xml = f"xl/worksheets/sheet{wb.sheetnames.index('Equity') + 1}.xml"
    with zipfile.ZipFile(path) as z:
        dash = z.read(dash_xml).decode()
        eq = z.read(eq_xml).decode()
        charts = "".join(
            z.read(c).decode() for c in z.namelist()
            if re.fullmatch(r"xl/charts/chart\d+\.xml", c))
    assert "dataBar" in dash                       # allocation value bars
    assert "iconSet" in eq                         # day-change ▲/▼ arrows
    assert "<c:areaChart>" in charts               # net worth by class, stacked
    assert "Net worth by class over time" in charts
