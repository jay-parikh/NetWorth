"""R6: FMV 31-01-2018 fallback, delisted/stale detection, amber flags."""

import zipfile
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from networth.fetch.amfi import AmfiData
from networth.fetch.bhavcopy import PriceData
from networth.generate import build_workbook
from networth.model import EquityRow, load_fmv
from networth.reader import read_workbook
from networth.sample_data import sample_portfolio
from networth.update import run

TODAY = date(2026, 7, 15)


def _empty_sources():
    # dual-source: R8's status escalation only trusts absence from BOTH exchanges
    return (PriceData(trade_date=TODAY, source="TEST", sources=["BSE", "NSE"]),
            AmfiData())


def test_fmv_dataset_loads():
    by_isin, by_symbol = load_fmv()
    assert len(by_isin) > 1500
    assert by_isin["INE002A01018"] == 964.5          # RELIANCE high on 31-01-2018
    assert by_symbol["HDFCBANK"] == 2013.5           # ISIN changed post-split


def test_fmv_fallback_by_isin_and_symbol(tmp_path):
    data = sample_portfolio()
    # blank cost, pre-2018 purchase → by-ISIN hit (RELIANCE, same ISIN today)
    data.equity.append(EquityRow(owner="Amit", scrip="RELIANCE INDUSTRIES LTD.",
                                 qty=10, cost_date=date(2016, 5, 1)))
    # HDFC BANK: today's ISIN differs from the 2018 one → symbol fallback
    data.equity.append(EquityRow(owner="Priya", scrip="HDFC BANK LTD.",
                                 qty=5, cost_date=date(2017, 1, 1)))
    # post-cutoff purchase must NOT be FMV-filled
    data.equity.append(EquityRow(owner="Rahul", scrip="ITC LTD.",
                                 qty=5, cost_date=date(2020, 1, 1)))
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))

    prices, amfi = _empty_sources()
    summary = run(path, price_data=prices, amfi_data=amfi, ca_data=[], today=TODAY)
    assert summary["fmv_filled"] == 2

    back = read_workbook(str(path))
    rel, hdfc, itc = back.equity[-3], back.equity[-2], back.equity[-1]
    assert rel.avg_cost == 964.5 and rel.fmv_used
    assert hdfc.avg_cost == 2013.5 and hdfc.fmv_used
    assert itc.avg_cost is None and not itc.fmv_used
    # flag survives a second regenerate untouched
    build_workbook(back, str(path))
    again = read_workbook(str(path))
    assert again.equity[-3].fmv_used and again.equity[-3].avg_cost == 964.5


def test_status_active_then_suspended_then_delisted(tmp_path):
    data = sample_portfolio()
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    isin_by_name = {n: i for _s, n, i in data.masters.stock_rows}
    wipro = isin_by_name["WIPRO LTD."]

    # round 1: everyone quoted → Active
    prices, amfi = _empty_sources()
    for row in data.equity:
        prices.prices[isin_by_name[row.scrip]] = {"close": 100.0, "prev": 99.0}
    run(path, price_data=prices, amfi_data=amfi, ca_data=[], today=TODAY)
    back = read_workbook(str(path))
    assert back.masters.stock_status[wipro][0] == "Active"

    # round 2, 30 days on: WIPRO missing from the feed → Suspended
    later = TODAY + timedelta(days=30)
    prices2, _ = _empty_sources()
    prices2.trade_date = later
    for row in data.equity:
        if row.scrip != "WIPRO LTD.":
            prices2.prices[isin_by_name[row.scrip]] = {"close": 101.0, "prev": 100.0}
    summary = run(path, price_data=prices2, amfi_data=AmfiData(), ca_data=[], today=later)
    assert summary["suspended"] == 1
    back = read_workbook(str(path))
    st, last = back.masters.stock_status[wipro]
    assert st == "Suspended" and last == TODAY   # last traded stays at round 1
    # price kept (manual-override semantics), only flagged
    wrow = next(r for r in back.equity if r.scrip == "WIPRO LTD.")
    assert wrow.close == 100.0 and wrow.close_date == TODAY

    # round 3, 200+ days on: still missing → Delisted
    much_later = TODAY + timedelta(days=200)
    prices3, _ = _empty_sources()
    prices3.trade_date = much_later
    for row in data.equity:
        if row.scrip != "WIPRO LTD.":
            prices3.prices[isin_by_name[row.scrip]] = {"close": 102.0, "prev": 101.0}
    run(path, price_data=prices3, amfi_data=AmfiData(), ca_data=[], today=much_later)
    back = read_workbook(str(path))
    assert back.masters.stock_status[wipro][0] == "Delisted"


def test_amber_conditional_formats_present(tmp_path):
    path = tmp_path / "wb.xlsx"
    build_workbook(sample_portfolio(), str(path))
    wb = load_workbook(path)
    with zipfile.ZipFile(path) as z:
        eq = z.read(f"xl/worksheets/sheet{wb.sheetnames.index('Equity') + 1}.xml").decode()
    assert "TODAY()-$H4&gt;7" in eq or "TODAY()-$H4>7" in eq
    assert "Stock_Master!$D:$D" in eq


def test_stock_master_status_columns(tmp_path):
    data = sample_portfolio()
    data.masters.stock_status["INE075A01022"] = ("Suspended", date(2026, 6, 1))
    path = tmp_path / "wb.xlsx"
    build_workbook(data, str(path))
    wb = load_workbook(path)
    sm = wb["Stock_Master"]
    assert sm["D3"].value == "Status"
    hit = [r for r in range(4, sm.max_row + 1)
           if sm.cell(r, 3).value == "INE075A01022"]
    assert sm.cell(hit[0], 4).value == "Suspended"
